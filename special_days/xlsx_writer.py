"""Excel (``.xlsx``) writer, built on openpyxl.

Produces a single "Special Dates" sheet: the ten headline columns (Event,
Start date, End date, City, Source, Nearest airport, Impact, Predicted
attendance, Bridge start, Bridge end) plus two per-day weight columns (Impact
by day, Impact by day (bridge)) as JSON strings. Bold + frozen header,
auto-filter, real date cells.
"""

from __future__ import annotations

import json

from .models import SpecialDate
from .output import HEADERS

_JSON_HEADERS = ["Impact by day", "Impact by day (bridge)"]
_DATE_COLUMNS = (2, 3, 9, 10)  # Start, End, Bridge start, Bridge end
_COLUMN_WIDTHS = {
    "A": 44, "B": 12, "C": 12, "D": 20, "E": 14, "F": 15, "G": 8,
    "H": 12, "I": 12, "J": 12, "K": 58, "L": 58,
}


def write_xlsx(rows: list[SpecialDate], path: str) -> None:
    """Write ``rows`` to an ``.xlsx`` file at ``path``."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Special Dates"

    header = list(HEADERS) + _JSON_HEADERS
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for sd in rows:
        sheet.append(
            [
                sd.event,
                sd.start_date,  # real date cell
                sd.end_date,  # real date cell
                sd.city,
                sd.source,
                sd.nearest_airport or "",
                sd.impact_score,  # numeric cell
                sd.predicted_attendance,  # numeric cell (blank without LLM)
                sd.bridge_start,  # real date cell
                sd.bridge_end,  # real date cell
                json.dumps(dict(sd.impact_by_day or ()), ensure_ascii=False),
                json.dumps(dict(sd.impact_by_day_bridge or ()), ensure_ascii=False),
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        for column in _DATE_COLUMNS:
            row[column - 1].number_format = "yyyy-mm-dd"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(header))}{sheet.max_row}"

    for column, width in _COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width

    workbook.save(path)
