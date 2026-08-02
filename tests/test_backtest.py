import os
import tempfile
import unittest
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from special_days import backtest


EVENT_HEADERS = [
    "Event",
    "Start date",
    "End date",
    "City",
    "Source",
    "Nearest airport",
    "Impact",
    "Predicted attendance",
    "Bridge start",
    "Bridge end",
    "Impact by day",
    "Impact by day (bridge)",
]

LF_HEADERS = [
    "YMD",
    "AIRPORT",
    "BOARDED_PAX",
    "CAPACITY",
    "LF",
    "flight_date",
    "LF_PCT",
]


def _write_workbook(path, headers, rows, title="Sheet1"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value


class BacktestWorkflowTest(unittest.TestCase):
    def test_workflow_enriches_and_expands_airports(self):
        with tempfile.TemporaryDirectory() as tmp:
            lf_path = os.path.join(tmp, "lf.xlsx")
            events_path = os.path.join(tmp, "events.xlsx")

            _write_workbook(
                lf_path,
                LF_HEADERS,
                [
                    [20250101, "IST", 100, 200, 0.50, datetime(2025, 1, 1), 50.0],
                    [20250102, "SAW", 120, 200, 0.60, datetime(2025, 1, 2), 60.0],
                    [20250102, "JFK", 130, 200, 0.65, datetime(2025, 1, 2), 65.0],
                    [20250102, "EWR", 140, 200, 0.70, datetime(2025, 1, 2), 70.0],
                ],
            )
            _write_workbook(
                events_path,
                EVENT_HEADERS,
                [
                    [
                        "Ramazan Bayramı",
                        date(2025, 1, 1),
                        date(2025, 1, 1),
                        "Nationwide (TR)",
                        "nager",
                        "",
                        90,
                        "",
                        date(2025, 1, 1),
                        date(2025, 1, 2),
                        "",
                        '{"2025-01-01": 90, "2025-01-02": 75}',
                    ],
                    [
                        "New York Expo",
                        date(2025, 1, 2),
                        date(2025, 1, 2),
                        "New York",
                        "eventseye",
                        "",
                        55,
                        1000,
                        "",
                        "",
                        '{"2025-01-02": 55}',
                        "",
                    ],
                ],
            )

            outputs = backtest.run_workflow(lf_path, events_path, tmp)

            enriched = load_workbook(outputs["enriched_workbook"])
            correlation = load_workbook(outputs["correlation_workbook"])
            try:
                event_rows = list(enriched["event_day_airport"].iter_rows(values_only=True))
                header = event_rows[0]
                airport_index = header.index("Airport")
                event_date_index = header.index("Event date")
                event_index = header.index("Event")
                values = event_rows[1:]
                event_tuples = {
                    (row[event_index], _as_date(row[event_date_index]), row[airport_index]) for row in values
                }

                self.assertIn(("Ramazan Bayramı", date(2025, 1, 1), "IST"), event_tuples)
                self.assertIn(("Ramazan Bayramı", date(2025, 1, 2), "SAW"), event_tuples)
                self.assertIn(("New York Expo", date(2025, 1, 2), "JFK"), event_tuples)
                self.assertIn(("New York Expo", date(2025, 1, 2), "EWR"), event_tuples)

                timeline_rows = list(correlation["daily_timeline"].iter_rows(values_only=True))
                timeline_header = timeline_rows[0]
                date_index = timeline_header.index("date")
                peak_index = timeline_header.index("event_peak_impact")
                high_names_index = timeline_header.index("high_impact_event_names")
                high_group_names_index = timeline_header.index("high_impact_event_group_names")
                by_date = {_as_date(row[date_index]): row for row in timeline_rows[1:]}

                self.assertEqual(by_date[date(2025, 1, 1)][peak_index], 90)
                self.assertIn("Ramazan Bayramı", by_date[date(2025, 1, 1)][high_names_index])
                self.assertEqual(by_date[date(2025, 1, 1)][high_group_names_index], "Ramazan Bayramı")
                self.assertGreater(len(correlation["daily_timeline"]._charts), 0)
                self.assertGreater(len(correlation["high_impact_timeline"]._charts), 0)
            finally:
                enriched.close()
                correlation.close()

    def test_workflow_routes_unknown_city_to_unresolved(self):
        with tempfile.TemporaryDirectory() as tmp:
            lf_path = os.path.join(tmp, "lf.xlsx")
            events_path = os.path.join(tmp, "events.xlsx")

            _write_workbook(
                lf_path,
                LF_HEADERS,
                [[20250101, "IST", 100, 200, 0.50, datetime(2025, 1, 1), 50.0]],
            )
            _write_workbook(
                events_path,
                EVENT_HEADERS,
                [
                    [
                        "Unknown Event",
                        date(2025, 1, 1),
                        date(2025, 1, 1),
                        "Atlantis",
                        "eventseye",
                        "",
                        40,
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                ],
            )

            outputs = backtest.run_workflow(lf_path, events_path, tmp)

            enriched = load_workbook(outputs["enriched_workbook"])
            correlation = load_workbook(outputs["correlation_workbook"])
            try:
                unresolved_rows = list(enriched["unresolved_events"].iter_rows(values_only=True))
                header = unresolved_rows[0]
                reason_index = header.index("unresolved_reason")
                self.assertEqual(unresolved_rows[1][reason_index], "city_not_found_or_not_in_lf")

                high_timeline_rows = list(correlation["high_impact_timeline"].iter_rows(values_only=True))
                if len(high_timeline_rows) > 1:
                    text_blob = " ".join(str(value) for row in high_timeline_rows[1:] for value in row if value is not None)
                    self.assertNotIn("Unknown Event", text_blob)
            finally:
                enriched.close()
                correlation.close()

    def test_spearman_correlation_handles_monotonic_series_and_constants(self):
        self.assertAlmostEqual(backtest.spearman_correlation([1, 2, 3], [10, 20, 30]), 1.0)
        self.assertIsNone(backtest.spearman_correlation([1, 1, 1], [10, 20, 30]))


if __name__ == "__main__":
    unittest.main()