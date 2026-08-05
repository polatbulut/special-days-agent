"""Event case studies: did load factor move on the routes around a given date?

This module answers one narrow question, one event at a time:

    For the routes that an event should have filled, did LF move around the
    event date -- and was the move demand (pax up) or capacity (seats up)?

It is deliberately *not* an inference engine. It carries no significance
machinery of any kind -- no test statistics, no fitted coefficients. The output
is a description of what the extract shows, nothing more.

Method
------
1.  LF rows are aggregated to ``(route, date)`` with ``lf = sum(pax) /
    sum(capacity)`` -- capacity weighted, never the mean of per-flight LFs.
    A 300-seat flight at 90% and a 60-seat flight at 20% give 0.783, not 0.55.
2.  The baseline for a date ``d`` on a route is built from the **same route on
    the same weekday** within ``+/- baseline_days`` of ``d``, excluding ``d``
    itself and every date within ``guard_days`` of the event span. Day-of-week
    alone swings LF by more than 4 pp, so a same-weekday baseline is mandatory.
    Fewer than ``MIN_BASELINE_OBS`` clean observations means the baseline is
    ``None`` for that date and it is reported as such. There is no fallback to
    an overall mean -- a thin baseline must look thin.
3.  The three baseline numbers come from **one construction**, so they
    reconcile: ``baseline_pax`` and ``baseline_capacity`` are the pool means
    and ``baseline_lf = baseline_pax / baseline_capacity``, which equals the
    capacity-weighted ``sum(pax)/sum(capacity)`` over the pool. Three
    independent medians would not compose -- ``median(pax)/median(cap)`` is not
    ``median(pax/cap)`` -- and would let ``delta_pp`` disagree with the pax and
    capacity moves printed beside it.
4.  Each date in the window carries ``lf``, ``baseline_lf``, ``delta_pp``,
    ``pax``, ``capacity`` and ``pax_vs_baseline_pct`` /
    ``cap_vs_baseline_pct``. THY up-gauges known peaks, so an LF number without
    its pax and capacity move beside it can read a real demand rise as a null.
5.  The three summary numbers are **capacity-weighted aggregates over the event
    span**, not means of daily percentages: with ``P``/``C`` the pax and
    capacity summed over the in-event days that have a baseline and ``P0``/``C0``
    the same sums of those days' baselines,
    ``mean_delta_pp = 100 * (P/C - P0/C0)``,
    ``mean_pax_vs_baseline_pct = 100 * (P - P0) / P0`` and
    ``mean_cap_vs_baseline_pct = 100 * (C - C0) / C0``. All three are computed
    over one common day set. An unweighted mean would give a 60-seat ferry day
    the same weight as a 1000-seat day -- the mean-of-LFs error, one level up.
    (The fields keep their ``mean_`` names for the CSV/report contract.)

Reporting window vs. baseline hygiene
-------------------------------------
``guard_days`` does exactly one job: keep contaminated dates out of the
baseline pool. It does **not** widen the measured window -- ``in_event`` is the
event span itself. Sharing one parameter between the two would mean that
raising the guard averaged a one-day concert spike over ever more non-event
days and reported it at a fraction of its size.

Verdict rule (descriptive only, never a significance claim)
-----------------------------------------------------------
The label is composed from the three numbers, so it can never contradict them.
With ``d`` = mean delta_pp, ``p`` = mean pax_vs_baseline_pct, ``c`` = mean
cap_vs_baseline_pct, and the fixed thresholds ``LF_PP_THRESHOLD`` = 2.0 pp,
``PCT_THRESHOLD`` = 5.0 %::

    no route-days in the window            -> "no LF data"
    route-days, but none inside the span   -> "event days not in extract"
    no baseline on any in-event day        -> "no baseline"
    all three flat                         -> "no movement"
    otherwise                              -> "<demand>, <capacity>, <LF>"

where ``<demand>`` is ``demand up|flat|down`` from ``p`` against
``PCT_THRESHOLD``, ``<capacity>`` is ``capacity up|flat|cut`` from ``c``
against the same threshold, and ``<LF>`` is ``LF up|flat|down`` from ``d``
against ``LF_PP_THRESHOLD``. Every one of the 26 remaining combinations is
reachable and each names all three dimensions -- a 20% capacity change is never
described as "capacity flat", and an 8 pp LF fall is never described as flat.

Usage
-----
    python -m special_days.case_studies \
        --events events.csv --lf lf_route_day.csv --out out/case_studies \
        [--window-days 14] [--baseline-days 42] [--guard-days 3]

``events.csv`` columns: ``event, start_date, end_date, city, airport, routes,
note``. Blank ``end_date`` means the same day as ``start_date``. Blank
``routes`` means every route in the LF extract whose **DST** is the event
airport -- the inbound legs, which are what an event in that city fills.

``run`` writes six artifacts: ``summary.csv``, ``timeline.csv``,
``coverage.txt``, ``case_studies.xlsx``, ``case_studies.html`` and
``charts/*.svg``. The Excel / HTML / SVG renderers live in
``special_days.case_report`` and consume the dataclasses exported here.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import median
from typing import Any

DEFAULT_WINDOW_DAYS = 14
DEFAULT_BASELINE_DAYS = 42
DEFAULT_GUARD_DAYS = 3

#: A baseline built on fewer observations than this is reported as missing.
MIN_BASELINE_OBS = 3

#: Verdict thresholds. Fixed constants, not fitted to any data.
LF_PP_THRESHOLD = 2.0
PCT_THRESHOLD = 5.0

#: Rows at or below this capacity are kept but flagged; real extracts contain
#: tiny-capacity rows (ferry legs, data errors) that distort nothing on their
#: own but should be visible to whoever reads the timeline. Applied at both
#: grains: the route-day total and its constituent flight rows.
LOW_CAPACITY_FLAG_THRESHOLD = 200

_EVENT_REQUIRED_COLUMNS = ("event", "start date", "end date", "city", "airport", "routes")

_ORG_HEADERS = ("org", "org airport", "origin", "origin airport", "dep", "dep airport", "from")
_DST_HEADERS = ("dst", "dst airport", "destination", "destination airport", "arr", "arr airport", "to")
_ROUTE_HEADERS = ("route", "leg", "od", "o d", "market", "route code")
_DATE_HEADERS = ("ymd", "flight date", "date", "dep date", "departure date")
_PAX_HEADERS = ("boarded pax", "pax", "passengers", "boarded", "boarded passengers")
_CAPACITY_HEADERS = ("capacity", "seats", "seat capacity", "capacity seats")


# --------------------------------------------------------------------------
# Parsing helpers (salvaged from the retired backtest module)
# --------------------------------------------------------------------------


def _turkish_normalized(value: str) -> str:
    """ASCII-fold Turkish text, e.g. ``"İstanbul Bayramı"`` -> ``"istanbul bayrami"``."""
    translated = value.translate(
        str.maketrans(
            {
                "İ": "I",
                "I": "I",
                "ı": "i",
                "Ş": "S",
                "ş": "s",
                "Ğ": "G",
                "ğ": "g",
                "Ü": "U",
                "ü": "u",
                "Ö": "O",
                "ö": "o",
                "Ç": "C",
                "ç": "c",
            }
        )
    )
    text = unicodedata.normalize("NFKD", translated)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


def _normalize_header(value: Any) -> str:
    """Fold a column header to the lookup form used by the candidate tuples.

    Turkish-folding is not optional here. Excel's ``UPPER()`` under a Turkish
    locale maps ``i`` -> ``I``, so a THY-sourced extract arrives headed
    ``CAPACITY`` / ``ORIGIN``; a plain ``.lower()`` leaves the dotted I behind
    and the column then reads as missing from a file where it is plainly
    present. ``_turkish_normalized`` also collapses whitespace and strips
    punctuation, which subsumes the underscore handling -- every candidate
    header is alphanumeric-plus-space, so none of them is affected.
    """
    return _turkish_normalized(str(value or "").replace("_", " "))


def _clean_iata(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"[^A-Za-z]", "", str(value).strip()).upper()
    if len(text) != 3:
        return None
    return text


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    if len(text) == 8 and text.isdigit():  # YYYYMMDD, the YMD column shape
        try:
            return date(int(text[:4]), int(text[4:6]), int(text[6:8]))
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _load_sheet_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    """Read an ``.xlsx`` sheet (or a ``.csv``) as ``(headers, list-of-dicts)``."""
    text_path = str(path)
    if text_path.lower().endswith(".csv"):
        with open(text_path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            headers = [str(name).strip() for name in (reader.fieldnames or [])]
            data_rows = [
                {key: row.get(key) for key in headers}
                for row in reader
                if any((row.get(key) or "").strip() for key in headers)
            ]
        return headers, data_rows

    from openpyxl import load_workbook

    workbook = load_workbook(text_path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        if any(value not in (None, "") for value in row.values()):
            data_rows.append(row)
    return headers, data_rows


# --------------------------------------------------------------------------
# Data structures (consumed by special_days.case_report)
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class EventSpec:
    """One hand-written event from ``events.csv``."""

    row_id: int
    event: str
    start_date: date
    end_date: date
    city: str
    airport: str | None
    routes: tuple[tuple[str, str], ...]  # explicit (org, dst) pairs; empty = default
    note: str = ""

    def slug(self) -> str:
        """Filename/sheet-safe identifier, e.g. ``"kanye-west-concert-2025-06-14"``.

        The date is the disambiguator, so its width is reserved before the name
        is truncated. Slicing the whole string would drop the date first and
        leave two long same-venue events with identical slugs.
        """
        base = _turkish_normalized(self.event).replace(" ", "-") or f"event-{self.row_id}"
        iso = self.start_date.isoformat()
        room = max(1, 60 - len(iso) - 1)
        return f"{base[:room].rstrip('-')}-{iso}"

    def spans(self, day: date, guard_days: int) -> bool:
        """True if ``day`` falls inside the event span widened by ``guard_days``."""
        return self.start_date - timedelta(days=guard_days) <= day <= self.end_date + timedelta(days=guard_days)


@dataclass(frozen=True)
class RouteDay:
    """One ``(route, date)`` cell of the LF extract, capacity-weighted."""

    org: str
    dst: str
    day: date
    pax: int
    capacity: int
    n_flights: int
    #: How many *constituent flight rows* were anomalous. Kept separately
    #: because aggregation hides them: one 60-seat leg at LF 1.4 plus one
    #: 300-seat leg at LF 0.5 totals 360 seats at LF 0.65, and neither the
    #: overbooking nor the tiny gauge shows in the day total.
    n_rows_lf_gt_1: int = 0
    n_rows_low_capacity: int = 0

    @property
    def route(self) -> str:
        return f"{self.org}-{self.dst}"

    @property
    def lf(self) -> float:
        return self.pax / self.capacity


@dataclass(frozen=True)
class DayPoint:
    """One date in one event x route timeline."""

    day: date
    weekday: str
    in_event: bool
    lf: float
    baseline_lf: float | None
    delta_pp: float | None
    pax: int
    capacity: int
    baseline_pax: float | None
    baseline_capacity: float | None
    pax_vs_baseline_pct: float | None
    cap_vs_baseline_pct: float | None
    n_baseline_obs: int
    n_flights: int
    data_flags: str = ""

    # Flat aliases for the renderers in special_days.case_report, which read a
    # denormalized point. Read-only, so there is exactly one source of truth.
    @property
    def date(self) -> date:
        return self.day

    @property
    def baseline(self) -> float | None:
        return self.baseline_lf

    @property
    def flags(self) -> str:
        return self.data_flags


@dataclass(frozen=True)
class EventRouteResult:
    """Everything measured for one event on one route.

    ``spec`` carries the full event definition; ``event``/``start_date``/
    ``end_date``/``city``/``note`` are flat read-only views onto it so the
    renderers never have to reach through two levels.
    """

    spec: EventSpec
    org: str
    dst: str
    timeline: list[DayPoint] = field(default_factory=list)
    mean_delta_pp: float | None = None
    mean_pax_vs_baseline_pct: float | None = None
    mean_cap_vs_baseline_pct: float | None = None
    # How far this route wanders from its own baseline on ORDINARY days in the
    # window (median |delta_pp| over non-event days). Not a significance test --
    # it is the yardstick a reader needs to tell a real move from routine noise.
    # Without it a +3 pp wobble and a +14 pp lift look like the same kind of fact.
    normal_swing_pp: float | None = None
    n_days: int = 0
    n_days_no_baseline: int = 0
    n_route_days: int = 0
    n_flights: int = 0
    verdict: str = "no LF data"

    @property
    def route(self) -> str:
        return f"{self.org}-{self.dst}"

    @property
    def event(self) -> str:
        return self.spec.event

    @property
    def start_date(self) -> date:
        return self.spec.start_date

    @property
    def end_date(self) -> date:
        return self.spec.end_date

    @property
    def city(self) -> str:
        return self.spec.city

    @property
    def note(self) -> str:
        return self.spec.note

    @property
    def mean_pax_pct(self) -> float | None:
        return self.mean_pax_vs_baseline_pct

    @property
    def mean_cap_pct(self) -> float | None:
        return self.mean_cap_vs_baseline_pct


@dataclass(frozen=True)
class Coverage:
    """What the run actually saw, for ``coverage.txt``."""

    lf_min_date: date | None
    lf_max_date: date | None
    n_routes: int
    n_route_days: int
    n_flight_rows: int
    n_skipped_bad_capacity: int
    n_skipped_unparseable: int
    n_events: int
    events_without_lf: list[str] = field(default_factory=list)
    #: Routes named explicitly in events.csv that the extract never contained.
    missing_routes: list[str] = field(default_factory=list)
    #: Dates that arrived as ``DD/MM/YYYY``-style text. Ambiguous with the US
    #: ordering, and a misread date silently reassigns the row's weekday --
    #: which is the whole baseline construction. Reported, never guessed at.
    n_slash_format_dates: int = 0


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------


def _resolve(by_normalized: dict[str, str], candidates: tuple[str, ...]) -> str | None:
    for name in candidates:
        original = by_normalized.get(name)
        if original is not None:
            return original
    return None


def _parse_route_token(token: str) -> tuple[str, str] | None:
    """Accept ``"SVO-IST"``, ``"SVO/IST"``, ``"SVO IST"`` or ``"SVOIST"``."""
    text = re.sub(r"[^A-Za-z]", "", str(token or "")).upper()
    if len(text) != 6:
        return None
    return text[:3], text[3:]


def load_events(path: str | Path) -> list[EventSpec]:
    """Read the hand-written event list. Missing columns are named explicitly."""
    _, raw_rows = _load_sheet_rows(path)
    if not raw_rows:
        raise ValueError(f"Events file has no data rows: {path}")

    by_normalized = {_normalize_header(name): name for name in raw_rows[0].keys()}
    missing = [name for name in _EVENT_REQUIRED_COLUMNS if name not in by_normalized]
    if missing:
        raise ValueError(
            "Events file is missing required column(s): "
            + ", ".join(name.replace(" ", "_") for name in missing)
        )
    col = {name.replace(" ", "_"): by_normalized[name] for name in _EVENT_REQUIRED_COLUMNS}
    note_col = by_normalized.get("note")

    events: list[EventSpec] = []
    for row_id, row in enumerate(raw_rows, start=1):
        name = str(row.get(col["event"], "") or "").strip()
        start = _parse_date(row.get(col["start_date"]))
        if not name or start is None:
            raise ValueError(f"Events file row {row_id}: 'event' and 'start_date' are required")
        end = _parse_date(row.get(col["end_date"])) or start
        if end < start:
            raise ValueError(f"Events file row {row_id} ({name}): end_date is before start_date")

        airport = _clean_iata(row.get(col["airport"]))
        routes: list[tuple[str, str]] = []
        for token in re.split(r"[;,]", str(row.get(col["routes"], "") or "")):
            if not token.strip():
                continue
            pair = _parse_route_token(token)
            if pair is None:
                raise ValueError(
                    f"Events file row {row_id} ({name}): cannot read route {token.strip()!r}; "
                    "use ORG-DST or ORGDST"
                )
            routes.append(pair)
        if airport is None and not routes:
            raise ValueError(
                f"Events file row {row_id} ({name}): needs an 'airport' IATA code or an explicit 'routes' list"
            )

        events.append(
            EventSpec(
                row_id=row_id,
                event=name,
                start_date=start,
                end_date=end,
                city=str(row.get(col["city"], "") or "").strip(),
                airport=airport,
                routes=tuple(dict.fromkeys(routes)),
                note=str(row.get(note_col, "") or "").strip() if note_col else "",
            )
        )
    return events


def load_route_days(path: str | Path, sheet_name: str | None = None) -> tuple[dict[tuple[str, str], dict[date, RouteDay]], Coverage]:
    """Aggregate the LF extract to capacity-weighted ``(route, date)`` cells."""
    _, raw_rows = _load_sheet_rows(path, sheet_name)
    if not raw_rows:
        raise ValueError(f"LF file has no data rows: {path}")

    by_normalized = {_normalize_header(name): name for name in raw_rows[0].keys()}
    org_col = _resolve(by_normalized, _ORG_HEADERS)
    dst_col = _resolve(by_normalized, _DST_HEADERS)
    route_col = _resolve(by_normalized, _ROUTE_HEADERS)
    if (org_col is None or dst_col is None) and route_col is None:
        raise ValueError("LF file is missing required column(s): ORG and DST (or a single ROUTE column)")
    date_col = _resolve(by_normalized, _DATE_HEADERS)
    if date_col is None:
        raise ValueError("LF file is missing required column: FLIGHT_DATE (or YMD / DATE)")
    pax_col = _resolve(by_normalized, _PAX_HEADERS)
    if pax_col is None:
        raise ValueError("LF file is missing required column: BOARDED_PAX (or PAX)")
    capacity_col = _resolve(by_normalized, _CAPACITY_HEADERS)
    if capacity_col is None:
        raise ValueError("LF file is missing required column: CAPACITY (or SEATS)")

    # Per (route, day): [pax, capacity, n_flights, n_rows_lf_gt_1, n_rows_low_cap]
    totals: dict[tuple[str, str], dict[date, list[float]]] = {}
    n_rows = 0
    n_bad_capacity = 0
    n_unparseable = 0
    n_slash_dates = 0
    min_date: date | None = None
    max_date: date | None = None

    for row in raw_rows:
        org = _clean_iata(row.get(org_col)) if org_col else None
        dst = _clean_iata(row.get(dst_col)) if dst_col else None
        if (org is None or dst is None) and route_col is not None:
            pair = _parse_route_token(str(row.get(route_col, "") or ""))
            if pair is not None:
                org, dst = pair
        raw_day = row.get(date_col)
        day = _parse_date(raw_day)
        pax = _parse_float(row.get(pax_col))
        capacity = _parse_float(row.get(capacity_col))
        if org is None or dst is None or day is None or pax is None or capacity is None:
            n_unparseable += 1
            continue
        if capacity <= 0:
            n_bad_capacity += 1
            continue
        if isinstance(raw_day, str) and "/" in raw_day:
            n_slash_dates += 1

        bucket = totals.setdefault((org, dst), {}).setdefault(day, [0.0, 0.0, 0.0, 0.0, 0.0])
        bucket[0] += pax
        bucket[1] += capacity
        bucket[2] += 1
        if pax / capacity > 1.0:
            bucket[3] += 1
        if capacity <= LOW_CAPACITY_FLAG_THRESHOLD:
            bucket[4] += 1
        n_rows += 1
        min_date = day if min_date is None or day < min_date else min_date
        max_date = day if max_date is None or day > max_date else max_date

    route_days: dict[tuple[str, str], dict[date, RouteDay]] = {}
    n_route_days = 0
    for (org, dst), by_day in totals.items():
        cells: dict[date, RouteDay] = {}
        for day, values in by_day.items():
            capacity_total = int(round(values[1]))
            if capacity_total <= 0:
                # A fractional capacity (0.4) clears the per-row ``<= 0`` gate
                # and then rounds to zero here. Dropping the cell is the only
                # way ``RouteDay.lf`` cannot raise ZeroDivisionError later.
                n_bad_capacity += int(values[2])
                n_rows -= int(values[2])
                continue
            cells[day] = RouteDay(
                org,
                dst,
                day,
                int(round(values[0])),
                capacity_total,
                int(values[2]),
                n_rows_lf_gt_1=int(values[3]),
                n_rows_low_capacity=int(values[4]),
            )
        if not cells:
            continue
        route_days[(org, dst)] = cells
        n_route_days += len(cells)

    coverage = Coverage(
        lf_min_date=min_date,
        lf_max_date=max_date,
        n_routes=len(route_days),
        n_route_days=n_route_days,
        n_flight_rows=n_rows,
        n_skipped_bad_capacity=n_bad_capacity,
        n_skipped_unparseable=n_unparseable,
        n_events=0,
        n_slash_format_dates=n_slash_dates,
    )
    return route_days, coverage


# --------------------------------------------------------------------------
# Measurement
# --------------------------------------------------------------------------


def _routes_for(event: EventSpec, route_days: dict[tuple[str, str], dict[date, RouteDay]]) -> list[tuple[str, str]]:
    """Explicit routes as written, else every inbound leg into the event airport."""
    if event.routes:
        return list(event.routes)
    if event.airport is None:
        return []
    return sorted(key for key in route_days if key[1] == event.airport)


def _flags_for(cell: RouteDay) -> str:
    """Anomaly flags for one route-day, at both the day and the flight-row grain.

    ``lf>1`` / ``low_capacity`` describe the day total; ``row_lf>1`` /
    ``row_low_capacity`` fire when the day total looks ordinary but at least one
    of the flights that built it did not.
    """
    flags: list[str] = []
    if cell.lf > 1.0:
        flags.append("lf>1")
    elif cell.n_rows_lf_gt_1:
        flags.append("row_lf>1")
    if cell.capacity <= LOW_CAPACITY_FLAG_THRESHOLD:
        flags.append("low_capacity")
    elif cell.n_rows_low_capacity:
        flags.append("row_low_capacity")
    return "|".join(flags)


def _pct_change(value: float, baseline: float | None) -> float | None:
    if baseline is None or baseline == 0:
        return None
    return 100.0 * (value - baseline) / baseline


def _direction(value: float, threshold: float, up: str, flat: str, down: str) -> str:
    if value >= threshold:
        return up
    if value <= -threshold:
        return down
    return flat


def _verdict(
    delta_pp: float | None,
    pax_pct: float | None,
    cap_pct: float | None,
    n_route_days: int,
    n_event_days: int | None = None,
) -> str:
    """Descriptive label from the fixed rule documented in the module docstring.

    The label is *composed* from the three numbers rather than picked out of a
    branch table, so it always names all three dimensions and can never assert
    "capacity flat" about a 20% capacity change or leave a 12 pp LF move
    unmentioned.
    """
    if n_route_days == 0:
        return "no LF data"
    if n_event_days == 0:
        # Window rows exist but none of them fall inside the event span: the
        # extract simply stops before (or starts after) the event. That is a
        # different failure from a baseline that came out too thin.
        return "event days not in extract"
    if delta_pp is None or pax_pct is None or cap_pct is None:
        return "no baseline"

    demand = _direction(pax_pct, PCT_THRESHOLD, "demand up", "demand flat", "demand down")
    capacity = _direction(cap_pct, PCT_THRESHOLD, "capacity up", "capacity flat", "capacity cut")
    load = _direction(delta_pp, LF_PP_THRESHOLD, "LF up", "LF flat", "LF down")
    if (demand, capacity, load) == ("demand flat", "capacity flat", "LF flat"):
        return "no movement"
    return f"{demand}, {capacity}, {load}"


def measure(
    event: EventSpec,
    org: str,
    dst: str,
    cells: dict[date, RouteDay],
    *,
    window_days: int = DEFAULT_WINDOW_DAYS,
    baseline_days: int = DEFAULT_BASELINE_DAYS,
    guard_days: int = DEFAULT_GUARD_DAYS,
) -> EventRouteResult:
    """Build the timeline and the three summary numbers for one event x route."""
    window_start = event.start_date - timedelta(days=window_days)
    window_end = event.end_date + timedelta(days=window_days)

    # The clean pool: every observed day on this route that the event (widened
    # by the guard) does not touch. Baselines are drawn only from here, so the
    # event can never be part of its own comparison.
    clean = [cell for day, cell in cells.items() if not event.spans(day, guard_days)]

    timeline: list[DayPoint] = []
    for day in sorted(d for d in cells if window_start <= d <= window_end):
        cell = cells[day]
        pool = [
            other
            for other in clean
            # ``other.day != day``: a day must not be a member of its own
            # comparison set. Without it every context day is shrunk toward zero
            # and n_baseline_obs overstates the evidence by one, which also lets
            # a day with two real peers slip past MIN_BASELINE_OBS.
            if other.day != day
            and other.day.weekday() == day.weekday()
            and abs((other.day - day).days) <= baseline_days
        ]
        pool_capacity = sum(other.capacity for other in pool)
        if len(pool) >= MIN_BASELINE_OBS and pool_capacity > 0:
            # One construction for all three: the pool means, with the LF read
            # off them. baseline_pax / baseline_capacity == sum(pax)/sum(cap),
            # so delta_pp is arithmetically implied by the pax and cap moves
            # printed next to it. Three separate medians would not compose.
            baseline_pax = sum(other.pax for other in pool) / len(pool)
            baseline_capacity = pool_capacity / len(pool)
            baseline_lf = baseline_pax / baseline_capacity
            delta_pp = 100.0 * (cell.lf - baseline_lf)
        else:
            baseline_lf = baseline_pax = baseline_capacity = delta_pp = None

        timeline.append(
            DayPoint(
                day=day,
                weekday=day.strftime("%a"),
                # The event span itself -- never widened by the guard. The guard
                # is baseline hygiene; using it here would average a one-day
                # spike over 2*guard+1 days and report it at a fraction of size.
                in_event=event.spans(day, 0),
                lf=cell.lf,
                baseline_lf=baseline_lf,
                delta_pp=delta_pp,
                pax=cell.pax,
                capacity=cell.capacity,
                baseline_pax=baseline_pax,
                baseline_capacity=baseline_capacity,
                pax_vs_baseline_pct=_pct_change(float(cell.pax), baseline_pax),
                cap_vs_baseline_pct=_pct_change(float(cell.capacity), baseline_capacity),
                n_baseline_obs=len(pool),
                n_flights=cell.n_flights,
                data_flags=_flags_for(cell),
            )
        )

    in_event = [point for point in timeline if point.in_event]
    scored = [point for point in in_event if point.delta_pp is not None]

    # Capacity-weighted aggregates over one common day set, not means of daily
    # percentages: an unweighted mean would let a 60-seat day at +30 pp outvote
    # a 1000-seat day at +1 pp, which is the mean-of-LFs error one level up.
    event_pax = float(sum(point.pax for point in scored))
    event_capacity = float(sum(point.capacity for point in scored))
    base_pax = sum(point.baseline_pax or 0.0 for point in scored)
    base_capacity = sum(point.baseline_capacity or 0.0 for point in scored)
    if scored and event_capacity > 0 and base_capacity > 0:
        mean_delta_pp = 100.0 * (event_pax / event_capacity - base_pax / base_capacity)
    else:
        mean_delta_pp = None
    mean_pax_pct = _pct_change(event_pax, base_pax) if scored else None
    mean_cap_pct = _pct_change(event_capacity, base_capacity) if scored else None

    # The route's ordinary swing, from the non-event days of this same window.
    # Deliberately the median of absolute deviations rather than a standard
    # deviation: one unrelated disruption in the window would inflate an sd and
    # make a genuine event look unremarkable.
    off_event = [
        abs(point.delta_pp)
        for point in timeline
        if not point.in_event and point.delta_pp is not None
    ]
    normal_swing_pp = median(off_event) if off_event else None

    return EventRouteResult(
        spec=event,
        org=org,
        dst=dst,
        timeline=timeline,
        mean_delta_pp=mean_delta_pp,
        mean_pax_vs_baseline_pct=mean_pax_pct,
        mean_cap_vs_baseline_pct=mean_cap_pct,
        normal_swing_pp=normal_swing_pp,
        n_days=len(in_event),
        n_days_no_baseline=len(in_event) - len(scored),
        n_route_days=len(timeline),
        n_flights=sum(point.n_flights for point in timeline),
        verdict=_verdict(mean_delta_pp, mean_pax_pct, mean_cap_pct, len(timeline), len(in_event)),
    )


def analyse(
    events: list[EventSpec],
    route_days: dict[tuple[str, str], dict[date, RouteDay]],
    *,
    window_days: int = DEFAULT_WINDOW_DAYS,
    baseline_days: int = DEFAULT_BASELINE_DAYS,
    guard_days: int = DEFAULT_GUARD_DAYS,
) -> tuple[list[EventRouteResult], list[str]]:
    """Measure every event x route. Returns ``(results, events_without_lf)``.

    Results come out in *events file order*, then route order within an event.
    They are NEVER sorted by effect size: the event list is fixed up front and
    whatever it shows -- flat, negative or large -- is what gets reported.
    Ranking by outcome would turn a fixed list into a cherry-picked one.
    """
    results: list[EventRouteResult] = []
    events_without_lf: list[str] = []

    for event in events:
        routes = _routes_for(event, route_days)
        produced = 0
        for org, dst in routes:
            cells = route_days.get((org, dst), {})
            result = measure(
                event,
                org,
                dst,
                cells,
                window_days=window_days,
                baseline_days=baseline_days,
                guard_days=guard_days,
            )
            # A route with no rows still emits a result, so a genuine null never
            # looks the same as a broken join.
            results.append(result)
            # Count in-event days, not window rows: an extract that covers the
            # run-up but stops before the event produces plenty of window rows
            # and still measures nothing.
            produced += result.n_days
        if not routes or produced == 0:
            events_without_lf.append(event.event)

    return results, events_without_lf


# --------------------------------------------------------------------------
# CSV output
# --------------------------------------------------------------------------


def _fmt(value: float | None, digits: int = 4) -> str:
    return "" if value is None else f"{value:.{digits}f}"


_SUMMARY_HEADER = [
    "event",
    "start_date",
    "end_date",
    "city",
    "route",
    "org",
    "dst",
    "mean_delta_pp",
    "normal_swing_pp",
    "mean_pax_vs_baseline_pct",
    "mean_cap_vs_baseline_pct",
    "n_days",
    "n_days_no_baseline",
    "n_route_days",
    "n_flights",
    "verdict",
    "note",
]

_TIMELINE_HEADER = [
    "event",
    "route",
    "date",
    "weekday",
    "in_event",
    "lf",
    "baseline_lf",
    "delta_pp",
    "pax",
    "capacity",
    "baseline_pax",
    "baseline_capacity",
    "pax_vs_baseline_pct",
    "cap_vs_baseline_pct",
    "n_baseline_obs",
    "n_flights",
    "data_flags",
]


def write_summary_csv(results: list[EventRouteResult], path: str | Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(_SUMMARY_HEADER)
        for result in results:
            writer.writerow(
                [
                    result.event,
                    result.start_date.isoformat(),
                    result.end_date.isoformat(),
                    result.city,
                    result.route,
                    result.org,
                    result.dst,
                    _fmt(result.mean_delta_pp, 2),
                    _fmt(result.normal_swing_pp, 2),
                    _fmt(result.mean_pax_vs_baseline_pct, 2),
                    _fmt(result.mean_cap_vs_baseline_pct, 2),
                    result.n_days,
                    result.n_days_no_baseline,
                    result.n_route_days,
                    result.n_flights,
                    result.verdict,
                    result.note,
                ]
            )


def write_timeline_csv(results: list[EventRouteResult], path: str | Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(_TIMELINE_HEADER)
        for result in results:
            for point in result.timeline:
                writer.writerow(
                    [
                        result.event,
                        result.route,
                        point.day.isoformat(),
                        point.weekday,
                        "1" if point.in_event else "0",
                        _fmt(point.lf),
                        _fmt(point.baseline_lf),
                        _fmt(point.delta_pp, 2),
                        point.pax,
                        point.capacity,
                        _fmt(point.baseline_pax, 1),
                        _fmt(point.baseline_capacity, 1),
                        _fmt(point.pax_vs_baseline_pct, 2),
                        _fmt(point.cap_vs_baseline_pct, 2),
                        point.n_baseline_obs,
                        point.n_flights,
                        point.data_flags,
                    ]
                )


def write_coverage_txt(coverage: Coverage, results: list[EventRouteResult], path: str | Path) -> None:
    lines = [
        "case studies coverage",
        "=====================",
        f"LF date range          : {coverage.lf_min_date} .. {coverage.lf_max_date}",
        f"routes in LF file      : {coverage.n_routes}",
        f"route-days in LF file  : {coverage.n_route_days}",
        f"flight rows read       : {coverage.n_flight_rows}",
        f"rows skipped (capacity <= 0) : {coverage.n_skipped_bad_capacity}",
        f"rows skipped (unparseable)   : {coverage.n_skipped_unparseable}",
        f"events processed       : {coverage.n_events}",
        f"event x route results  : {len(results)}",
        f"events with no measured event day: {len(coverage.events_without_lf)}",
    ]
    lines.extend(f"  - {name}" for name in coverage.events_without_lf)
    lines.append(f"routes requested but absent from the extract: {len(coverage.missing_routes)}")
    lines.extend(f"  - {name}" for name in coverage.missing_routes)
    if coverage.n_slash_format_dates:
        lines.append(
            f"WARNING: {coverage.n_slash_format_dates} date(s) arrived as slash-separated text "
            "and were read as DD/MM/YYYY. If the extract came from a US-locale export they are "
            "wrong by up to 11 months, and the same-weekday baseline is built on the weekday. "
            "Re-export as YYYY-MM-DD to remove the ambiguity."
        )
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------
# Entry points
# --------------------------------------------------------------------------


def run(
    events_path: str | Path,
    lf_path: str | Path,
    out_dir: str | Path,
    *,
    window_days: int = DEFAULT_WINDOW_DAYS,
    baseline_days: int = DEFAULT_BASELINE_DAYS,
    guard_days: int = DEFAULT_GUARD_DAYS,
    lf_sheet: str | None = None,
    reports: bool = True,
) -> list[EventRouteResult]:
    """Measure every event x route and write every artifact to ``out_dir``.

    Always writes ``summary.csv``, ``timeline.csv`` and ``coverage.txt``. With
    ``reports`` (the default) it also writes ``case_studies.xlsx``,
    ``case_studies.html`` and one SVG per result under ``charts/`` -- the
    renderers are part of the deliverable, not an optional extra, so the CLI
    must not leave half of it unwritten.
    """
    events = load_events(events_path)
    route_days, coverage = load_route_days(lf_path, lf_sheet)
    results, events_without_lf = analyse(
        events,
        route_days,
        window_days=window_days,
        baseline_days=baseline_days,
        guard_days=guard_days,
    )
    missing_routes = [
        f"{event.event}: {org}-{dst}"
        for event in events
        for org, dst in _routes_for(event, route_days)
        if (org, dst) not in route_days
    ]
    coverage = replace(
        coverage,
        n_events=len(events),
        events_without_lf=events_without_lf,
        missing_routes=missing_routes,
    )

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    write_summary_csv(results, out / "summary.csv")
    write_timeline_csv(results, out / "timeline.csv")
    write_coverage_txt(coverage, results, out / "coverage.txt")

    if reports:
        from . import case_report

        case_report.write_workbook(results, str(out / "case_studies.xlsx"))
        case_report.write_html(results, str(out / "case_studies.html"))
        case_report.write_svgs(results, str(out / "charts"))
    return results


def _positive_int(value: str) -> int:
    number = int(value)
    if number <= 0:
        raise argparse.ArgumentTypeError("must be greater than 0")
    return number


def _non_negative_int(value: str) -> int:
    number = int(value)
    if number < 0:
        raise argparse.ArgumentTypeError("must be 0 or greater")
    return number


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="special_days.case_studies",
        description="Per-event LF case studies on route-day data (descriptive, no significance tests).",
    )
    parser.add_argument("--events", required=True, help="Hand-written event list (.csv or .xlsx).")
    parser.add_argument("--lf", required=True, help="Route-day LF extract (.csv or .xlsx).")
    parser.add_argument("--out", required=True, help="Output directory for summary/timeline/coverage.")
    parser.add_argument("--lf-sheet", default=None, help="Sheet name when --lf is an .xlsx workbook.")
    parser.add_argument(
        "--window-days",
        type=_positive_int,
        default=DEFAULT_WINDOW_DAYS,
        help=f"Days shown either side of the event span (default: {DEFAULT_WINDOW_DAYS}).",
    )
    parser.add_argument(
        "--baseline-days",
        type=_positive_int,
        default=DEFAULT_BASELINE_DAYS,
        help=f"Same-weekday baseline reach either side of each date (default: {DEFAULT_BASELINE_DAYS}).",
    )
    parser.add_argument(
        "--guard-days",
        type=_non_negative_int,
        default=DEFAULT_GUARD_DAYS,
        help=(
            "Days around the event span excluded from every baseline "
            f"(default: {DEFAULT_GUARD_DAYS}). Baseline hygiene only -- it never "
            "widens the measured window."
        ),
    )
    parser.add_argument(
        "--csv-only",
        action="store_true",
        help="Skip the xlsx/html/svg renderers and write only the three text files.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        results = run(
            args.events,
            args.lf,
            args.out,
            window_days=args.window_days,
            baseline_days=args.baseline_days,
            guard_days=args.guard_days,
            lf_sheet=args.lf_sheet,
            reports=not args.csv_only,
        )
    except (ValueError, FileNotFoundError, ZeroDivisionError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2

    print(f"Wrote {len(results)} event x route result(s) to {os.path.join(str(args.out), '')}")
    for result in results:  # input order, never ranked
        print(f"  {result.event} | {result.route} | {_fmt(result.mean_delta_pp, 2)} pp | {result.verdict}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
