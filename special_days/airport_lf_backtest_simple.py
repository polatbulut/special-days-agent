"""Simple airport-specific LF backtest using an already-enriched events workbook.

This module is intentionally narrower than ``special_days.backtest``.
It assumes the events workbook already has airport-date enrichment enabled via
the ``event_day_airport`` sheet and then answers one question only:

* when an event happens at airport X on date D, does LF at airport X on date D
  move with it?

The join grain is therefore strictly ``AIRPORT`` plus ``flight_date``.
"""

from __future__ import annotations

import argparse
import html
import math
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

LF_HEADERS = [
    "YMD",
    "AIRPORT",
    "BOARDED_PAX",
    "CAPACITY",
    "LF",
    "flight_date",
    "LF_PCT",
]

EVENT_DAY_AIRPORT_HEADERS = [
    "row_id",
    "Event",
    "Source",
    "City",
    "Airport",
    "Event date",
    "Impact",
    "Predicted attendance",
    "Start date",
    "End date",
    "Bridge start",
    "Bridge end",
    "Effective start",
    "Effective end",
    "Mapping method",
    "Curve source",
]

_DATE_FORMAT = "yyyy-mm-dd"
_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
_ONE_DAY = timedelta(days=1)
_SUMMER_MONTHS = {6, 7, 8}
_FOCUS_PERIODS = {
    "bayram": {
        "label": "Bayram",
        "basis": "event_named_period",
        "color": "#34a853",
    },
    "summer_vacation": {
        "label": "Summer vacation",
        "basis": "event_named_period",
        "color": "#f59e0b",
    },
    "school_break": {
        "label": "School break",
        "basis": "event_named_period",
        "color": "#0ea5e9",
    },
    "summer_months": {
        "label": "Summer months (Jun-Aug)",
        "basis": "calendar_season",
        "color": "#fcd34d",
    },
}
_MONTH_NAMES = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec",
}


def _iso_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("must be a date in YYYY-MM-DD format") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="special_days.airport_lf_backtest_simple",
        description=(
            "Airport-specific LF backtest using special_days_enriched. "
            "Join grain is AIRPORT + flight_date only."
        ),
    )
    parser.add_argument("--lf", required=True, help="Path to the LF workbook (.xlsx).")
    parser.add_argument(
        "--events-enriched",
        required=True,
        help="Path to the special_days_enriched workbook (.xlsx).",
    )
    parser.add_argument(
        "--out-dir",
        default="out/airport_lf_backtest_simple",
        help="Output directory for the workbook (default: out/airport_lf_backtest_simple).",
    )
    parser.add_argument("--lf-sheet", default=None, help="Optional LF sheet name.")
    parser.add_argument(
        "--events-sheet",
        default="event_day_airport",
        help="Sheet in the enriched workbook containing airport-day event rows (default: event_day_airport).",
    )
    parser.add_argument(
        "--start-date",
        type=_iso_date,
        default=None,
        help="Optional inclusive start date YYYY-MM-DD. Defaults to the first LF date.",
    )
    parser.add_argument(
        "--end-date",
        type=_iso_date,
        default=None,
        help="Optional inclusive end date YYYY-MM-DD. Defaults to the last LF date.",
    )
    parser.add_argument(
        "--airport",
        default=None,
        help="Single airport to export as a timeline sheet and graph.",
    )
    parser.add_argument(
        "--airports",
        default=None,
        help=(
            "Comma-separated airports to export as separate timeline sheets and graphs. "
            "Defaults to all airports with matched event days."
        ),
    )
    return parser


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("_", " "))


def _turkish_normalized(value: str) -> str:
    translated = value.translate(
        str.maketrans(
            {
                "İ": "I",
                "I": "I",
                "ı": "i",
                "Ş": "S",
                "ş": "s",
                "Ğ": "G",
                "ğ": "g",
                "Ü": "U",
                "ü": "u",
                "Ö": "O",
                "ö": "o",
                "Ç": "C",
                "ç": "c",
            }
        )
    )
    text = unicodedata.normalize("NFKD", translated)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


def _clean_iata(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"[^A-Za-z]", "", str(value).strip()).upper()
    if len(text) != 3:
        return None
    return text


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_airports(value: str | None) -> list[str] | None:
    if value is None:
        return None
    airports: list[str] = []
    seen: set[str] = set()
    for part in value.split(","):
        airport = _clean_iata(part)
        if not airport or airport in seen:
            continue
        seen.add(airport)
        airports.append(airport)
    return airports or None


def _split_event_names(value: Any) -> list[str]:
    return [part.strip() for part in str(value or "").split("|") if part.strip()]


def _focus_period_keys_for_event_name(event_name: str) -> list[str]:
    normalized = _turkish_normalized(event_name)
    keys: list[str] = []
    if "bayram" in normalized:
        keys.append("bayram")
    if "yaz tatili" in normalized or "summer vacation" in normalized:
        keys.append("summer_vacation")
    if "yariyil tatili" in normalized or "ara tatili" in normalized or "school break" in normalized:
        keys.append("school_break")
    return keys


def _focus_period_keys_from_event_names(event_names: Any) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for event_name in _split_event_names(event_names):
        for key in _focus_period_keys_for_event_name(event_name):
            if key in seen:
                continue
            seen.add(key)
            keys.append(key)
    return keys


def _focus_period_labels(keys: list[str]) -> str:
    return " | ".join(_FOCUS_PERIODS[key]["label"] for key in keys)


def _row_focus_period_keys(row: dict[str, Any]) -> list[str]:
    return _focus_period_keys_from_event_names(row.get("event_names", ""))


def _row_has_focus_period(row: dict[str, Any], period_key: str) -> bool:
    if period_key == "summer_months":
        flight_date = row.get("flight_date")
        return isinstance(flight_date, date) and flight_date.month in _SUMMER_MONTHS
    return period_key in _row_focus_period_keys(row)


def _row_focus_period_names(row: dict[str, Any], period_key: str) -> list[str]:
    if period_key == "summer_months":
        return []
    names: list[str] = []
    seen: set[str] = set()
    for event_name in _split_event_names(row.get("event_names", "")):
        if period_key in _focus_period_keys_for_event_name(event_name) and event_name not in seen:
            seen.add(event_name)
            names.append(event_name)
    return names


def _load_sheet_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        try:
            sheet = workbook[sheet_name] if sheet_name else workbook.active
        except KeyError as exc:
            raise ValueError(f"Workbook is missing sheet: {sheet_name}") from exc
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        if any(value not in (None, "") for value in row.values()):
            data_rows.append(row)
    return headers, data_rows


def _required_lf_columns(rows: list[dict[str, Any]]) -> dict[str, str]:
    if not rows:
        return {}
    by_normalized = {_normalize_header(name): name for name in rows[0].keys()}
    resolved: dict[str, str] = {}

    airport = by_normalized.get("airport")
    if airport is None:
        raise ValueError("LF workbook is missing required column: AIRPORT")
    resolved["airport"] = airport

    flight_date = by_normalized.get("flight date") or by_normalized.get("flightdate")
    ymd = by_normalized.get("ymd")
    if flight_date is None and ymd is None:
        raise ValueError("LF workbook must contain either flight_date or YMD")
    if flight_date is not None:
        resolved["flight_date"] = flight_date
    if ymd is not None:
        resolved["ymd"] = ymd

    lf = by_normalized.get("lf")
    lf_pct = by_normalized.get("lf pct")
    if lf is None and lf_pct is None:
        raise ValueError("LF workbook must contain either LF or LF_PCT")
    if lf is not None:
        resolved["lf"] = lf
    if lf_pct is not None:
        resolved["lf_pct"] = lf_pct

    for optional in ("boarded pax", "capacity"):
        original = by_normalized.get(optional)
        if original is not None:
            resolved[optional.replace(" ", "_")] = original
    return resolved


def load_lf_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    headers, raw_rows = _load_sheet_rows(path, sheet_name)
    col = _required_lf_columns(raw_rows)
    parsed_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        airport = _clean_iata(row.get(col["airport"]))
        if not airport:
            continue

        flight_date = None
        if "flight_date" in col:
            flight_date = _parse_date(row.get(col["flight_date"]))
        if flight_date is None and "ymd" in col:
            ymd_text = str(row.get(col["ymd"], "") or "").strip()
            if len(ymd_text) == 8 and ymd_text.isdigit():
                flight_date = date(int(ymd_text[:4]), int(ymd_text[4:6]), int(ymd_text[6:8]))
        if flight_date is None:
            continue

        lf = _parse_float(row.get(col.get("lf", "")))
        lf_pct = _parse_float(row.get(col.get("lf_pct", "")))
        if lf is None and lf_pct is not None:
            lf = lf_pct / 100.0
        if lf_pct is None and lf is not None:
            lf_pct = round(lf * 100.0, 2)

        parsed = dict(row)
        parsed["AIRPORT"] = airport
        parsed["flight_date"] = flight_date
        parsed["LF"] = lf
        parsed["LF_PCT"] = lf_pct
        parsed_rows.append(parsed)
    return headers, parsed_rows


def _required_event_day_airport_columns(rows: list[dict[str, Any]]) -> dict[str, str]:
    if not rows:
        raise ValueError(
            "Enrichment is not enabled correctly: event_day_airport sheet is empty or missing data rows."
        )
    by_normalized = {_normalize_header(name): name for name in rows[0].keys()}
    required = {
        "airport": "airport",
        "event date": "event_date",
        "event": "event",
    }
    resolved: dict[str, str] = {}
    missing: list[str] = []
    for header, alias in required.items():
        original = by_normalized.get(header)
        if original is None:
            missing.append(header)
        else:
            resolved[alias] = original

    if missing:
        raise ValueError(
            "Enrichment is not enabled correctly: event_day_airport sheet is missing required columns: "
            + ", ".join(missing)
        )

    optional_map = {
        "impact": "impact",
        "source": "source",
        "predicted attendance": "predicted_attendance",
    }
    for header, alias in optional_map.items():
        original = by_normalized.get(header)
        if original is not None:
            resolved[alias] = original
    return resolved


def load_enriched_event_rows(path: str | Path, sheet_name: str = "event_day_airport") -> tuple[list[str], list[dict[str, Any]]]:
    try:
        headers, raw_rows = _load_sheet_rows(path, sheet_name)
    except ValueError as exc:
        raise ValueError(
            "Enrichment is not enabled correctly: expected a special_days_enriched workbook with an "
            "event_day_airport sheet."
        ) from exc

    col = _required_event_day_airport_columns(raw_rows)
    parsed_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        airport = _clean_iata(row.get(col["airport"]))
        event_date = _parse_date(row.get(col["event_date"]))
        if not airport or event_date is None:
            continue
        parsed_rows.append(
            {
                "Airport": airport,
                "Event date": event_date,
                "Event": str(row.get(col["event"], "") or "").strip(),
                "Impact": _parse_float(row.get(col.get("impact", ""))) or 0.0,
                "Source": str(row.get(col.get("source", ""), "") or "").strip(),
                "Predicted attendance": _parse_float(row.get(col.get("predicted_attendance", ""))),
            }
        )
    return headers, parsed_rows


def aggregate_event_features(event_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[date, str], dict[str, Any]] = {}
    for row in event_rows:
        key = (row["Event date"], row["Airport"])
        agg = grouped.get(key)
        if agg is None:
            agg = {
                "event_date": row["Event date"],
                "airport": row["Airport"],
                "event_weight_sum": 0.0,
                "event_weight_max": 0.0,
                "event_count": 0,
                "event_names": set(),
                "sources": set(),
            }
            grouped[key] = agg

        impact = float(row["Impact"] or 0.0)
        agg["event_weight_sum"] += impact
        agg["event_weight_max"] = max(agg["event_weight_max"], impact)
        agg["event_count"] += 1
        if row["Event"]:
            agg["event_names"].add(row["Event"])
        if row["Source"]:
            agg["sources"].add(row["Source"])

    features: list[dict[str, Any]] = []
    for key in sorted(grouped):
        agg = grouped[key]
        features.append(
            {
                "event_date": agg["event_date"],
                "airport": agg["airport"],
                "event_weight_sum": round(agg["event_weight_sum"], 4),
                "event_weight_max": round(agg["event_weight_max"], 4),
                "event_count": agg["event_count"],
                "event_flag": 1 if agg["event_count"] else 0,
                "event_names": " | ".join(sorted(agg["event_names"])),
                "sources": ",".join(sorted(agg["sources"])),
            }
        )
    return features


def join_lf_with_events(lf_rows: list[dict[str, Any]], features: list[dict[str, Any]]) -> list[dict[str, Any]]:
    feature_map = {(row["event_date"], row["airport"]): row for row in features}
    joined: list[dict[str, Any]] = []
    for row in lf_rows:
        feature = feature_map.get((row["flight_date"], row["AIRPORT"]))
        joined_row = dict(row)
        joined_row.update(
            {
                "event_date": feature["event_date"] if feature else None,
                "event_weight_sum": feature["event_weight_sum"] if feature else 0.0,
                "event_weight_max": feature["event_weight_max"] if feature else 0.0,
                "event_count": feature["event_count"] if feature else 0,
                "event_flag": feature["event_flag"] if feature else 0,
                "event_names": feature["event_names"] if feature else "",
                "sources": feature["sources"] if feature else "",
            }
        )
        joined.append(joined_row)
    joined.sort(key=lambda row: (row["flight_date"], row["AIRPORT"]))
    return joined


def _mean(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def _sample_variance(values: list[float]) -> float | None:
    if len(values) < 2:
        return None
    mean_value = sum(values) / len(values)
    return sum((value - mean_value) ** 2 for value in values) / (len(values) - 1)


def _rank(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    index = 0
    while index < len(indexed):
        end = index
        while end + 1 < len(indexed) and indexed[end + 1][1] == indexed[index][1]:
            end += 1
        rank = (index + end + 2) / 2.0
        for offset in range(index, end + 1):
            ranks[indexed[offset][0]] = rank
        index = end + 1
    return ranks


def pearson_correlation(x: list[float], y: list[float]) -> float | None:
    if len(x) < 2 or len(y) < 2 or len(x) != len(y):
        return None
    mean_x = sum(x) / len(x)
    mean_y = sum(y) / len(y)
    centered_x = [value - mean_x for value in x]
    centered_y = [value - mean_y for value in y]
    sum_x = sum(value * value for value in centered_x)
    sum_y = sum(value * value for value in centered_y)
    denom = math.sqrt(sum_x * sum_y)
    if denom == 0:
        return None
    cov = sum(a * b for a, b in zip(centered_x, centered_y))
    return cov / denom


def spearman_correlation(x: list[float], y: list[float]) -> float | None:
    if len(x) < 2 or len(y) < 2 or len(x) != len(y):
        return None
    return pearson_correlation(_rank(x), _rank(y))


def _normal_p_value_from_t(t_stat: float | None) -> float | None:
    if t_stat is None:
        return None
    return math.erfc(abs(t_stat) / math.sqrt(2.0))


def correlation_t_stat(correlation: float | None, n_rows: int) -> tuple[float | None, float | None]:
    if correlation is None or n_rows < 3:
        return None, None
    if abs(correlation) >= 1.0:
        return math.copysign(math.inf, correlation), 0.0
    denom = 1.0 - (correlation * correlation)
    if denom <= 0:
        return None, None
    t_stat = correlation * math.sqrt((n_rows - 2) / denom)
    return t_stat, _normal_p_value_from_t(t_stat)


def welch_t_test(sample_a: list[float], sample_b: list[float]) -> tuple[float | None, float | None]:
    if len(sample_a) < 2 or len(sample_b) < 2:
        return None, None
    mean_a = _mean(sample_a)
    mean_b = _mean(sample_b)
    if mean_a is None or mean_b is None:
        return None, None
    var_a = _sample_variance(sample_a)
    var_b = _sample_variance(sample_b)
    if var_a is None or var_b is None:
        return None, None
    denom = math.sqrt((var_a / len(sample_a)) + (var_b / len(sample_b)))
    if denom == 0:
        return None, None
    t_stat = (mean_a - mean_b) / denom
    return t_stat, _normal_p_value_from_t(t_stat)


def _rounded(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def _airport_summary_row(airport: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    lf_rows = [row for row in rows if row.get("LF_PCT") is not None]
    lf_values = [float(row["LF_PCT"]) for row in lf_rows]
    event_values = [float(row["LF_PCT"]) for row in lf_rows if row["event_flag"]]
    non_event_values = [float(row["LF_PCT"]) for row in lf_rows if not row["event_flag"]]

    weight_pairs = [(float(row["LF_PCT"]), float(row["event_weight_sum"])) for row in lf_rows]
    flag_pairs = [(float(row["LF_PCT"]), float(row["event_flag"])) for row in lf_rows]

    corr_weight_sum_pearson = pearson_correlation(
        [pair[0] for pair in weight_pairs], [pair[1] for pair in weight_pairs]
    )
    corr_weight_sum_spearman = spearman_correlation(
        [pair[0] for pair in weight_pairs], [pair[1] for pair in weight_pairs]
    )
    corr_event_flag_pearson = pearson_correlation(
        [pair[0] for pair in flag_pairs], [pair[1] for pair in flag_pairs]
    )
    corr_event_flag_spearman = spearman_correlation(
        [pair[0] for pair in flag_pairs], [pair[1] for pair in flag_pairs]
    )

    avg_event_lf = _mean(event_values)
    avg_non_event_lf = _mean(non_event_values)
    event_uplift = None
    if avg_event_lf is not None and avg_non_event_lf is not None:
        event_uplift = avg_event_lf - avg_non_event_lf

    uplift_t_stat, uplift_p_value = welch_t_test(event_values, non_event_values)
    corr_t_stat, corr_p_value = correlation_t_stat(corr_weight_sum_pearson, len(weight_pairs))
    event_names = sorted({name for row in rows for name in str(row.get("event_names", "")).split(" | ") if name})

    return {
        "AIRPORT": airport,
        "n_rows": len(rows),
        "event_days": sum(1 for row in rows if row["event_flag"]),
        "avg_lf_pct": _rounded(_mean(lf_values), 4),
        "avg_event_lf_pct": _rounded(avg_event_lf, 4),
        "avg_non_event_lf_pct": _rounded(avg_non_event_lf, 4),
        "event_uplift_pct_pts": _rounded(event_uplift, 4),
        "corr_weight_sum_pearson": _rounded(corr_weight_sum_pearson),
        "corr_weight_sum_spearman": _rounded(corr_weight_sum_spearman),
        "corr_event_flag_pearson": _rounded(corr_event_flag_pearson),
        "corr_event_flag_spearman": _rounded(corr_event_flag_spearman),
        "corr_weight_sum_t_stat": _rounded(corr_t_stat),
        "corr_weight_sum_p_value": _rounded(corr_p_value),
        "uplift_t_stat": _rounded(uplift_t_stat),
        "uplift_p_value": _rounded(uplift_p_value),
        "event_names": " | ".join(event_names),
    }


def summarize_by_airport(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        grouped[row["AIRPORT"]].append(row)
    summary = [_airport_summary_row(airport, grouped[airport]) for airport in sorted(grouped)]
    summary.sort(
        key=lambda row: (
            abs(row["corr_weight_sum_spearman"] or 0.0),
            row["event_days"],
            row["n_rows"],
        ),
        reverse=True,
    )
    return summary


def summarize_overall(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_airport_summary_row("ALL", joined_rows)]


def build_airport_timeline(joined_rows: list[dict[str, Any]], airport: str) -> list[dict[str, Any]]:
    timeline = [row for row in joined_rows if row["AIRPORT"] == airport]
    timeline.sort(key=lambda row: row["flight_date"])
    timeline_rows: list[dict[str, Any]] = []
    for row in timeline:
        flight_date = row["flight_date"]
        focus_period_keys = _row_focus_period_keys(row)
        timeline_rows.append(
            {
                "flight_date": flight_date,
                "AIRPORT": row["AIRPORT"],
                "month": flight_date.month,
                "month_name": _MONTH_NAMES[flight_date.month],
                "LF_PCT": row["LF_PCT"],
                "event_weight_sum": row["event_weight_sum"],
                "event_weight_max": row["event_weight_max"],
                "event_count": row["event_count"],
                "event_flag": row["event_flag"],
                "summer_month_flag": 1 if flight_date.month in _SUMMER_MONTHS else 0,
                "bayram_flag": 1 if "bayram" in focus_period_keys else 0,
                "summer_vacation_flag": 1 if "summer_vacation" in focus_period_keys else 0,
                "school_break_flag": 1 if "school_break" in focus_period_keys else 0,
                "focus_period_keys": ",".join(focus_period_keys),
                "focus_period_labels": _focus_period_labels(focus_period_keys),
                "event_names": row["event_names"],
                "sources": row["sources"],
            }
        )
    return timeline_rows


def resolve_timeline_airports(
    summary_rows: list[dict[str, Any]],
    requested_airport: str | None,
    requested_airports: list[str] | None,
) -> list[str]:
    if requested_airport and requested_airports:
        raise ValueError("Use either airport or airports, not both.")

    summary_airports = [str(row["AIRPORT"]) for row in summary_rows if row["AIRPORT"] != "ALL"]
    summary_airports_set = set(summary_airports)

    if requested_airports:
        missing = [airport for airport in requested_airports if airport not in summary_airports_set]
        if missing:
            raise ValueError(
                "Requested airports are not present in the joined LF data: " + ", ".join(sorted(missing))
            )
        return [airport for airport in summary_airports if airport in set(requested_airports)]

    if requested_airport:
        requested = requested_airport.upper()
        if requested not in summary_airports_set:
            raise ValueError(f"Requested airport is not present in the joined LF data: {requested}")
        return [requested]

    default_airports: list[str] = []
    if "IST" in summary_airports_set:
        default_airports.append("IST")
    default_airports.extend(
        str(row["AIRPORT"])
        for row in summary_rows
        if row["AIRPORT"] not in {"ALL", "IST"} and row["event_days"] > 0
    )
    if not default_airports:
        raise ValueError("No airport has any matched event days after the airport-date join.")
    return default_airports


def build_graph_index_rows(
    summary_rows: list[dict[str, Any]],
    timeline_airports: list[str],
    graph_files: dict[str, Path],
) -> list[dict[str, Any]]:
    summary_by_airport = {str(row["AIRPORT"]): row for row in summary_rows}
    rows: list[dict[str, Any]] = []
    for airport in timeline_airports:
        summary = summary_by_airport[airport]
        rows.append(
            {
                "AIRPORT": airport,
                "timeline_sheet": f"timeline_{airport}",
                "graph_file": graph_files[airport].name,
                "event_days": summary["event_days"],
                "event_uplift_pct_pts": summary["event_uplift_pct_pts"],
                "corr_weight_sum_spearman": summary["corr_weight_sum_spearman"],
                "corr_weight_sum_p_value": summary["corr_weight_sum_p_value"],
            }
        )
    return rows


def build_methodology_rows(summary_rows: list[dict[str, Any]], timeline_airports: list[str]) -> list[dict[str, Any]]:
    non_ist_event_airports = [
        str(row["AIRPORT"])
        for row in summary_rows
        if row["AIRPORT"] not in {"ALL", "IST", "SAW"} and row["event_days"] > 0
    ]
    return [
        {
            "topic": "join_grain",
            "detail": (
                "This workflow only treats an LF row when AIRPORT and flight_date match an airport-day event row "
                "from special_days_enriched.event_day_airport."
            ),
        },
        {
            "topic": "why_non_ist_matters",
            "detail": (
                "Non-IST airports are the cleaner proof slice because a local event in FRA or CGN should align with the "
                "same local airport's LF if the event signal is real. That is harder to fake with generic Turkey-wide seasonality."
            ),
        },
        {
            "topic": "ist_hub_caveat",
            "detail": (
                "IST is a transfer hub with many unrelated demand drivers. A positive IST-only result is weaker evidence than "
                "seeing aligned signal repeated at non-IST destination airports."
            ),
        },
        {
            "topic": "ist_benchmark",
            "detail": (
                "IST is still exported as a benchmark because Bayram and summer patterns should be visible there, but non-IST "
                "airports remain the cleaner proof that local events are affecting LF."
            ),
        },
        {
            "topic": "proof_standard",
            "detail": (
                "Repeated airport-specific uplift and correlation across multiple non-IST airports is stronger evidence that "
                "special events affect LF than a single hub-level result."
            ),
        },
        {
            "topic": "airport_day_caveat",
            "detail": (
                "Airport-day LF is good screening evidence. Route-direction LF would be a stricter causal test because some "
                "event demand can lead or lag around the actual event date."
            ),
        },
        {
            "topic": "timeline_airports",
            "detail": ", ".join(timeline_airports),
        },
        {
            "topic": "non_ist_airports_with_events",
            "detail": ", ".join(non_ist_event_airports) if non_ist_event_airports else "None",
        },
    ]


def _period_summary_row(
    airport: str,
    period_key: str,
    period_rows: list[dict[str, Any]],
    non_period_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    period_values = [float(row["LF_PCT"]) for row in period_rows if row.get("LF_PCT") is not None]
    non_period_values = [float(row["LF_PCT"]) for row in non_period_rows if row.get("LF_PCT") is not None]
    avg_period_lf = _mean(period_values)
    avg_non_period_lf = _mean(non_period_values)
    uplift = None
    if avg_period_lf is not None and avg_non_period_lf is not None:
        uplift = avg_period_lf - avg_non_period_lf
    t_stat, p_value = welch_t_test(period_values, non_period_values)
    event_names = sorted(
        {
            name
            for row in period_rows
            for name in _row_focus_period_names(row, period_key)
        }
    )
    return {
        "AIRPORT": airport,
        "period_key": period_key,
        "period_label": _FOCUS_PERIODS[period_key]["label"],
        "period_basis": _FOCUS_PERIODS[period_key]["basis"],
        "period_days": len(period_rows),
        "avg_period_lf_pct": _rounded(avg_period_lf, 4),
        "avg_non_period_lf_pct": _rounded(avg_non_period_lf, 4),
        "uplift_pct_pts": _rounded(uplift, 4),
        "avg_period_event_weight_sum": _rounded(
            _mean([float(row.get("event_weight_sum") or 0.0) for row in period_rows]),
            4,
        ),
        "t_stat": _rounded(t_stat),
        "p_value": _rounded(p_value),
        "event_names": " | ".join(event_names),
    }


def summarize_focus_periods(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        grouped[row["AIRPORT"]].append(row)

    rows: list[dict[str, Any]] = []
    for airport in sorted(grouped):
        airport_rows = grouped[airport]
        for period_key in _FOCUS_PERIODS:
            period_rows = [row for row in airport_rows if _row_has_focus_period(row, period_key)]
            non_period_rows = [row for row in airport_rows if not _row_has_focus_period(row, period_key)]
            if not period_rows or not non_period_rows:
                continue
            rows.append(_period_summary_row(airport, period_key, period_rows, non_period_rows))
    return rows


def _window_row(
    airport: str,
    period_key: str,
    window_rows: list[dict[str, Any]],
    non_period_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    period_values = [float(row["LF_PCT"]) for row in window_rows if row.get("LF_PCT") is not None]
    avg_period_lf = _mean(period_values)
    avg_non_period_lf = _mean([float(row["LF_PCT"]) for row in non_period_rows if row.get("LF_PCT") is not None])
    uplift = None
    if avg_period_lf is not None and avg_non_period_lf is not None:
        uplift = avg_period_lf - avg_non_period_lf
    event_names = sorted(
        {
            name
            for row in window_rows
            for name in _row_focus_period_names(row, period_key)
        }
    )
    return {
        "AIRPORT": airport,
        "period_key": period_key,
        "period_label": _FOCUS_PERIODS[period_key]["label"],
        "period_basis": _FOCUS_PERIODS[period_key]["basis"],
        "window_start": window_rows[0]["flight_date"],
        "window_end": window_rows[-1]["flight_date"],
        "window_days": len(window_rows),
        "avg_window_lf_pct": _rounded(avg_period_lf, 4),
        "avg_non_period_lf_pct": _rounded(avg_non_period_lf, 4),
        "uplift_pct_pts": _rounded(uplift, 4),
        "avg_window_event_weight_sum": _rounded(
            _mean([float(row.get("event_weight_sum") or 0.0) for row in window_rows]),
            4,
        ),
        "event_names": " | ".join(event_names),
    }


def build_focus_period_windows(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        grouped[row["AIRPORT"]].append(row)

    windows: list[dict[str, Any]] = []
    for airport in sorted(grouped):
        airport_rows = sorted(grouped[airport], key=lambda row: row["flight_date"])
        for period_key in _FOCUS_PERIODS:
            non_period_rows = [row for row in airport_rows if not _row_has_focus_period(row, period_key)]
            if not non_period_rows:
                continue
            current_window: list[dict[str, Any]] = []
            previous_date: date | None = None
            for row in airport_rows:
                if _row_has_focus_period(row, period_key):
                    if current_window and previous_date is not None and row["flight_date"] != previous_date + _ONE_DAY:
                        windows.append(_window_row(airport, period_key, current_window, non_period_rows))
                        current_window = []
                    current_window.append(row)
                    previous_date = row["flight_date"]
                else:
                    if current_window:
                        windows.append(_window_row(airport, period_key, current_window, non_period_rows))
                        current_window = []
                    previous_date = None
            if current_window:
                windows.append(_window_row(airport, period_key, current_window, non_period_rows))
    return windows


def summarize_monthly_seasonality(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        grouped[(row["AIRPORT"], row["flight_date"].month)].append(row)

    rows: list[dict[str, Any]] = []
    for airport, month in sorted(grouped):
        month_rows = grouped[(airport, month)]
        rows.append(
            {
                "AIRPORT": airport,
                "month": month,
                "month_name": _MONTH_NAMES[month],
                "summer_month_flag": 1 if month in _SUMMER_MONTHS else 0,
                "n_rows": len(month_rows),
                "event_days": sum(1 for row in month_rows if row.get("event_flag")),
                "bayram_days": sum(1 for row in month_rows if _row_has_focus_period(row, "bayram")),
                "summer_vacation_days": sum(1 for row in month_rows if _row_has_focus_period(row, "summer_vacation")),
                "school_break_days": sum(1 for row in month_rows if _row_has_focus_period(row, "school_break")),
                "avg_lf_pct": _rounded(_mean([float(row["LF_PCT"]) for row in month_rows if row.get("LF_PCT") is not None]), 4),
                "avg_event_weight_sum": _rounded(
                    _mean([float(row.get("event_weight_sum") or 0.0) for row in month_rows]),
                    4,
                ),
            }
        )
    return rows


def _sheet_headers(rows: list[dict[str, Any]], preferred: list[str]) -> list[str]:
    if rows:
        extras = [key for key in rows[0].keys() if key not in preferred]
    else:
        extras = []
    return preferred + extras


def _write_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, Any]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    for index, (name, headers, rows) in enumerate(sheets):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = name[:31]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
        _format_sheet_dates(sheet)
        _autosize(sheet)
    workbook.save(path)


def _format_sheet_dates(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = _DATETIME_FORMAT
            elif isinstance(cell.value, date):
                cell.number_format = _DATE_FORMAT


def _autosize(sheet) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            if value is None:
                continue
            widths[index] = min(max(widths.get(index, 0), len(str(value)) + 2), 60)
    for index, width in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def _svg_y(value: float, chart_top: int, chart_height: int, y_max: float) -> float:
    return chart_top + chart_height - ((value / y_max) * chart_height)


def _svg_x(index: int, total: int, chart_left: int, chart_width: int) -> float:
    if total <= 1:
        return chart_left + (chart_width / 2.0)
    return chart_left + (index * (chart_width / float(total - 1)))


def _format_svg_points(
    values: list[tuple[float, float]],
) -> str:
    return " ".join(f"{x:.2f},{y:.2f}" for x, y in values)


def _write_airport_svg(path: Path, airport: str, timeline_rows: list[dict[str, Any]]) -> None:
    width = 1280
    height = 720
    chart_left = 80
    chart_right = 40
    chart_top = 60
    chart_bottom = 120
    chart_width = width - chart_left - chart_right
    chart_height = height - chart_top - chart_bottom

    max_value = 100.0
    for row in timeline_rows:
        for key in ("LF_PCT", "event_weight_sum", "event_weight_max"):
            value = row.get(key)
            if value is not None:
                max_value = max(max_value, float(value))
    y_max = float(int(math.ceil(max_value / 20.0) * 20.0))

    lf_points: list[tuple[float, float]] = []
    event_points: list[tuple[float, float]] = []
    event_markers: list[str] = []
    axis_labels: list[str] = []
    grid_lines: list[str] = []
    period_overlays: list[str] = []

    for grid_value in range(0, int(y_max) + 1, 20):
        y = _svg_y(float(grid_value), chart_top, chart_height, y_max)
        grid_lines.append(
            f'<line x1="{chart_left}" y1="{y:.2f}" x2="{chart_left + chart_width}" y2="{y:.2f}" '
            'stroke="#d9dde3" stroke-width="1" />'
        )
        axis_labels.append(
            f'<text x="{chart_left - 12}" y="{y + 5:.2f}" text-anchor="end" '
            'font-family="Arial" font-size="12" fill="#3a4654">'
            f'{grid_value}</text>'
        )

    tick_step = max(1, len(timeline_rows) // 10) if timeline_rows else 1
    x_tick_labels: list[str] = []
    overlay_step = chart_width / float(max(len(timeline_rows) - 1, 1))

    for period_key in ("bayram", "summer_vacation", "summer_months"):
        start_index: int | None = None
        for index, row in enumerate(timeline_rows):
            in_period = (
                row["summer_month_flag"] == 1 if period_key == "summer_months" else row[f"{period_key}_flag"] == 1
            )
            if in_period and start_index is None:
                start_index = index
            if (not in_period or index == len(timeline_rows) - 1) and start_index is not None:
                end_index = index if in_period and index == len(timeline_rows) - 1 else index - 1
                x_start = _svg_x(start_index, len(timeline_rows), chart_left, chart_width) - (overlay_step / 2.0)
                x_end = _svg_x(end_index, len(timeline_rows), chart_left, chart_width) + (overlay_step / 2.0)
                rect_x = max(chart_left, x_start)
                rect_width = min(chart_left + chart_width, x_end) - rect_x
                if rect_width > 0:
                    period_info = _FOCUS_PERIODS[period_key]
                    label = period_info["label"]
                    start_date = timeline_rows[start_index]["flight_date"].isoformat()
                    end_date = timeline_rows[end_index]["flight_date"].isoformat()
                    period_overlays.append(
                        f'<rect x="{rect_x:.2f}" y="{chart_top}" width="{rect_width:.2f}" height="{chart_height}" '
                        f'fill="{period_info["color"]}" opacity="0.12">'
                        f'<title>{html.escape(label)} | {start_date} to {end_date}</title></rect>'
                    )
                start_index = None

    for index, row in enumerate(timeline_rows):
        x = _svg_x(index, len(timeline_rows), chart_left, chart_width)
        lf_value = float(row["LF_PCT"] or 0.0)
        event_value = float(row["event_weight_sum"] or 0.0)
        lf_points.append((x, _svg_y(lf_value, chart_top, chart_height, y_max)))
        event_points.append((x, _svg_y(event_value, chart_top, chart_height, y_max)))

        if row["event_flag"]:
            tooltip = html.escape(
                f"{row['flight_date'].isoformat()} | {airport} | LF={lf_value:.2f} | EventWeight={event_value:.2f} | {row['event_names']}"
            )
            marker_y = _svg_y(event_value, chart_top, chart_height, y_max)
            event_markers.append(
                f'<line x1="{x:.2f}" y1="{chart_top + chart_height:.2f}" x2="{x:.2f}" y2="{marker_y:.2f}" '
                'stroke="#e68619" stroke-width="2" opacity="0.45" />'
                f'<circle cx="{x:.2f}" cy="{marker_y:.2f}" r="5" fill="#e68619">'
                f'<title>{tooltip}</title></circle>'
            )

        if index % tick_step == 0 or index == len(timeline_rows) - 1:
            label = row["flight_date"].isoformat()
            x_tick_labels.append(
                f'<text x="{x:.2f}" y="{chart_top + chart_height + 25:.2f}" transform="rotate(45 {x:.2f},{chart_top + chart_height + 25:.2f})" '
                'text-anchor="start" font-family="Arial" font-size="11" fill="#3a4654">'
                f'{html.escape(label)}</text>'
            )

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
<rect x="0" y="0" width="{width}" height="{height}" fill="#f6f8fb" />
<text x="{chart_left}" y="34" font-family="Arial" font-size="24" font-weight="700" fill="#1f2933">Airport LF vs special events: {html.escape(airport)}</text>
<text x="{chart_left}" y="54" font-family="Arial" font-size="13" fill="#52606d">Airport-specific join only: AIRPORT + flight_date</text>
{''.join(period_overlays)}
{''.join(grid_lines)}
<line x1="{chart_left}" y1="{chart_top}" x2="{chart_left}" y2="{chart_top + chart_height}" stroke="#52606d" stroke-width="1.5" />
<line x1="{chart_left}" y1="{chart_top + chart_height}" x2="{chart_left + chart_width}" y2="{chart_top + chart_height}" stroke="#52606d" stroke-width="1.5" />
<polyline fill="none" stroke="#2563eb" stroke-width="3" points="{_format_svg_points(lf_points)}" />
<polyline fill="none" stroke="#e68619" stroke-width="3" stroke-dasharray="8 5" points="{_format_svg_points(event_points)}" />
{''.join(event_markers)}
{''.join(axis_labels)}
{''.join(x_tick_labels)}
<rect x="{chart_left}" y="{height - 48}" width="16" height="4" fill="#2563eb" />
<text x="{chart_left + 24}" y="{height - 42}" font-family="Arial" font-size="13" fill="#1f2933">LF_PCT</text>
<rect x="{chart_left + 120}" y="{height - 48}" width="16" height="4" fill="#e68619" />
<text x="{chart_left + 144}" y="{height - 42}" font-family="Arial" font-size="13" fill="#1f2933">event_weight_sum</text>
<rect x="{chart_left + 310}" y="{height - 48}" width="16" height="10" fill="#34a853" opacity="0.20" />
<text x="{chart_left + 334}" y="{height - 40}" font-family="Arial" font-size="13" fill="#1f2933">Bayram</text>
<rect x="{chart_left + 420}" y="{height - 48}" width="16" height="10" fill="#f59e0b" opacity="0.20" />
<text x="{chart_left + 444}" y="{height - 40}" font-family="Arial" font-size="13" fill="#1f2933">Summer vacation</text>
<rect x="{chart_left + 610}" y="{height - 48}" width="16" height="10" fill="#fcd34d" opacity="0.20" />
<text x="{chart_left + 634}" y="{height - 40}" font-family="Arial" font-size="13" fill="#1f2933">Summer months</text>
<text x="{width - 300}" y="{height - 42}" font-family="Arial" font-size="12" fill="#52606d">Orange markers show event dates and carry event-name tooltips.</text>
</svg>
'''
    path.write_text(svg, encoding="utf-8")


def _write_dashboard_html(
    path: Path,
    summary_rows: list[dict[str, Any]],
    timeline_airports: list[str],
    graph_files: dict[str, Path],
    methodology_rows: list[dict[str, Any]],
    focus_period_summary_rows: list[dict[str, Any]],
) -> None:
    summary_by_airport = {str(row["AIRPORT"]): row for row in summary_rows}
    table_rows = "".join(
        "<tr>"
        f"<td>{html.escape(airport)}</td>"
        f"<td>{summary_by_airport[airport]['event_days']}</td>"
        f"<td>{summary_by_airport[airport]['avg_event_lf_pct']}</td>"
        f"<td>{summary_by_airport[airport]['avg_non_event_lf_pct']}</td>"
        f"<td>{summary_by_airport[airport]['event_uplift_pct_pts']}</td>"
        f"<td>{summary_by_airport[airport]['corr_weight_sum_spearman']}</td>"
        f"<td>{summary_by_airport[airport]['corr_weight_sum_p_value']}</td>"
        "</tr>"
        for airport in timeline_airports
    )
    methodology_html = "".join(
        f"<li><strong>{html.escape(str(row['topic']))}</strong>: {html.escape(str(row['detail']))}</li>"
        for row in methodology_rows
    )
    graph_sections = "".join(
        f"<section><h2>{html.escape(airport)}</h2><img src=\"graphs/{html.escape(graph_files[airport].name)}\" alt=\"{html.escape(airport)} graph\" /></section>"
        for airport in timeline_airports
    )
    focus_summary_by_airport: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in focus_period_summary_rows:
        focus_summary_by_airport[str(row["AIRPORT"])].append(row)
    focus_sections = "".join(
        "<section>"
        f"<h2>{html.escape(airport)} period effects</h2>"
        "<table><thead><tr><th>period_label</th><th>period_basis</th><th>period_days</th><th>avg_period_lf_pct</th><th>avg_non_period_lf_pct</th><th>uplift_pct_pts</th><th>p_value</th></tr></thead><tbody>"
        + "".join(
            "<tr>"
            f"<td>{html.escape(str(row['period_label']))}</td>"
            f"<td>{html.escape(str(row['period_basis']))}</td>"
            f"<td>{row['period_days']}</td>"
            f"<td>{row['avg_period_lf_pct']}</td>"
            f"<td>{row['avg_non_period_lf_pct']}</td>"
            f"<td>{row['uplift_pct_pts']}</td>"
            f"<td>{row['p_value']}</td>"
            "</tr>"
            for row in focus_summary_by_airport.get(airport, [])
        )
        + "</tbody></table></section>"
        for airport in timeline_airports
    )
    html_text = f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Airport LF backtest dashboard</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; background: #f6f8fb; color: #1f2933; }}
    h1, h2 {{ margin-bottom: 8px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0 28px; background: white; }}
    th, td {{ border: 1px solid #d9dde3; padding: 8px 10px; text-align: left; font-size: 14px; }}
    th {{ background: #eaf1ff; }}
    ul {{ background: white; padding: 18px 24px; border: 1px solid #d9dde3; }}
    section {{ margin: 28px 0; background: white; padding: 16px; border: 1px solid #d9dde3; }}
    img {{ max-width: 100%; height: auto; border: 1px solid #d9dde3; background: white; }}
  </style>
</head>
<body>
  <h1>Airport-specific LF vs special-events dashboard</h1>
  <p>This dashboard uses only AIRPORT + flight_date joins from the enriched airport-day event sheet.</p>
  <h2>Why non-IST airports matter</h2>
  <ul>{methodology_html}</ul>
  <h2>Selected airports</h2>
  <table>
    <thead>
      <tr>
        <th>AIRPORT</th>
        <th>event_days</th>
        <th>avg_event_lf_pct</th>
        <th>avg_non_event_lf_pct</th>
        <th>event_uplift_pct_pts</th>
        <th>corr_weight_sum_spearman</th>
        <th>corr_weight_sum_p_value</th>
      </tr>
    </thead>
    <tbody>{table_rows}</tbody>
  </table>
    {focus_sections}
  {graph_sections}
</body>
</html>
'''
    path.write_text(html_text, encoding="utf-8")


def _add_airport_timeline_chart(path: Path, sheet_name: str, airport: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference

    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        if sheet.max_row <= 1:
            workbook.save(path)
            return

        headers = {cell.value: index + 1 for index, cell in enumerate(sheet[1])}
        date_col = headers.get("flight_date")
        if date_col is None:
            workbook.save(path)
            return

        chart = LineChart()
        chart.title = f"Airport LF vs special-event weights: {airport}"
        chart.y_axis.title = "LF % / event weight"
        chart.x_axis.title = "Date"
        chart.height = 10
        chart.width = 22
        chart.legend.position = "r"

        for header in ("LF_PCT", "event_weight_sum", "event_weight_max"):
            column = headers.get(header)
            if column is None:
                continue
            values = Reference(sheet, min_col=column, min_row=1, max_row=sheet.max_row)
            chart.add_data(values, titles_from_data=True)

        categories = Reference(sheet, min_col=date_col, min_row=2, max_row=sheet.max_row)
        chart.set_categories(categories)
        sheet.add_chart(chart, "K2")
        workbook.save(path)
    finally:
        workbook.close()


def run_workflow(
    lf_path: str | Path,
    events_enriched_path: str | Path,
    out_dir: str | Path,
    *,
    lf_sheet: str | None = None,
    events_sheet: str = "event_day_airport",
    start_date: date | None = None,
    end_date: date | None = None,
    airport: str | None = None,
    airports: list[str] | None = None,
) -> dict[str, Any]:
    _, lf_rows = load_lf_rows(lf_path, lf_sheet)
    if not lf_rows:
        raise ValueError("LF workbook has no valid LF rows after parsing.")

    inferred_start = min(row["flight_date"] for row in lf_rows)
    inferred_end = max(row["flight_date"] for row in lf_rows)
    analysis_start = start_date or inferred_start
    analysis_end = end_date or inferred_end
    if analysis_end < analysis_start:
        raise ValueError("end_date must be on or after start_date")

    lf_rows = [row for row in lf_rows if analysis_start <= row["flight_date"] <= analysis_end]
    _, event_rows = load_enriched_event_rows(events_enriched_path, events_sheet)
    event_rows = [row for row in event_rows if analysis_start <= row["Event date"] <= analysis_end]

    features = aggregate_event_features(event_rows)
    joined_rows = join_lf_with_events(lf_rows, features)
    airport_summary = summarize_by_airport(joined_rows)
    overall_summary = summarize_overall(joined_rows)
    focus_period_summary_rows = summarize_focus_periods(joined_rows)
    focus_period_window_rows = build_focus_period_windows(joined_rows)
    monthly_seasonality_rows = summarize_monthly_seasonality(joined_rows)
    timeline_airports = resolve_timeline_airports(airport_summary, airport, airports)
    methodology_rows = build_methodology_rows(airport_summary, timeline_airports)

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    graph_dir = out_path / "graphs"
    graph_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"{analysis_start.isoformat()}_{analysis_end.isoformat()}"
    workbook_path = out_path / f"airport_lf_backtest_simple_{suffix}.xlsx"
    dashboard_path = out_path / f"airport_lf_backtest_simple_{suffix}.html"

    timeline_sheets: list[tuple[str, list[str], list[dict[str, Any]]]] = []
    graph_files: dict[str, Path] = {}
    for timeline_airport in timeline_airports:
        timeline_rows = build_airport_timeline(joined_rows, timeline_airport)
        timeline_sheets.append(
            (
                f"timeline_{timeline_airport}",
                [
                    "flight_date",
                    "AIRPORT",
                    "month",
                    "month_name",
                    "LF_PCT",
                    "event_weight_sum",
                    "event_weight_max",
                    "event_count",
                    "event_flag",
                    "summer_month_flag",
                    "bayram_flag",
                    "summer_vacation_flag",
                    "school_break_flag",
                    "focus_period_keys",
                    "focus_period_labels",
                    "event_names",
                    "sources",
                ],
                timeline_rows,
            )
        )
        graph_path = graph_dir / f"{timeline_airport}_lf_vs_events.svg"
        _write_airport_svg(graph_path, timeline_airport, timeline_rows)
        graph_files[timeline_airport] = graph_path

    graph_index_rows = build_graph_index_rows(airport_summary, timeline_airports, graph_files)

    _write_workbook(
        workbook_path,
        [
            (
                "methodology_notes",
                ["topic", "detail"],
                methodology_rows,
            ),
            (
                "overall_correlation",
                [
                    "AIRPORT",
                    "n_rows",
                    "event_days",
                    "avg_lf_pct",
                    "avg_event_lf_pct",
                    "avg_non_event_lf_pct",
                    "event_uplift_pct_pts",
                    "corr_weight_sum_pearson",
                    "corr_weight_sum_spearman",
                    "corr_event_flag_pearson",
                    "corr_event_flag_spearman",
                    "corr_weight_sum_t_stat",
                    "corr_weight_sum_p_value",
                    "uplift_t_stat",
                    "uplift_p_value",
                    "event_names",
                ],
                overall_summary,
            ),
            (
                "airport_summary",
                [
                    "AIRPORT",
                    "n_rows",
                    "event_days",
                    "avg_lf_pct",
                    "avg_event_lf_pct",
                    "avg_non_event_lf_pct",
                    "event_uplift_pct_pts",
                    "corr_weight_sum_pearson",
                    "corr_weight_sum_spearman",
                    "corr_event_flag_pearson",
                    "corr_event_flag_spearman",
                    "corr_weight_sum_t_stat",
                    "corr_weight_sum_p_value",
                    "uplift_t_stat",
                    "uplift_p_value",
                    "event_names",
                ],
                airport_summary,
            ),
            (
                "graph_index",
                [
                    "AIRPORT",
                    "timeline_sheet",
                    "graph_file",
                    "event_days",
                    "event_uplift_pct_pts",
                    "corr_weight_sum_spearman",
                    "corr_weight_sum_p_value",
                ],
                graph_index_rows,
            ),
            (
                "focus_period_summary",
                [
                    "AIRPORT",
                    "period_key",
                    "period_label",
                    "period_basis",
                    "period_days",
                    "avg_period_lf_pct",
                    "avg_non_period_lf_pct",
                    "uplift_pct_pts",
                    "avg_period_event_weight_sum",
                    "t_stat",
                    "p_value",
                    "event_names",
                ],
                focus_period_summary_rows,
            ),
            (
                "focus_period_windows",
                [
                    "AIRPORT",
                    "period_key",
                    "period_label",
                    "period_basis",
                    "window_start",
                    "window_end",
                    "window_days",
                    "avg_window_lf_pct",
                    "avg_non_period_lf_pct",
                    "uplift_pct_pts",
                    "avg_window_event_weight_sum",
                    "event_names",
                ],
                focus_period_window_rows,
            ),
            (
                "monthly_seasonality",
                [
                    "AIRPORT",
                    "month",
                    "month_name",
                    "summer_month_flag",
                    "n_rows",
                    "event_days",
                    "bayram_days",
                    "summer_vacation_days",
                    "school_break_days",
                    "avg_lf_pct",
                    "avg_event_weight_sum",
                ],
                monthly_seasonality_rows,
            ),
            (
                "joined_airport_day",
                _sheet_headers(
                    joined_rows,
                    LF_HEADERS
                    + [
                        "event_date",
                        "event_weight_sum",
                        "event_weight_max",
                        "event_count",
                        "event_flag",
                        "event_names",
                        "sources",
                    ],
                ),
                joined_rows,
            ),
        ]
        + timeline_sheets,
    )

    for timeline_airport in timeline_airports:
        _add_airport_timeline_chart(workbook_path, f"timeline_{timeline_airport}", timeline_airport)
    _write_dashboard_html(
        dashboard_path,
        airport_summary,
        timeline_airports,
        graph_files,
        methodology_rows,
        focus_period_summary_rows,
    )
    return {
        "workbook": workbook_path,
        "dashboard_html": dashboard_path,
        "graph_files": graph_files,
        "timeline_airports": timeline_airports,
        "analysis_start": analysis_start,
        "analysis_end": analysis_end,
    }


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    parsed_airports = _parse_airports(args.airports)
    outputs = run_workflow(
        lf_path=args.lf,
        events_enriched_path=args.events_enriched,
        out_dir=args.out_dir,
        lf_sheet=args.lf_sheet,
        events_sheet=args.events_sheet,
        start_date=args.start_date,
        end_date=args.end_date,
        airport=args.airport,
        airports=parsed_airports,
    )
    print(f"Wrote workbook: {outputs['workbook']}")
    print(f"Wrote dashboard: {outputs['dashboard_html']}")
    print(f"Timeline airports: {', '.join(outputs['timeline_airports'])}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())