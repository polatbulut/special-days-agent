import os
import tempfile
import unittest
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from special_days import airport_lf_backtest_simple


LF_HEADERS = [
    "YMD",
    "AIRPORT",
    "BOARDED_PAX",
    "CAPACITY",
    "LF",
    "flight_date",
    "LF_PCT",
]

EVENT_DAY_HEADERS = [
    "row_id",
    "Event",
    "Source",
    "City",
    "Airport",
    "Event date",
    "Impact",
    "Predicted attendance",
    "Start date",
    "End date",
    "Bridge start",
    "Bridge end",
    "Effective start",
    "Effective end",
    "Mapping method",
    "Curve source",
]


def _write_single_sheet_workbook(path, headers, rows, title="Sheet1"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _write_multi_sheet_workbook(path, sheets):
    workbook = Workbook()
    first = True
    for title, headers, rows in sheets:
        sheet = workbook.active if first else workbook.create_sheet()
        first = False
        sheet.title = title
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


class AirportLfBacktestSimpleTest(unittest.TestCase):
    def test_workflow_shows_ist_bayram_and_summer_effects_alongside_non_ist_airports(self):
        with tempfile.TemporaryDirectory() as tmp:
            lf_path = os.path.join(tmp, "lf.xlsx")
            events_path = os.path.join(tmp, "special_days_enriched.xlsx")

            _write_single_sheet_workbook(
                lf_path,
                LF_HEADERS,
                [
                    [20250101, "FRA", 70, 100, 0.70, datetime(2025, 1, 1), 70.0],
                    [20250102, "FRA", 72, 100, 0.72, datetime(2025, 1, 2), 72.0],
                    [20250103, "FRA", 95, 100, 0.95, datetime(2025, 1, 3), 95.0],
                    [20250104, "FRA", 96, 100, 0.96, datetime(2025, 1, 4), 96.0],
                    [20250105, "FRA", 71, 100, 0.71, datetime(2025, 1, 5), 71.0],
                    [20250106, "FRA", 73, 100, 0.73, datetime(2025, 1, 6), 73.0],
                    [20250101, "CGN", 80, 100, 0.80, datetime(2025, 1, 1), 80.0],
                    [20250102, "CGN", 79, 100, 0.79, datetime(2025, 1, 2), 79.0],
                    [20250103, "CGN", 60, 100, 0.60, datetime(2025, 1, 3), 60.0],
                    [20250104, "CGN", 62, 100, 0.62, datetime(2025, 1, 4), 62.0],
                    [20250105, "CGN", 81, 100, 0.81, datetime(2025, 1, 5), 81.0],
                    [20250106, "CGN", 78, 100, 0.78, datetime(2025, 1, 6), 78.0],
                    [20250201, "IST", 70, 100, 0.70, datetime(2025, 2, 1), 70.0],
                    [20250202, "IST", 72, 100, 0.72, datetime(2025, 2, 2), 72.0],
                    [20250329, "IST", 92, 100, 0.92, datetime(2025, 3, 29), 92.0],
                    [20250330, "IST", 94, 100, 0.94, datetime(2025, 3, 30), 94.0],
                    [20250715, "IST", 90, 100, 0.90, datetime(2025, 7, 15), 90.0],
                    [20250716, "IST", 91, 100, 0.91, datetime(2025, 7, 16), 91.0],
                    [20251001, "IST", 74, 100, 0.74, datetime(2025, 10, 1), 74.0],
                    [20251002, "IST", 73, 100, 0.73, datetime(2025, 10, 2), 73.0],
                ],
                title="lf_airport_day",
            )

            _write_multi_sheet_workbook(
                events_path,
                [
                    (
                        "event_day_airport",
                        EVENT_DAY_HEADERS,
                        [
                            [1, "Frankfurt Expo", "eventseye", "Frankfurt", "FRA", date(2025, 1, 3), 60, 1000, date(2025, 1, 3), date(2025, 1, 3), "", "", date(2025, 1, 3), date(2025, 1, 3), "nearest_airport", "row_impact"],
                            [2, "Frankfurt Expo", "eventseye", "Frankfurt", "FRA", date(2025, 1, 4), 80, 1000, date(2025, 1, 4), date(2025, 1, 4), "", "", date(2025, 1, 4), date(2025, 1, 4), "nearest_airport", "row_impact"],
                            [3, "Cologne Fair", "eventseye", "Cologne", "CGN", date(2025, 1, 3), 70, 900, date(2025, 1, 3), date(2025, 1, 3), "", "", date(2025, 1, 3), date(2025, 1, 3), "nearest_airport", "row_impact"],
                            [4, "Cologne Fair", "eventseye", "Cologne", "CGN", date(2025, 1, 4), 75, 900, date(2025, 1, 4), date(2025, 1, 4), "", "", date(2025, 1, 4), date(2025, 1, 4), "nearest_airport", "row_impact"],
                            [5, "Ramazan Bayramı 1. Gün", "nager", "Nationwide (TR)", "IST", date(2025, 3, 29), 95, "", date(2025, 3, 29), date(2025, 3, 29), "", "", date(2025, 3, 29), date(2025, 3, 29), "nearest_airport", "row_impact"],
                            [6, "Ramazan Bayramı 2. Gün", "nager", "Nationwide (TR)", "IST", date(2025, 3, 30), 95, "", date(2025, 3, 30), date(2025, 3, 30), "", "", date(2025, 3, 30), date(2025, 3, 30), "nearest_airport", "row_impact"],
                            [7, "Yaz tatili", "meb", "Nationwide (TR)", "IST", date(2025, 7, 15), 65, "", date(2025, 7, 15), date(2025, 7, 15), "", "", date(2025, 7, 15), date(2025, 7, 15), "nearest_airport", "row_impact"],
                            [8, "Yaz tatili", "meb", "Nationwide (TR)", "IST", date(2025, 7, 16), 65, "", date(2025, 7, 16), date(2025, 7, 16), "", "", date(2025, 7, 16), date(2025, 7, 16), "nearest_airport", "row_impact"],
                        ],
                    ),
                ],
            )

            outputs = airport_lf_backtest_simple.run_workflow(
                lf_path,
                events_path,
                tmp,
            )

            workbook = load_workbook(outputs["workbook"])
            try:
                summary_rows = list(workbook["airport_summary"].iter_rows(values_only=True))
                header = summary_rows[0]
                airport_index = header.index("AIRPORT")
                uplift_index = header.index("event_uplift_pct_pts")
                spearman_index = header.index("corr_weight_sum_spearman")
                event_days_index = header.index("event_days")

                summary_by_airport = {row[airport_index]: row for row in summary_rows[1:]}
                self.assertEqual(summary_by_airport["FRA"][event_days_index], 2)
                self.assertGreater(summary_by_airport["FRA"][uplift_index], 0)
                self.assertGreater(summary_by_airport["FRA"][spearman_index], 0)
                self.assertLess(summary_by_airport["CGN"][uplift_index], 0)
                self.assertGreater(summary_by_airport["IST"][uplift_index], 0)

                methodology_rows = list(workbook["methodology_notes"].iter_rows(values_only=True))
                self.assertIn("why_non_ist_matters", [row[0] for row in methodology_rows[1:]])
                self.assertIn("ist_benchmark", [row[0] for row in methodology_rows[1:]])

                focus_rows = list(workbook["focus_period_summary"].iter_rows(values_only=True))
                focus_header = focus_rows[0]
                focus_airport_index = focus_header.index("AIRPORT")
                focus_period_key_index = focus_header.index("period_key")
                focus_uplift_index = focus_header.index("uplift_pct_pts")
                focus_lookup = {
                    (row[focus_airport_index], row[focus_period_key_index]): row for row in focus_rows[1:]
                }
                self.assertIn(("IST", "bayram"), focus_lookup)
                self.assertIn(("IST", "summer_months"), focus_lookup)
                self.assertGreater(focus_lookup[("IST", "bayram")][focus_uplift_index], 0)
                self.assertGreater(focus_lookup[("IST", "summer_months")][focus_uplift_index], 0)

                seasonality_rows = list(workbook["monthly_seasonality"].iter_rows(values_only=True))
                seasonality_header = seasonality_rows[0]
                seasonality_airport_index = seasonality_header.index("AIRPORT")
                seasonality_month_index = seasonality_header.index("month")
                seasonality_summer_flag_index = seasonality_header.index("summer_month_flag")
                seasonality_avg_index = seasonality_header.index("avg_lf_pct")
                seasonality_lookup = {
                    (row[seasonality_airport_index], row[seasonality_month_index]): row for row in seasonality_rows[1:]
                }
                self.assertEqual(seasonality_lookup[("IST", 7)][seasonality_summer_flag_index], 1)
                self.assertGreater(seasonality_lookup[("IST", 7)][seasonality_avg_index], seasonality_lookup[("IST", 2)][seasonality_avg_index])

                timeline_ist = workbook["timeline_IST"]
                timeline_fra = workbook["timeline_FRA"]
                timeline_cgn = workbook["timeline_CGN"]
                self.assertGreater(len(timeline_ist._charts), 0)
                self.assertGreater(len(timeline_fra._charts), 0)
                self.assertGreater(len(timeline_cgn._charts), 0)
                timeline_rows = list(timeline_ist.iter_rows(values_only=True))
                timeline_header = timeline_rows[0]
                timeline_period_labels_index = timeline_header.index("focus_period_labels")
                timeline_summer_index = timeline_header.index("summer_month_flag")
                self.assertEqual(timeline_rows[1][1], "IST")
                self.assertIn("Bayram", " ".join(str(row[timeline_period_labels_index] or "") for row in timeline_rows[1:]))
                self.assertIn(1, [row[timeline_summer_index] for row in timeline_rows[1:]])
            finally:
                workbook.close()

            self.assertEqual(outputs["timeline_airports"][0], "IST")
            self.assertIn("FRA", outputs["timeline_airports"])
            self.assertIn("CGN", outputs["timeline_airports"])
            self.assertTrue(os.path.exists(outputs["dashboard_html"]))
            self.assertTrue(os.path.exists(outputs["graph_files"]["IST"]))
            self.assertTrue(os.path.exists(outputs["graph_files"]["FRA"]))
            self.assertTrue(os.path.exists(outputs["graph_files"]["CGN"]))

    def test_workflow_requires_enrichment_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            lf_path = os.path.join(tmp, "lf.xlsx")
            events_path = os.path.join(tmp, "not_enriched.xlsx")

            _write_single_sheet_workbook(
                lf_path,
                LF_HEADERS,
                [[20250101, "FRA", 70, 100, 0.70, datetime(2025, 1, 1), 70.0]],
            )
            _write_single_sheet_workbook(
                events_path,
                ["Event", "Airport"],
                [["Anything", "FRA"]],
                title="raw_filtered",
            )

            with self.assertRaisesRegex(ValueError, "Enrichment is not enabled"):
                airport_lf_backtest_simple.run_workflow(lf_path, events_path, tmp)


if __name__ == "__main__":
    unittest.main()