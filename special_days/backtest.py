"""Excel-based LF backtest workflow for historical special events.

This module reads two workbooks:

* an LF workbook at airport-day grain
* a special-days workbook at event-span grain

It produces two new workbooks:

* ``special_days_enriched_<start>_<end>.xlsx``
* ``lf_vs_special_events_<start>_<end>.xlsx``

The workflow is intentionally conservative:

* only exact nearest-airport matches are trusted directly
* ``Nationwide (TR)`` rows expand to Turkish airports present in LF
* blank-airport rows map by exact normalized city only
* anything else goes to ``unresolved_events`` for manual review
"""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from .dataset import load_airports

DEFAULT_START_DATE = date(2025, 1, 1)
DEFAULT_END_DATE = date(2026, 12, 31)

EVENT_HEADERS = [
    "Event",
    "Start date",
    "End date",
    "City",
    "Source",
    "Nearest airport",
    "Impact",
    "Predicted attendance",
    "Bridge start",
    "Bridge end",
    "Impact by day",
    "Impact by day (bridge)",
]

LF_HEADERS = [
    "YMD",
    "AIRPORT",
    "BOARDED_PAX",
    "CAPACITY",
    "LF",
    "flight_date",
    "LF_PCT",
]

_DATE_FORMAT = "yyyy-mm-dd"
_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
_ONE_DAY = timedelta(days=1)


@dataclass
class EventRecord:
    row_id: int
    event: str
    start_date: date
    end_date: date
    city: str
    source: str
    nearest_airport: str | None
    impact: int | None
    predicted_attendance: int | None
    bridge_start: date | None
    bridge_end: date | None
    impact_by_day: dict[date, int]
    impact_by_day_bridge: dict[date, int]
    original: dict[str, Any]

    @property
    def effective_start(self) -> date:
        if self.bridge_start and self.bridge_end:
            return self.bridge_start
        return self.start_date

    @property
    def effective_end(self) -> date:
        if self.bridge_start and self.bridge_end:
            return self.bridge_end
        return self.end_date

    def overlaps(self, start: date, end: date) -> bool:
        return self.effective_end >= start and self.effective_start <= end


@dataclass
class MappingResult:
    row_id: int
    method: str
    airports: list[str]
    reason: str | None = None


def _iso_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("must be a date in YYYY-MM-DD format") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="special_days.backtest",
        description="Enrich event airports and correlate airport-day LF against 2025-2026 special events.",
    )
    parser.add_argument("--lf", required=True, help="Path to the LF workbook (.xlsx).")
    parser.add_argument("--events", required=True, help="Path to the special-events workbook (.xlsx).")
    parser.add_argument(
        "--out-dir",
        default="out/backtest",
        help="Output directory for the enriched and correlation workbooks (default: out/backtest).",
    )
    parser.add_argument("--lf-sheet", default=None, help="Optional LF sheet name. Defaults to the active sheet.")
    parser.add_argument(
        "--events-sheet",
        default=None,
        help="Optional events sheet name. Defaults to the active sheet.",
    )
    parser.add_argument(
        "--start-date",
        type=_iso_date,
        default=DEFAULT_START_DATE,
        help="Inclusive analysis start date YYYY-MM-DD (default: 2025-01-01).",
    )
    parser.add_argument(
        "--end-date",
        type=_iso_date,
        default=DEFAULT_END_DATE,
        help="Inclusive analysis end date YYYY-MM-DD (default: 2026-12-31).",
    )
    parser.add_argument(
        "--high-impact-threshold",
        type=int,
        default=80,
        help="Daily impact threshold used to flag high-impact dates and events (default: 80).",
    )
    return parser


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("_", " "))


def _turkish_normalized(value: str) -> str:
    translated = value.translate(
        str.maketrans(
            {
                "İ": "I",
                "I": "I",
                "ı": "i",
                "Ş": "S",
                "ş": "s",
                "Ğ": "G",
                "ğ": "g",
                "Ü": "U",
                "ü": "u",
                "Ö": "O",
                "ö": "o",
                "Ç": "C",
                "ç": "c",
            }
        )
    )
    text = unicodedata.normalize("NFKD", translated)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


def _clean_iata(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"[^A-Za-z]", "", str(value).strip()).upper()
    if len(text) != 3:
        return None
    return text


def _event_group_name(event_name: str) -> str:
    text = re.sub(r"\s+", " ", event_name.strip())
    text = re.sub(r"\s+\d+\.\s*G[uü]n$", "", text, flags=re.IGNORECASE)
    return text.strip()


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_curve_json(value: Any) -> dict[date, int]:
    if value is None or value == "":
        return {}
    if isinstance(value, dict):
        raw = value
    else:
        text = str(value).strip()
        if not text:
            return {}
        try:
            raw = json.loads(text)
        except json.JSONDecodeError:
            return {}
    curve: dict[date, int] = {}
    for key, weight in raw.items():
        curve_date = _parse_date(key)
        if curve_date is None:
            continue
        parsed_weight = _parse_int(weight)
        if parsed_weight is None:
            continue
        curve[curve_date] = parsed_weight
    return curve


def _load_sheet_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        if any(value not in (None, "") for value in row.values()):
            data_rows.append(row)
    return headers, data_rows


def _required_event_columns(rows: list[dict[str, Any]]) -> dict[str, str]:
    if not rows:
        return {}
    by_normalized = {_normalize_header(name): name for name in rows[0].keys()}
    required = {
        "event": "event",
        "start date": "start_date",
        "end date": "end_date",
        "city": "city",
        "source": "source",
    }
    resolved: dict[str, str] = {}
    missing: list[str] = []
    for header, alias in required.items():
        original = by_normalized.get(header)
        if original is None:
            missing.append(header)
        else:
            resolved[alias] = original
    if missing:
        raise ValueError(f"Events workbook is missing required columns: {', '.join(missing)}")
    optional_map = {
        "nearest airport": "nearest_airport",
        "impact": "impact",
        "predicted attendance": "predicted_attendance",
        "bridge start": "bridge_start",
        "bridge end": "bridge_end",
        "impact by day": "impact_by_day",
        "impact by day (bridge)": "impact_by_day_bridge",
    }
    for header, alias in optional_map.items():
        original = by_normalized.get(header)
        if original is not None:
            resolved[alias] = original
    return resolved


def _required_lf_columns(rows: list[dict[str, Any]]) -> dict[str, str]:
    if not rows:
        return {}
    by_normalized = {_normalize_header(name): name for name in rows[0].keys()}
    resolved: dict[str, str] = {}

    airport = by_normalized.get("airport")
    if airport is None:
        raise ValueError("LF workbook is missing required column: AIRPORT")
    resolved["airport"] = airport

    flight_date = by_normalized.get("flight date") or by_normalized.get("flightdate")
    ymd = by_normalized.get("ymd")
    if flight_date is None and ymd is None:
        raise ValueError("LF workbook must contain either flight_date or YMD")
    if flight_date is not None:
        resolved["flight_date"] = flight_date
    if ymd is not None:
        resolved["ymd"] = ymd

    lf = by_normalized.get("lf")
    lf_pct = by_normalized.get("lf pct")
    if lf is None and lf_pct is None:
        raise ValueError("LF workbook must contain either LF or LF_PCT")
    if lf is not None:
        resolved["lf"] = lf
    if lf_pct is not None:
        resolved["lf_pct"] = lf_pct

    for optional in ("boarded pax", "capacity"):
        original = by_normalized.get(optional)
        if original is not None:
            resolved[optional.replace(" ", "_")] = original
    return resolved


def load_event_records(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[EventRecord]]:
    headers, raw_rows = _load_sheet_rows(path, sheet_name)
    col = _required_event_columns(raw_rows)
    events: list[EventRecord] = []
    for row_id, row in enumerate(raw_rows, start=1):
        start = _parse_date(row.get(col["start_date"]))
        end = _parse_date(row.get(col["end_date"]))
        if start is None or end is None:
            continue
        events.append(
            EventRecord(
                row_id=row_id,
                event=str(row.get(col["event"], "") or "").strip(),
                start_date=start,
                end_date=end,
                city=str(row.get(col["city"], "") or "").strip(),
                source=str(row.get(col["source"], "") or "").strip(),
                nearest_airport=_clean_iata(row.get(col.get("nearest_airport", ""))),
                impact=_parse_int(row.get(col.get("impact", ""))),
                predicted_attendance=_parse_int(row.get(col.get("predicted_attendance", ""))),
                bridge_start=_parse_date(row.get(col.get("bridge_start", ""))),
                bridge_end=_parse_date(row.get(col.get("bridge_end", ""))),
                impact_by_day=_parse_curve_json(row.get(col.get("impact_by_day", ""))),
                impact_by_day_bridge=_parse_curve_json(row.get(col.get("impact_by_day_bridge", ""))),
                original=row,
            )
        )
    return headers, events


def load_lf_rows(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    headers, raw_rows = _load_sheet_rows(path, sheet_name)
    col = _required_lf_columns(raw_rows)
    parsed_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        airport = _clean_iata(row.get(col["airport"]))
        if not airport:
            continue

        flight_date = None
        if "flight_date" in col:
            flight_date = _parse_date(row.get(col["flight_date"]))
        if flight_date is None and "ymd" in col:
            ymd_text = str(row.get(col["ymd"], "") or "").strip()
            if len(ymd_text) == 8 and ymd_text.isdigit():
                flight_date = date(int(ymd_text[:4]), int(ymd_text[4:6]), int(ymd_text[6:8]))
        if flight_date is None:
            continue

        lf = _parse_float(row.get(col.get("lf", "")))
        lf_pct = _parse_float(row.get(col.get("lf_pct", "")))
        if lf is None and lf_pct is not None:
            lf = lf_pct / 100.0
        if lf_pct is None and lf is not None:
            lf_pct = round(lf * 100.0, 2)

        parsed = dict(row)
        parsed["AIRPORT"] = airport
        parsed["flight_date"] = flight_date
        parsed["LF"] = lf
        parsed["LF_PCT"] = lf_pct
        parsed_rows.append(parsed)
    return headers, parsed_rows


def _airports_by_city() -> dict[str, list[dict[str, Any]]]:
    lookup: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for airport in load_airports():
        lookup[_turkish_normalized(airport["city"])].append(airport)
    return dict(lookup)


def resolve_airports(
    events: list[EventRecord],
    lf_airports: set[str],
) -> tuple[list[MappingResult], list[dict[str, Any]]]:
    city_lookup = _airports_by_city()
    tr_airports = sorted(
        airport["iata"] for airport in load_airports() if airport["country"] == "TR" and airport["iata"] in lf_airports
    )
    mappings: list[MappingResult] = []
    unresolved: list[dict[str, Any]] = []

    for event in events:
        direct = event.nearest_airport
        if direct:
            if direct in lf_airports:
                mappings.append(MappingResult(event.row_id, "nearest_airport", [direct]))
            else:
                mappings.append(MappingResult(event.row_id, "unresolved", [], "nearest_airport_not_in_lf"))
                unresolved.append(_unresolved_row(event, "nearest_airport_not_in_lf"))
            continue

        city_key = _turkish_normalized(event.city)
        if city_key == "nationwide tr":
            if tr_airports:
                mappings.append(MappingResult(event.row_id, "nationwide_tr", tr_airports))
            else:
                mappings.append(MappingResult(event.row_id, "unresolved", [], "no_tr_airports_in_lf"))
                unresolved.append(_unresolved_row(event, "no_tr_airports_in_lf"))
            continue

        candidates = city_lookup.get(city_key, [])
        matched_airports = sorted(airport["iata"] for airport in candidates if airport["iata"] in lf_airports)
        if matched_airports:
            mappings.append(MappingResult(event.row_id, "city_exact", matched_airports))
            continue

        reason = "missing_city_and_airport" if not city_key else "city_not_found_or_not_in_lf"
        mappings.append(MappingResult(event.row_id, "unresolved", [], reason))
        unresolved.append(_unresolved_row(event, reason))

    return mappings, unresolved


def _unresolved_row(event: EventRecord, reason: str) -> dict[str, Any]:
    row = {header: event.original.get(header) for header in EVENT_HEADERS if header in event.original}
    row.update(
        {
            "row_id": event.row_id,
            "effective_start": event.effective_start,
            "effective_end": event.effective_end,
            "unresolved_reason": reason,
        }
    )
    return row


def _clip_range(start: date, end: date, min_date: date, max_date: date) -> tuple[date, date] | None:
    clipped_start = max(start, min_date)
    clipped_end = min(end, max_date)
    if clipped_end < clipped_start:
        return None
    return clipped_start, clipped_end


def _date_iter(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += _ONE_DAY


def expand_events_to_airport_days(
    events: list[EventRecord],
    mappings: list[MappingResult],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    mapping_by_row = {mapping.row_id: mapping for mapping in mappings}
    expanded: list[dict[str, Any]] = []
    for event in events:
        mapping = mapping_by_row.get(event.row_id)
        if mapping is None or not mapping.airports:
            continue
        clipped = _clip_range(event.effective_start, event.effective_end, start, end)
        if clipped is None:
            continue
        clipped_start, clipped_end = clipped
        curve = event.impact_by_day_bridge or event.impact_by_day
        curve_source = "impact_by_day_bridge" if event.impact_by_day_bridge else "impact_by_day" if event.impact_by_day else "row_impact"
        for event_date in _date_iter(clipped_start, clipped_end):
            impact = curve.get(event_date, event.impact if event.impact is not None else 0)
            for airport in mapping.airports:
                expanded.append(
                    {
                        "row_id": event.row_id,
                        "Event": event.event,
                        "Source": event.source,
                        "City": event.city,
                        "Airport": airport,
                        "Event date": event_date,
                        "Impact": impact,
                        "Predicted attendance": event.predicted_attendance,
                        "Start date": event.start_date,
                        "End date": event.end_date,
                        "Bridge start": event.bridge_start,
                        "Bridge end": event.bridge_end,
                        "Effective start": event.effective_start,
                        "Effective end": event.effective_end,
                        "Mapping method": mapping.method,
                        "Curve source": curve_source,
                    }
                )
    expanded.sort(key=lambda row: (row["Event date"], row["Airport"], row["Event"], row["row_id"]))
    return expanded


def aggregate_airport_day_features(
    rows: list[dict[str, Any]],
    high_impact_threshold: int,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[date, str], dict[str, Any]] = {}
    for row in rows:
        key = (row["Event date"], row["Airport"])
        agg = grouped.get(key)
        if agg is None:
            agg = {
                "event_date": row["Event date"],
                "airport": row["Airport"],
                "max_impact": row["Impact"] if row["Impact"] is not None else 0,
                "sum_impact": 0,
                "sum_predicted_attendance": 0,
                "n_events": 0,
                "high_impact_event_count": 0,
                "event_names": set(),
                "high_impact_event_names": set(),
                "sources": set(),
            }
            grouped[key] = agg
        impact = row["Impact"] if row["Impact"] is not None else 0
        if impact > agg["max_impact"]:
            agg["max_impact"] = impact
        agg["sum_impact"] += impact
        attendance = row["Predicted attendance"] or 0
        agg["sum_predicted_attendance"] += attendance
        agg["n_events"] += 1
        agg["event_names"].add(row["Event"])
        if impact >= high_impact_threshold:
            agg["high_impact_event_count"] += 1
            agg["high_impact_event_names"].add(row["Event"])
        agg["sources"].add(row["Source"])

    result: list[dict[str, Any]] = []
    for key in sorted(grouped):
        agg = grouped[key]
        result.append(
            {
                "event_date": agg["event_date"],
                "airport": agg["airport"],
                "max_impact": agg["max_impact"],
                "sum_impact": agg["sum_impact"],
                "sum_predicted_attendance": agg["sum_predicted_attendance"],
                "n_events": agg["n_events"],
                "high_impact_event_count": agg["high_impact_event_count"],
                "high_impact_flag": 1 if agg["high_impact_event_count"] else 0,
                "special_event_flag": 1 if agg["n_events"] else 0,
                "event_names": " | ".join(sorted(agg["event_names"])),
                "high_impact_event_names": " | ".join(sorted(agg["high_impact_event_names"])),
                "sources": ",".join(sorted(agg["sources"])),
            }
        )
    return result


def join_lf_with_events(
    lf_rows: list[dict[str, Any]],
    features: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    feature_map = {(row["event_date"], row["airport"]): row for row in features}
    joined: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    for row in lf_rows:
        feature = feature_map.get((row["flight_date"], row["AIRPORT"]))
        joined_row = dict(row)
        joined_row.update(
            {
                "event_date": feature["event_date"] if feature else None,
                "max_impact": feature["max_impact"] if feature else 0,
                "sum_impact": feature["sum_impact"] if feature else 0,
                "sum_predicted_attendance": feature["sum_predicted_attendance"] if feature else 0,
                "n_events": feature["n_events"] if feature else 0,
                "high_impact_event_count": feature["high_impact_event_count"] if feature else 0,
                "high_impact_flag": feature["high_impact_flag"] if feature else 0,
                "special_event_flag": feature["special_event_flag"] if feature else 0,
                "event_names": feature["event_names"] if feature else "",
                "high_impact_event_names": feature["high_impact_event_names"] if feature else "",
                "sources": feature["sources"] if feature else "",
            }
        )
        joined.append(joined_row)
        if feature is None:
            unmatched.append(dict(joined_row))
    joined.sort(key=lambda row: (row["flight_date"], row["AIRPORT"]))
    unmatched.sort(key=lambda row: (row["flight_date"], row["AIRPORT"]))
    return joined, unmatched


def _rank(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(indexed):
        j = i
        while j + 1 < len(indexed) and indexed[j + 1][1] == indexed[i][1]:
            j += 1
        rank = (i + j + 2) / 2.0
        for offset in range(i, j + 1):
            ranks[indexed[offset][0]] = rank
        i = j + 1
    return ranks


def _pearson(x: list[float], y: list[float]) -> float | None:
    if len(x) < 2 or len(y) < 2 or len(x) != len(y):
        return None
    mean_x = sum(x) / len(x)
    mean_y = sum(y) / len(y)
    centered_x = [value - mean_x for value in x]
    centered_y = [value - mean_y for value in y]
    sum_x = sum(value * value for value in centered_x)
    sum_y = sum(value * value for value in centered_y)
    denom = math.sqrt(sum_x * sum_y)
    if denom == 0:
        return None
    cov = sum(a * b for a, b in zip(centered_x, centered_y))
    return cov / denom


def spearman_correlation(x: list[float], y: list[float]) -> float | None:
    if len(x) < 2 or len(y) < 2 or len(x) != len(y):
        return None
    return _pearson(_rank(x), _rank(y))


def correlation_rows(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    metrics = [
        ("max_impact", "LF_PCT"),
        ("sum_impact", "LF_PCT"),
        ("sum_predicted_attendance", "LF_PCT"),
        ("n_events", "LF_PCT"),
        ("high_impact_event_count", "LF_PCT"),
        ("high_impact_flag", "LF_PCT"),
        ("special_event_flag", "LF_PCT"),
    ]
    rows: list[dict[str, Any]] = []
    for feature_name, target_name in metrics:
        pairs = [
            (float(row[target_name]), float(row[feature_name]))
            for row in joined_rows
            if row.get(target_name) is not None and row.get(feature_name) is not None
        ]
        value = spearman_correlation([pair[0] for pair in pairs], [pair[1] for pair in pairs]) if pairs else None
        rows.append(
            {
                "feature": feature_name,
                "target": target_name,
                "n_rows": len(pairs),
                "spearman": value,
            }
        )
    return rows


def correlation_by_airport(joined_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        grouped[row["AIRPORT"]].append(row)
    result: list[dict[str, Any]] = []
    for airport in sorted(grouped):
        rows = grouped[airport]
        result.append(
            {
                "AIRPORT": airport,
                "n_rows": len(rows),
                "days_with_events": sum(1 for row in rows if row["special_event_flag"]),
                "days_with_high_impact_events": sum(1 for row in rows if row["high_impact_flag"]),
                "avg_lf_pct": _mean([row["LF_PCT"] for row in rows if row["LF_PCT"] is not None]),
                "corr_max_impact": _feature_corr(rows, "max_impact"),
                "corr_sum_impact": _feature_corr(rows, "sum_impact"),
                "corr_sum_predicted_attendance": _feature_corr(rows, "sum_predicted_attendance"),
                "corr_n_events": _feature_corr(rows, "n_events"),
                "corr_high_impact_event_count": _feature_corr(rows, "high_impact_event_count"),
                "corr_high_impact_flag": _feature_corr(rows, "high_impact_flag"),
                "corr_special_event_flag": _feature_corr(rows, "special_event_flag"),
            }
        )
    return result


def daily_timeline_rows(
    expanded_rows: list[dict[str, Any]],
    joined_rows: list[dict[str, Any]],
    start: date,
    end: date,
    high_impact_threshold: int,
) -> list[dict[str, Any]]:
    event_daily: dict[date, dict[str, Any]] = {}
    seen_event_days: set[tuple[int, date]] = set()
    for row in expanded_rows:
        event_date = row["Event date"]
        row_id = row["row_id"]
        dedupe_key = (row_id, event_date)
        if dedupe_key in seen_event_days:
            continue
        seen_event_days.add(dedupe_key)

        impact = row["Impact"] if row["Impact"] is not None else 0
        event_name = row["Event"]
        event_group = _event_group_name(event_name)
        agg = event_daily.get(event_date)
        if agg is None:
            agg = {
                "date": event_date,
                "event_peak_impact": 0,
                "event_sum_impact": 0,
                "n_unique_events": 0,
                "high_impact_event_count": 0,
                "event_names": set(),
                "event_group_names": set(),
                "high_impact_event_names": set(),
                "high_impact_event_group_names": set(),
                "sources": set(),
            }
            event_daily[event_date] = agg
        if impact > agg["event_peak_impact"]:
            agg["event_peak_impact"] = impact
        agg["event_sum_impact"] += impact
        agg["n_unique_events"] += 1
        agg["event_names"].add(event_name)
        agg["event_group_names"].add(event_group)
        agg["sources"].add(row["Source"])
        if impact >= high_impact_threshold:
            agg["high_impact_event_count"] += 1
            agg["high_impact_event_names"].add(event_name)
            agg["high_impact_event_group_names"].add(event_group)

    lf_by_date: dict[date, list[dict[str, Any]]] = defaultdict(list)
    for row in joined_rows:
        lf_by_date[row["flight_date"]].append(row)

    timeline: list[dict[str, Any]] = []
    for current_date in _date_iter(start, end):
        lf_rows_for_date = lf_by_date.get(current_date, [])
        event_lf_rows = [row for row in lf_rows_for_date if row["special_event_flag"]]
        event_summary = event_daily.get(current_date, {})
        peak_impact = event_summary.get("event_peak_impact", 0)
        high_impact_count = event_summary.get("high_impact_event_count", 0)
        timeline.append(
            {
                "date": current_date,
                "avg_lf_pct_all": _mean([row["LF_PCT"] for row in lf_rows_for_date if row["LF_PCT"] is not None]),
                "avg_lf_pct_event_airports": _mean([row["LF_PCT"] for row in event_lf_rows if row["LF_PCT"] is not None]),
                "max_lf_pct_event_airports": max((row["LF_PCT"] for row in event_lf_rows if row["LF_PCT"] is not None), default=None),
                "airports_with_events": len(event_lf_rows),
                "event_peak_impact": peak_impact,
                "event_sum_impact": event_summary.get("event_sum_impact", 0),
                "n_unique_events": event_summary.get("n_unique_events", 0),
                "high_impact_event_count": high_impact_count,
                "high_impact_flag": 1 if high_impact_count else 0,
                "high_impact_signal": peak_impact if high_impact_count else 0,
                "event_names": " | ".join(sorted(event_summary.get("event_names", set()))),
                "event_group_names": " | ".join(sorted(event_summary.get("event_group_names", set()))),
                "high_impact_event_names": " | ".join(sorted(event_summary.get("high_impact_event_names", set()))),
                "high_impact_event_group_names": " | ".join(sorted(event_summary.get("high_impact_event_group_names", set()))),
                "sources": ",".join(sorted(event_summary.get("sources", set()))),
            }
        )
    return timeline


def _mean(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def _feature_corr(rows: list[dict[str, Any]], feature_name: str) -> float | None:
    pairs = [
        (float(row["LF_PCT"]), float(row[feature_name]))
        for row in rows
        if row.get("LF_PCT") is not None and row.get(feature_name) is not None
    ]
    if not pairs:
        return None
    return spearman_correlation([pair[0] for pair in pairs], [pair[1] for pair in pairs])


def _mapping_rows(events: list[EventRecord], mappings: list[MappingResult]) -> list[dict[str, Any]]:
    by_row = {mapping.row_id: mapping for mapping in mappings}
    rows: list[dict[str, Any]] = []
    for event in events:
        mapping = by_row[event.row_id]
        rows.append(
            {
                "row_id": event.row_id,
                "Event": event.event,
                "Source": event.source,
                "City": event.city,
                "Nearest airport": event.nearest_airport,
                "Effective start": event.effective_start,
                "Effective end": event.effective_end,
                "Mapping method": mapping.method,
                "Matched airports": ", ".join(mapping.airports),
                "Reason": mapping.reason,
            }
        )
    return rows


def _filtered_raw_rows(events: list[EventRecord]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for event in events:
        row = {header: event.original.get(header) for header in EVENT_HEADERS if header in event.original}
        row.update(
            {
                "row_id": event.row_id,
                "Effective start": event.effective_start,
                "Effective end": event.effective_end,
            }
        )
        rows.append(row)
    return rows


def _coverage_summary(
    filtered_events: list[EventRecord],
    mappings: list[MappingResult],
    expanded_rows: list[dict[str, Any]],
    lf_rows: list[dict[str, Any]],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    rows = [
        {"metric": "analysis_start", "value": start},
        {"metric": "analysis_end", "value": end},
        {"metric": "filtered_events", "value": len(filtered_events)},
        {"metric": "resolved_events", "value": sum(1 for mapping in mappings if mapping.airports)},
        {"metric": "unresolved_events", "value": sum(1 for mapping in mappings if not mapping.airports)},
        {"metric": "event_day_airport_rows", "value": len(expanded_rows)},
        {"metric": "lf_rows", "value": len(lf_rows)},
        {"metric": "lf_airports", "value": len({row['AIRPORT'] for row in lf_rows})},
    ]
    by_method: dict[str, int] = defaultdict(int)
    for mapping in mappings:
        by_method[mapping.method] += 1
    for method in sorted(by_method):
        rows.append({"metric": f"mapping_method:{method}", "value": by_method[method]})
    by_source: dict[str, int] = defaultdict(int)
    for event in filtered_events:
        by_source[event.source] += 1
    for source in sorted(by_source):
        rows.append({"metric": f"source:{source}", "value": by_source[source]})
    return rows


def _sheet_headers(rows: list[dict[str, Any]], preferred: list[str]) -> list[str]:
    if rows:
        extras = [key for key in rows[0].keys() if key not in preferred]
    else:
        extras = []
    return preferred + extras


def _write_workbook(path: Path, sheets: list[tuple[str, list[str], list[dict[str, Any]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    for index, (name, headers, rows) in enumerate(sheets):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = name[:31]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
        _format_sheet_dates(sheet)
        _autosize(sheet)
    workbook.save(path)


def _format_sheet_dates(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = _DATETIME_FORMAT
            elif isinstance(cell.value, date):
                cell.number_format = _DATE_FORMAT


def _autosize(sheet) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            if value is None:
                continue
            length = len(str(value))
            widths[index] = min(max(widths.get(index, 0), length + 2), 60)
    for index, width in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def _add_timeline_chart(path: Path, sheet_name: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference

    workbook = load_workbook(path)
    try:
        sheet = workbook[sheet_name]
        if sheet.max_row <= 1:
            workbook.save(path)
            return

        headers = {cell.value: index + 1 for index, cell in enumerate(sheet[1])}
        date_col = headers.get("date")
        if date_col is None:
            workbook.save(path)
            return

        chart = LineChart()
        chart.title = "LF vs High-Impact Event Timeline"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Date"
        chart.height = 10
        chart.width = 22

        for header in ("avg_lf_pct_all", "avg_lf_pct_event_airports", "event_peak_impact", "high_impact_signal"):
            column = headers.get(header)
            if column is None:
                continue
            values = Reference(sheet, min_col=column, min_row=1, max_row=sheet.max_row)
            chart.add_data(values, titles_from_data=True)

        categories = Reference(sheet, min_col=date_col, min_row=2, max_row=sheet.max_row)
        chart.set_categories(categories)
        chart.legend.position = "r"
        sheet.add_chart(chart, "O2")
        workbook.save(path)
    finally:
        workbook.close()


def run_workflow(
    lf_path: str | Path,
    events_path: str | Path,
    out_dir: str | Path,
    *,
    start_date: date = DEFAULT_START_DATE,
    end_date: date = DEFAULT_END_DATE,
    lf_sheet: str | None = None,
    events_sheet: str | None = None,
    high_impact_threshold: int = 80,
) -> dict[str, Path]:
    _, lf_rows = load_lf_rows(lf_path, lf_sheet)
    lf_rows = [row for row in lf_rows if start_date <= row["flight_date"] <= end_date]
    _, event_records = load_event_records(events_path, events_sheet)

    filtered_events = [event for event in event_records if event.overlaps(start_date, end_date)]
    lf_airports = {row["AIRPORT"] for row in lf_rows}
    mappings, unresolved = resolve_airports(filtered_events, lf_airports)
    expanded_rows = expand_events_to_airport_days(filtered_events, mappings, start_date, end_date)
    features = aggregate_airport_day_features(expanded_rows, high_impact_threshold)
    joined_rows, unmatched_lf_rows = join_lf_with_events(lf_rows, features)
    overall_corr = correlation_rows(joined_rows)
    by_airport_corr = correlation_by_airport(joined_rows)
    timeline_rows = daily_timeline_rows(expanded_rows, joined_rows, start_date, end_date, high_impact_threshold)
    high_impact_timeline_rows = [row for row in timeline_rows if row["high_impact_flag"]]

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    suffix = f"{start_date.isoformat()}_{end_date.isoformat()}"
    enriched_path = out_path / f"special_days_enriched_{suffix}.xlsx"
    correlation_path = out_path / f"lf_vs_special_events_{suffix}.xlsx"

    _write_workbook(
        enriched_path,
        [
            (
                "raw_filtered",
                _sheet_headers(_filtered_raw_rows(filtered_events), EVENT_HEADERS + ["row_id", "Effective start", "Effective end"]),
                _filtered_raw_rows(filtered_events),
            ),
            (
                "airport_mapping",
                [
                    "row_id",
                    "Event",
                    "Source",
                    "City",
                    "Nearest airport",
                    "Effective start",
                    "Effective end",
                    "Mapping method",
                    "Matched airports",
                    "Reason",
                ],
                _mapping_rows(filtered_events, mappings),
            ),
            (
                "event_day_airport",
                [
                    "row_id",
                    "Event",
                    "Source",
                    "City",
                    "Airport",
                    "Event date",
                    "Impact",
                    "Predicted attendance",
                    "Start date",
                    "End date",
                    "Bridge start",
                    "Bridge end",
                    "Effective start",
                    "Effective end",
                    "Mapping method",
                    "Curve source",
                ],
                expanded_rows,
            ),
            (
                "unresolved_events",
                _sheet_headers(unresolved, EVENT_HEADERS + ["row_id", "effective_start", "effective_end", "unresolved_reason"]),
                unresolved,
            ),
            (
                "coverage_summary",
                ["metric", "value"],
                _coverage_summary(filtered_events, mappings, expanded_rows, lf_rows, start_date, end_date),
            ),
        ],
    )

    _write_workbook(
        correlation_path,
        [
            (
                "lf_raw",
                _sheet_headers(lf_rows, LF_HEADERS),
                lf_rows,
            ),
            (
                "events_airport_day",
                [
                    "event_date",
                    "airport",
                    "max_impact",
                    "sum_impact",
                    "sum_predicted_attendance",
                    "n_events",
                    "high_impact_event_count",
                    "high_impact_flag",
                    "special_event_flag",
                    "event_names",
                    "high_impact_event_names",
                    "sources",
                ],
                features,
            ),
            (
                "joined_airport_day",
                _sheet_headers(
                    joined_rows,
                    LF_HEADERS + [
                        "event_date",
                        "max_impact",
                        "sum_impact",
                        "sum_predicted_attendance",
                        "n_events",
                        "high_impact_event_count",
                        "high_impact_flag",
                        "special_event_flag",
                        "event_names",
                        "high_impact_event_names",
                        "sources",
                    ],
                ),
                joined_rows,
            ),
            (
                "daily_timeline",
                [
                    "date",
                    "avg_lf_pct_all",
                    "avg_lf_pct_event_airports",
                    "max_lf_pct_event_airports",
                    "airports_with_events",
                    "event_peak_impact",
                    "event_sum_impact",
                    "n_unique_events",
                    "high_impact_event_count",
                    "high_impact_flag",
                    "high_impact_signal",
                    "event_names",
                    "event_group_names",
                    "high_impact_event_names",
                    "high_impact_event_group_names",
                    "sources",
                ],
                timeline_rows,
            ),
            (
                "high_impact_timeline",
                [
                    "date",
                    "avg_lf_pct_all",
                    "avg_lf_pct_event_airports",
                    "max_lf_pct_event_airports",
                    "airports_with_events",
                    "event_peak_impact",
                    "event_sum_impact",
                    "n_unique_events",
                    "high_impact_event_count",
                    "high_impact_flag",
                    "high_impact_signal",
                    "event_names",
                    "event_group_names",
                    "high_impact_event_names",
                    "high_impact_event_group_names",
                    "sources",
                ],
                high_impact_timeline_rows,
            ),
            (
                "corr_overall",
                ["feature", "target", "n_rows", "spearman"],
                overall_corr,
            ),
            (
                "corr_by_airport",
                [
                    "AIRPORT",
                    "n_rows",
                    "days_with_events",
                    "days_with_high_impact_events",
                    "avg_lf_pct",
                    "corr_max_impact",
                    "corr_sum_impact",
                    "corr_sum_predicted_attendance",
                    "corr_n_events",
                    "corr_high_impact_event_count",
                    "corr_high_impact_flag",
                    "corr_special_event_flag",
                ],
                by_airport_corr,
            ),
            (
                "unmatched_lf_days",
                _sheet_headers(
                    unmatched_lf_rows,
                    LF_HEADERS + [
                        "event_date",
                        "max_impact",
                        "sum_impact",
                        "sum_predicted_attendance",
                        "n_events",
                        "high_impact_event_count",
                        "high_impact_flag",
                        "special_event_flag",
                        "event_names",
                        "high_impact_event_names",
                        "sources",
                    ],
                ),
                unmatched_lf_rows,
            ),
        ],
    )
    _add_timeline_chart(correlation_path, "daily_timeline")
    _add_timeline_chart(correlation_path, "high_impact_timeline")

    return {
        "enriched_workbook": enriched_path,
        "correlation_workbook": correlation_path,
    }


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    outputs = run_workflow(
        lf_path=args.lf,
        events_path=args.events,
        out_dir=args.out_dir,
        start_date=args.start_date,
        end_date=args.end_date,
        lf_sheet=args.lf_sheet,
        events_sheet=args.events_sheet,
        high_impact_threshold=args.high_impact_threshold,
    )
    print(f"Wrote enriched workbook: {outputs['enriched_workbook']}")
    print(f"Wrote correlation workbook: {outputs['correlation_workbook']}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())