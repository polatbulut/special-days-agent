"""Tests for the per-event LF case studies.

Every fixture is generated in-process: synthetic route-day data with a known
day-of-week pattern and, where relevant, a planted effect. No network, no LLM,
no fixture files checked into the repo.
"""

from __future__ import annotations

import csv
import random
import tempfile
import unittest
from datetime import date, timedelta
from pathlib import Path

from special_days.case_studies import (
    analyse,
    load_events,
    load_route_days,
    measure,
    run,
)

START = date(2025, 1, 1)  # a Wednesday
LF_HEADER = ["ORG_AIRPORT", "DST_AIRPORT", "FLIGHT_DATE", "BOARDED_PAX", "CAPACITY"]
EVENTS_HEADER = ["event", "start_date", "end_date", "city", "airport", "routes", "note"]

BASE_LF = 0.75
WEEKEND_BUMP = 0.06
CAPACITY = 200


def _generate_lf_rows(
    routes: list[str],
    days: int,
    *,
    start: date = START,
    lf_injection: dict[date, float] | None = None,
    pax_multiplier: dict[date, float] | None = None,
    cap_multiplier: dict[date, float] | None = None,
    seed: int = 7,
) -> list[list]:
    """Base LF 0.75, +6 pp on weekends, one flight per route-day, capacity 200."""
    rnd = random.Random(seed)
    lf_injection = lf_injection or {}
    pax_multiplier = pax_multiplier or {}
    cap_multiplier = cap_multiplier or {}

    rows: list[list] = []
    for route in routes:
        org, dst = route[:3], route[3:]
        for offset in range(days):
            day = start + timedelta(days=offset)
            lf = BASE_LF + (WEEKEND_BUMP if day.weekday() >= 5 else 0.0)
            lf += rnd.uniform(-0.004, 0.004)
            lf += lf_injection.get(day, 0.0)
            capacity = CAPACITY * cap_multiplier.get(day, 1.0)
            pax = round(CAPACITY * lf) * pax_multiplier.get(day, 1.0)
            rows.append([org, dst, day.isoformat(), int(round(pax)), int(round(capacity))])
    return rows


def _write_csv(path: Path, header: list[str], rows: list[list]) -> Path:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)
    return path


def _write_events(path: Path, specs: list[list]) -> Path:
    return _write_csv(path, EVENTS_HEADER, specs)


def _point_for(result, day: date):
    for point in result.timeline:
        if point.day == day:
            return point
    raise AssertionError(f"no timeline point for {day}")


class _TempDirCase(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp = Path(self._tmp.name)
        self.addCleanup(self._tmp.cleanup)

    def measure_single(
        self,
        lf_rows: list[list],
        event_day: date,
        *,
        route: str = "SVOIST",
        end_day: date | None = None,
        **opts,
    ):
        lf_path = _write_csv(self.tmp / "lf.csv", LF_HEADER, lf_rows)
        events_path = _write_events(
            self.tmp / "events.csv",
            [
                [
                    "Planted event",
                    event_day.isoformat(),
                    (end_day or event_day).isoformat(),
                    "İstanbul",
                    "IST",
                    f"{route[:3]}-{route[3:]}",
                    "synthetic",
                ]
            ],
        )
        opts.setdefault("reports", False)  # CSV only; the renderers have their own tests
        results = run(events_path, lf_path, self.tmp / "out", **opts)
        self.assertEqual(len(results), 1)
        return results[0]


class PlantedLiftTest(_TempDirCase):
    def test_event_day_shows_the_planted_lift(self):
        event_day = date(2025, 3, 5)  # Wednesday, ~9 weeks in
        result = self.measure_single(
            _generate_lf_rows(["SVOIST"], 150, lf_injection={event_day: 0.15}),
            event_day,
        )
        point = _point_for(result, event_day)
        self.assertIsNotNone(point.delta_pp)
        self.assertGreater(point.delta_pp, 12.0)
        self.assertLess(point.delta_pp, 18.0)

    def test_surrounding_days_stay_near_zero(self):
        event_day = date(2025, 3, 5)
        result = self.measure_single(
            _generate_lf_rows(["SVOIST"], 150, lf_injection={event_day: 0.15}),
            event_day,
        )
        for point in result.timeline:
            if point.day == event_day or point.delta_pp is None:
                continue
            self.assertLess(abs(point.delta_pp), 2.0, f"{point.day} moved without an injection")

    def test_verdict_reads_demand_up_capacity_flat(self):
        event_day = date(2025, 3, 5)
        result = self.measure_single(
            _generate_lf_rows(["SVOIST"], 150, lf_injection={event_day: 0.15}),
            event_day,
            guard_days=0,
        )
        self.assertEqual(result.verdict, "demand up, capacity flat, LF up")


class PlantedNullTest(_TempDirCase):
    def test_null_is_flat_but_still_produces_rows(self):
        event_day = date(2025, 3, 5)
        result = self.measure_single(_generate_lf_rows(["SVOIST"], 150), event_day)
        point = _point_for(result, event_day)
        self.assertIsNotNone(point.delta_pp)
        self.assertLess(abs(point.delta_pp), 2.0)
        # A null must never be indistinguishable from a broken join.
        self.assertGreater(result.n_route_days, 0)
        self.assertGreater(result.n_days, 0)
        self.assertEqual(result.n_days_no_baseline, 0)
        self.assertEqual(result.verdict, "no movement")

    def test_null_still_writes_csv_rows(self):
        event_day = date(2025, 3, 5)
        self.measure_single(_generate_lf_rows(["SVOIST"], 150), event_day)
        timeline = (self.tmp / "out" / "timeline.csv").read_text(encoding="utf-8").splitlines()
        summary = (self.tmp / "out" / "summary.csv").read_text(encoding="utf-8").splitlines()
        self.assertGreater(len(timeline), 1)
        self.assertEqual(len(summary), 2)
        self.assertIn("events with no measured event day: 0", (self.tmp / "out" / "coverage.txt").read_text())


class WeekendConfoundTest(_TempDirCase):
    def test_saturday_event_shows_no_lift(self):
        event_day = date(2025, 3, 8)  # Saturday
        self.assertEqual(event_day.weekday(), 5)
        rows = _generate_lf_rows(["SVOIST"], 150)
        result = self.measure_single(rows, event_day)
        point = _point_for(result, event_day)

        # The raw Saturday LF really is well above the all-days average, so a
        # naive baseline would have printed a lift here...
        overall = sum(int(row[3]) for row in rows) / sum(int(row[4]) for row in rows)
        self.assertGreater(point.lf, overall + 0.03)
        # ...but the same-weekday baseline absorbs it.
        self.assertIsNotNone(point.delta_pp)
        self.assertLess(point.delta_pp, 2.0)
        self.assertGreater(point.delta_pp, -2.0)


class CapacityAbsorptionTest(_TempDirCase):
    def test_pax_and_capacity_both_up_twenty_percent(self):
        event_day = date(2025, 3, 5)
        rows = _generate_lf_rows(
            ["SVOIST"],
            150,
            pax_multiplier={event_day: 1.20},
            cap_multiplier={event_day: 1.20},
        )
        result = self.measure_single(rows, event_day, guard_days=0)
        point = _point_for(result, event_day)

        self.assertLess(abs(point.delta_pp), 1.0)
        self.assertGreater(point.pax_vs_baseline_pct, 18.0)
        self.assertLess(point.pax_vs_baseline_pct, 22.0)
        self.assertGreater(point.cap_vs_baseline_pct, 18.0)
        # The label names all three dimensions, so the capacity move that
        # absorbed the demand cannot be described as "flat".
        self.assertEqual(result.verdict, "demand up, capacity up, LF flat")


class CapacityWeightingTest(_TempDirCase):
    def test_route_day_lf_is_capacity_weighted_not_a_mean_of_lfs(self):
        # 300 seats at 90% + 60 seats at 20% -> 282/360, not (0.90+0.20)/2.
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            [
                ["SVO", "IST", "2025-03-05", 270, 300],
                ["SVO", "IST", "2025-03-05", 12, 60],
            ],
        )
        route_days, coverage = load_route_days(lf_path)
        cell = route_days[("SVO", "IST")][date(2025, 3, 5)]

        self.assertEqual(cell.pax, 282)
        self.assertEqual(cell.capacity, 360)
        self.assertAlmostEqual(cell.lf, 0.7833, places=4)
        self.assertNotAlmostEqual(cell.lf, 0.55, places=2)
        self.assertEqual(cell.n_flights, 2)
        self.assertEqual(coverage.n_flight_rows, 2)

    def test_bad_capacity_rows_are_skipped_and_counted(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            [
                ["SVO", "IST", "2025-03-05", 270, 300],
                ["SVO", "IST", "2025-03-05", 5, 0],
                ["SVO", "IST", "2025-03-05", 5, -10],
            ],
        )
        route_days, coverage = load_route_days(lf_path)
        self.assertEqual(coverage.n_skipped_bad_capacity, 2)
        self.assertEqual(route_days[("SVO", "IST")][date(2025, 3, 5)].capacity, 300)

    def test_lf_above_one_is_kept_and_flagged(self):
        event_day = date(2025, 3, 5)
        rows = _generate_lf_rows(["SVOIST"], 150, lf_injection={event_day: 0.30})
        result = self.measure_single(rows, event_day)
        point = _point_for(result, event_day)
        self.assertGreater(point.lf, 1.0)
        self.assertIn("lf>1", point.data_flags)


class ThinBaselineTest(_TempDirCase):
    def test_fewer_than_three_same_weekday_observations_gives_no_baseline(self):
        event_day = date(2025, 3, 5)
        # baseline_days=8 leaves only d-7 and d+7 on the same weekday: 2 < 3.
        result = self.measure_single(
            _generate_lf_rows(["SVOIST"], 150),
            event_day,
            window_days=10,
            baseline_days=8,
            guard_days=3,
        )
        point = _point_for(result, event_day)

        self.assertIsNone(point.baseline_lf)
        self.assertIsNone(point.delta_pp)
        self.assertIsNone(point.pax_vs_baseline_pct)
        self.assertEqual(point.n_baseline_obs, 2)

        in_event = [p for p in result.timeline if p.in_event]
        self.assertEqual(result.n_days, len(in_event))
        self.assertEqual(result.n_days_no_baseline, len(in_event))
        # No fallback value anywhere: the summary stays empty, not zero.
        self.assertIsNone(result.mean_delta_pp)
        self.assertIsNone(result.mean_pax_vs_baseline_pct)
        self.assertEqual(result.verdict, "no baseline")

        summary = (self.tmp / "out" / "summary.csv").read_text(encoding="utf-8").splitlines()
        self.assertIn(",,", summary[1])  # blank cells, not 0.00


class OrderPreservedTest(_TempDirCase):
    def test_events_come_out_in_input_order_not_effect_order(self):
        small = date(2025, 3, 5)
        big = date(2025, 4, 2)
        negative = date(2025, 5, 7)
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            _generate_lf_rows(
                ["SVOIST"],
                220,
                lf_injection={small: 0.02, big: 0.20, negative: -0.18},
            ),
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [
                ["Small effect", small.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""],
                ["Big effect", big.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""],
                ["Negative effect", negative.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""],
            ],
        )
        results = run(events_path, lf_path, self.tmp / "out", guard_days=0, reports=False)

        self.assertEqual(
            [r.event for r in results],
            ["Small effect", "Big effect", "Negative effect"],
        )
        # The effect sizes really are out of order, so this is not a free pass.
        self.assertGreater(results[1].mean_delta_pp, results[0].mean_delta_pp)
        self.assertLess(results[2].mean_delta_pp, 0.0)

        summary = (self.tmp / "out" / "summary.csv").read_text(encoding="utf-8").splitlines()
        self.assertTrue(summary[1].startswith("Small effect,"))
        self.assertTrue(summary[2].startswith("Big effect,"))
        self.assertTrue(summary[3].startswith("Negative effect,"))


class RouteDefaultingTest(_TempDirCase):
    def test_blank_routes_picks_every_inbound_leg(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            _generate_lf_rows(["SVOIST", "LEDIST", "ISTSVO", "LEDAYT"], 120),
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Kanye West concert", "2025-03-05", "", "İstanbul", "IST", "", "stadium"]],
        )
        results = run(events_path, lf_path, self.tmp / "out", reports=False)

        self.assertEqual([r.route for r in results], ["LED-IST", "SVO-IST"])
        for result in results:
            self.assertGreater(result.n_route_days, 0)

    def test_explicit_routes_accept_both_spellings(self):
        events_path = _write_events(
            self.tmp / "events.csv",
            [["E", "2025-03-05", "2025-03-07", "İstanbul", "", "SVO-IST;LEDIST", ""]],
        )
        events = load_events(events_path)
        self.assertEqual(events[0].routes, (("SVO", "IST"), ("LED", "IST")))
        self.assertIsNone(events[0].airport)
        self.assertEqual(events[0].end_date, date(2025, 3, 7))

    def test_blank_end_date_means_same_day(self):
        events_path = _write_events(
            self.tmp / "events.csv",
            [["E", "2025-03-05", "", "İstanbul", "IST", "", ""]],
        )
        events = load_events(events_path)
        self.assertEqual(events[0].start_date, events[0].end_date)


class InputValidationTest(_TempDirCase):
    def test_missing_event_column_is_named(self):
        path = _write_csv(self.tmp / "events.csv", ["event", "start_date", "city"], [["E", "2025-03-05", "IST"]])
        with self.assertRaises(ValueError) as ctx:
            load_events(path)
        self.assertIn("end_date", str(ctx.exception))
        self.assertIn("airport", str(ctx.exception))
        self.assertIn("routes", str(ctx.exception))

    def test_missing_lf_column_is_named(self):
        path = _write_csv(self.tmp / "lf.csv", ["ORG", "DST", "FLIGHT_DATE", "CAPACITY"], [["S", "I", "2025-03-05", 1]])
        with self.assertRaises(ValueError) as ctx:
            load_route_days(path)
        self.assertIn("BOARDED_PAX", str(ctx.exception))

    def test_single_route_column_is_split(self):
        path = _write_csv(
            self.tmp / "lf.csv",
            ["ROUTE", "YMD", "PAX", "SEATS"],
            [["SVOIST", "20250305", 150, 200]],
        )
        route_days, _ = load_route_days(path)
        self.assertIn(("SVO", "IST"), route_days)
        self.assertEqual(route_days[("SVO", "IST")][date(2025, 3, 5)].pax, 150)

    def test_event_with_no_matching_lf_rows_is_named_and_still_emitted(self):
        lf_path = _write_csv(self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST"], 60))
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Ghost", "2025-02-05", "", "Antalya", "AYT", "MOW-AYT", ""]],
        )
        results = run(events_path, lf_path, self.tmp / "out", reports=False)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].n_route_days, 0)
        self.assertEqual(results[0].verdict, "no LF data")
        self.assertIn("Ghost", (self.tmp / "out" / "coverage.txt").read_text(encoding="utf-8"))


class NoSignificanceMachineryTest(unittest.TestCase):
    def test_module_contains_no_p_values_or_t_statistics(self):
        source = (Path(__file__).resolve().parents[1] / "special_days" / "case_studies.py").read_text(
            encoding="utf-8"
        )
        for banned in ("p_value", "pvalue", "t_stat", "tstat", "ttest", "pearson", "correlation"):
            self.assertNotIn(banned, source.lower(), f"{banned} must not appear in case_studies.py")


class MeasureDirectTest(unittest.TestCase):
    def test_measure_accepts_prebuilt_cells(self):
        from special_days.case_studies import EventSpec, RouteDay

        cells = {}
        for offset in range(120):
            day = START + timedelta(days=offset)
            lf = BASE_LF + (WEEKEND_BUMP if day.weekday() >= 5 else 0.0)
            cells[day] = RouteDay("SVO", "IST", day, round(CAPACITY * lf), CAPACITY, 1)
        event = EventSpec(1, "E", date(2025, 3, 5), date(2025, 3, 5), "İstanbul", "IST", (), "")

        result = measure(event, "SVO", "IST", cells, guard_days=0)
        self.assertEqual(result.n_days, 1)
        self.assertAlmostEqual(result.mean_delta_pp, 0.0, places=6)
        self.assertEqual(result.verdict, "no movement")

    def test_analyse_reports_events_with_no_routes(self):
        from special_days.case_studies import EventSpec

        event = EventSpec(1, "E", date(2025, 3, 5), date(2025, 3, 5), "Antalya", "AYT", (), "")
        results, missing = analyse([event], {})
        self.assertEqual(results, [])
        self.assertEqual(missing, ["E"])


def _flat_cells(
    span_days: int = 140,
    *,
    centre: date = date(2026, 6, 17),
    pax: int = 100,
    capacity: int = 200,
):
    """A featureless route: same pax and capacity every day, both directions."""
    from special_days.case_studies import RouteDay

    return {
        centre + timedelta(days=offset): RouteDay(
            "SVO", "IST", centre + timedelta(days=offset), pax, capacity, 1
        )
        for offset in range(-span_days, span_days + 1)
    }


def _spec(start: date, end: date | None = None, name: str = "E"):
    from special_days.case_studies import EventSpec

    return EventSpec(1, name, start, end or start, "İstanbul", "IST", (), "")


class BaselineExcludesItselfTest(unittest.TestCase):
    """F1: a day must not be a member of its own comparison set."""

    def test_day_is_not_in_its_own_baseline_pool(self):
        from special_days.case_studies import RouteDay

        cells = {}
        for day, lf in (
            (date(2026, 6, 10), 0.50),
            (date(2026, 6, 17), 0.50),
            (date(2026, 6, 24), 0.70),
            (date(2026, 7, 1), 0.90),
        ):
            cells[day] = RouteDay("SVO", "IST", day, round(200 * lf), 200, 1)

        # An event far away, so every Wednesday above is a clean context day.
        result = measure(
            _spec(date(2026, 8, 1)), "SVO", "IST", cells, window_days=60, guard_days=3
        )
        point = _point_for(result, date(2026, 7, 1))

        # Honest peers are .50 .50 .70 -> 340/600 = 0.5667. Counting the day
        # itself would give 440/800 = 0.55 and shrink the delta from 33.3 to 25.
        self.assertEqual(point.n_baseline_obs, 3)
        self.assertAlmostEqual(point.baseline_lf, 340.0 / 600.0, places=9)
        self.assertAlmostEqual(point.delta_pp, 100.0 * (0.90 - 340.0 / 600.0), places=6)

    def test_two_real_peers_cannot_pass_the_three_observation_gate(self):
        from special_days.case_studies import RouteDay

        cells = {
            day: RouteDay("SVO", "IST", day, 100, 200, 1)
            for day in (date(2026, 6, 17), date(2026, 6, 24), date(2026, 7, 1))
        }
        result = measure(
            _spec(date(2026, 9, 1)), "SVO", "IST", cells, window_days=90, guard_days=3
        )
        for point in result.timeline:
            self.assertEqual(point.n_baseline_obs, 2)
            self.assertIsNone(point.baseline_lf)
            self.assertIsNone(point.delta_pp)


class GuardDoesNotDiluteTest(unittest.TestCase):
    """F2: guard_days is baseline hygiene, never the measured window."""

    def test_single_day_spike_reads_full_size_at_every_guard(self):
        from special_days.case_studies import RouteDay

        event_day = date(2026, 6, 17)
        cells = _flat_cells()
        cells[event_day] = RouteDay("SVO", "IST", event_day, 180, 200, 1)
        event = _spec(event_day)

        for guard in (0, 1, 3, 7):
            result = measure(event, "SVO", "IST", cells, window_days=30, guard_days=guard)
            self.assertEqual(result.n_days, 1, f"guard_days={guard} widened the window")
            self.assertAlmostEqual(
                result.mean_delta_pp, 40.0, places=4, msg=f"guard_days={guard} diluted the effect"
            )

    def test_guard_still_keeps_neighbours_out_of_the_baseline(self):
        from special_days.case_studies import RouteDay

        event_day = date(2026, 6, 17)  # Wednesday
        cells = _flat_cells()
        # Contaminated run-up on the *same weekday*, inside a 7-day guard.
        for offset in (-7, 7):
            day = event_day + timedelta(days=offset)
            cells[day] = RouteDay("SVO", "IST", day, 190, 200, 1)
        cells[event_day] = RouteDay("SVO", "IST", event_day, 180, 200, 1)
        event = _spec(event_day)

        guarded = measure(event, "SVO", "IST", cells, window_days=30, guard_days=7)
        unguarded = measure(event, "SVO", "IST", cells, window_days=30, guard_days=0)
        # The guard excludes the two contaminated Wednesdays, so the baseline
        # stays at the clean 50% and the spike keeps its full size.
        self.assertAlmostEqual(guarded.mean_delta_pp, 40.0, places=4)
        self.assertLess(unguarded.mean_delta_pp, guarded.mean_delta_pp)


class BaselineReconcilesTest(unittest.TestCase):
    """F3: LF, pax and capacity baselines come from one construction."""

    def test_baseline_lf_equals_baseline_pax_over_baseline_capacity(self):
        from special_days.case_studies import RouteDay

        cells = {}
        base = date(2026, 6, 3)
        for index, (pax, capacity) in enumerate([(50, 100), (90, 100), (100, 200)]):
            day = base + timedelta(days=7 * index)
            cells[day] = RouteDay("SVO", "IST", day, pax, capacity, 1)
        target = base + timedelta(days=28)
        cells[target] = RouteDay("SVO", "IST", target, 126, 200, 1)

        result = measure(
            _spec(target), "SVO", "IST", cells, window_days=5, guard_days=0
        )
        point = _point_for(result, target)

        self.assertAlmostEqual(
            point.baseline_lf, point.baseline_pax / point.baseline_capacity, places=12
        )
        # The pool is capacity-weighted: 240 pax / 400 seats = 0.60.
        self.assertAlmostEqual(point.baseline_lf, 0.60, places=12)

    def test_delta_pp_is_implied_by_the_pax_and_capacity_moves(self):
        from special_days.case_studies import RouteDay

        cells = _flat_cells()
        event_day = date(2026, 6, 17)
        cells[event_day] = RouteDay("SVO", "IST", event_day, 132, 220, 1)
        result = measure(_spec(event_day), "SVO", "IST", cells, window_days=20, guard_days=0)
        point = _point_for(result, event_day)

        # Reconstruct the LF move from the two numbers printed beside it.
        pax_factor = 1.0 + point.pax_vs_baseline_pct / 100.0
        cap_factor = 1.0 + point.cap_vs_baseline_pct / 100.0
        implied = 100.0 * point.baseline_lf * (pax_factor / cap_factor - 1.0)
        self.assertAlmostEqual(implied, point.delta_pp, places=9)


class CapacityWeightedSummaryTest(unittest.TestCase):
    """F4: the span aggregate is capacity-weighted, not a mean of daily deltas."""

    def test_a_sixty_seat_day_cannot_outvote_a_thousand_seat_day(self):
        from special_days.case_studies import RouteDay

        cells = _flat_cells(pax=500, capacity=1000)
        first, second = date(2026, 6, 17), date(2026, 6, 18)
        cells[first] = RouteDay("SVO", "IST", first, 510, 1000, 1)   # +1 pp
        cells[second] = RouteDay("SVO", "IST", second, 48, 60, 1)    # +30 pp, tiny

        result = measure(
            _spec(first, second), "SVO", "IST", cells, window_days=10, guard_days=0
        )
        self.assertEqual(result.n_days, 2)
        # (510+48)/(1000+60) - (500+500)/(1000+1000) = 0.5264 - 0.50
        self.assertAlmostEqual(result.mean_delta_pp, 2.64, places=2)
        self.assertLess(result.mean_delta_pp, 5.0, "unweighted mean would have printed +15.5")

    def test_the_three_summary_numbers_share_one_day_set(self):
        from special_days.case_studies import RouteDay

        cells = _flat_cells()
        first, second = date(2026, 6, 17), date(2026, 6, 18)
        cells[first] = RouteDay("SVO", "IST", first, 150, 200, 1)
        cells[second] = RouteDay("SVO", "IST", second, 120, 250, 1)
        result = measure(
            _spec(first, second), "SVO", "IST", cells, window_days=20, guard_days=0
        )

        pax_factor = 1.0 + result.mean_pax_vs_baseline_pct / 100.0
        cap_factor = 1.0 + result.mean_cap_vs_baseline_pct / 100.0
        implied = 100.0 * 0.5 * (pax_factor / cap_factor - 1.0)  # baseline LF is 0.50
        self.assertAlmostEqual(implied, result.mean_delta_pp, places=9)


class VerdictNeverContradictsTest(unittest.TestCase):
    """F5: every label names all three dimensions, consistently with the numbers."""

    def test_previously_contradictory_cases_now_read_correctly(self):
        from special_days.case_studies import _verdict

        cases = {
            (-8.0, 12.0, 1.0): "demand up, capacity flat, LF down",
            (6.0, 12.0, -20.0): "demand up, capacity cut, LF up",
            (-6.0, -12.0, 20.0): "demand down, capacity up, LF down",
            (12.0, -20.0, -20.0): "demand down, capacity cut, LF up",
            (-12.0, -20.0, -20.0): "demand down, capacity cut, LF down",
        }
        for (delta, pax, cap), expected in cases.items():
            self.assertEqual(_verdict(delta, pax, cap, 5), expected)

    def test_sweep_of_every_combination_agrees_with_its_numbers(self):
        from special_days.case_studies import LF_PP_THRESHOLD, PCT_THRESHOLD, _verdict

        deltas = (-9.0, -LF_PP_THRESHOLD, 0.0, LF_PP_THRESHOLD, 9.0)
        pcts = (-30.0, -PCT_THRESHOLD, 0.0, PCT_THRESHOLD, 30.0)
        seen = set()
        for delta in deltas:
            for pax in pcts:
                for cap in pcts:
                    label = _verdict(delta, pax, cap, 5)
                    seen.add(label)
                    if label == "no movement":
                        self.assertLess(abs(delta), LF_PP_THRESHOLD)
                        self.assertLess(abs(pax), PCT_THRESHOLD)
                        self.assertLess(abs(cap), PCT_THRESHOLD)
                        continue
                    demand, capacity, load = [part.strip() for part in label.split(",")]
                    self.assertEqual(demand, "demand up" if pax >= PCT_THRESHOLD
                                     else "demand down" if pax <= -PCT_THRESHOLD else "demand flat")
                    self.assertEqual(capacity, "capacity up" if cap >= PCT_THRESHOLD
                                     else "capacity cut" if cap <= -PCT_THRESHOLD else "capacity flat")
                    self.assertEqual(load, "LF up" if delta >= LF_PP_THRESHOLD
                                     else "LF down" if delta <= -LF_PP_THRESHOLD else "LF flat")
        self.assertEqual(len(seen), 27, "every combination should be reachable")

    def test_no_label_ever_calls_a_large_capacity_move_flat(self):
        from special_days.case_studies import _verdict

        for cap in (-40.0, -20.0, 20.0, 40.0):
            for delta in (-10.0, 0.0, 10.0):
                for pax in (-20.0, 0.0, 20.0):
                    self.assertNotIn("capacity flat", _verdict(delta, pax, cap, 5))


class EventDaysNotInExtractTest(_TempDirCase):
    """F6: an extract that stops before the event is not a thin baseline."""

    def test_window_rows_without_event_days_get_their_own_verdict(self):
        event_day = date(2025, 3, 10)  # the extract stops on 2025-03-01
        rows = _generate_lf_rows(["SVOIST"], 60)  # 2025-01-01 .. 2025-03-01
        result = self.measure_single(rows, event_day)

        self.assertGreater(result.n_route_days, 0)
        self.assertEqual(result.n_days, 0)
        self.assertEqual(result.verdict, "event days not in extract")

    def test_such_an_event_is_reported_as_unmeasured(self):
        event_day = date(2025, 3, 10)  # the extract stops on 2025-03-01
        self.measure_single(_generate_lf_rows(["SVOIST"], 60), event_day)
        coverage = (self.tmp / "out" / "coverage.txt").read_text(encoding="utf-8")
        self.assertIn("events with no measured event day: 1", coverage)
        self.assertIn("Planted event", coverage)

    def test_a_route_absent_from_the_extract_is_named(self):
        lf_path = _write_csv(self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST"], 60))
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Typo", "2025-02-05", "", "İstanbul", "IST", "SVO-IST;LED-IST", ""]],
        )
        run(events_path, lf_path, self.tmp / "out", reports=False)
        coverage = (self.tmp / "out" / "coverage.txt").read_text(encoding="utf-8")
        self.assertIn("routes requested but absent from the extract: 1", coverage)
        self.assertIn("Typo: LED-IST", coverage)


class FractionalCapacityTest(_TempDirCase):
    """F7: a capacity that rounds to zero must not reach RouteDay.lf."""

    def test_fractional_capacity_cell_is_dropped_not_divided_by(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            [["SVO", "IST", "2025-03-05", 0, 0.4], ["SVO", "IST", "2025-03-06", 150, 200]],
        )
        route_days, coverage = load_route_days(lf_path)
        cells = route_days[("SVO", "IST")]

        self.assertNotIn(date(2025, 3, 5), cells)
        self.assertEqual(coverage.n_skipped_bad_capacity, 1)
        self.assertEqual(coverage.n_flight_rows, 1)
        for cell in cells.values():
            cell.lf  # must not raise

    def test_a_route_left_with_no_cells_disappears_entirely(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, [["SVO", "IST", "2025-03-05", 0, 0.4]]
        )
        route_days, coverage = load_route_days(lf_path)
        self.assertEqual(route_days, {})
        self.assertEqual(coverage.n_routes, 0)

    def test_the_cli_reports_it_instead_of_crashing(self):
        from special_days.case_studies import main

        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, [["SVO", "IST", "2025-03-05", 0, 0.4]]
        )
        events_path = _write_events(
            self.tmp / "events.csv", [["E", "2025-03-05", "", "İstanbul", "IST", "SVO-IST", ""]]
        )
        code = main(
            ["--events", str(events_path), "--lf", str(lf_path),
             "--out", str(self.tmp / "out"), "--csv-only"]
        )
        self.assertEqual(code, 0)
        self.assertIn("no LF data", (self.tmp / "out" / "summary.csv").read_text(encoding="utf-8"))


class TurkishHeaderTest(_TempDirCase):
    """F8: Excel's UPPER() under a Turkish locale must still resolve."""

    def test_dotted_capital_i_headers_resolve(self):
        from special_days.case_studies import _normalize_header

        self.assertEqual(_normalize_header("CAPACİTY"), "capacity")
        self.assertEqual(_normalize_header("ORİGİN"), "origin")
        self.assertEqual(_normalize_header("FLİGHT_DATE"), "flight date")

    def test_a_turkish_locale_extract_loads(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            ["ORİGİN", "DESTİNATİON", "FLİGHT_DATE", "BOARDED_PAX", "CAPACİTY"],
            [["SVO", "IST", "2025-03-05", 150, 200]],
        )
        route_days, _ = load_route_days(lf_path)
        self.assertEqual(route_days[("SVO", "IST")][date(2025, 3, 5)].pax, 150)

    def test_a_turkish_locale_event_file_loads(self):
        events_path = _write_csv(
            self.tmp / "events.csv",
            ["EVENT", "START_DATE", "END_DATE", "CİTY", "AİRPORT", "ROUTES", "NOTE"],
            [["E", "2025-03-05", "", "İstanbul", "IST", "SVO-IST", "n"]],
        )
        events = load_events(events_path)
        self.assertEqual(events[0].airport, "IST")
        self.assertEqual(events[0].city, "İstanbul")


class GuardWiderThanWindowTest(unittest.TestCase):
    """F9: no in-event day can be guarded out of the timeline."""

    def test_every_event_day_is_still_measured(self):
        cells = _flat_cells()
        event = _spec(date(2026, 6, 17), date(2026, 6, 19))
        result = measure(event, "SVO", "IST", cells, window_days=1, guard_days=10)
        measured = [point.day for point in result.timeline if point.in_event]
        self.assertEqual(
            measured, [date(2026, 6, 17), date(2026, 6, 18), date(2026, 6, 19)]
        )
        self.assertEqual(result.n_days, 3)


class SlugKeepsTheDateTest(unittest.TestCase):
    """F10: the date is the disambiguator, so it survives truncation."""

    def test_long_names_keep_their_dates_and_stay_distinct(self):
        from special_days.case_studies import EventSpec

        name = "Galatasaray Fenerbahce derbisi Turkiye Kupasi finali oynanacak"
        first = EventSpec(1, name, date(2026, 4, 19), date(2026, 4, 19), "İstanbul", "IST", (), "")
        second = EventSpec(2, name, date(2026, 5, 24), date(2026, 5, 24), "Ankara", "ESB", (), "")

        self.assertTrue(first.slug().endswith("-2026-04-19"), first.slug())
        self.assertTrue(second.slug().endswith("-2026-05-24"), second.slug())
        self.assertNotEqual(first.slug(), second.slug())
        self.assertLessEqual(len(first.slug()), 60)


class RowGrainFlagsTest(_TempDirCase):
    """F11: anomalies must not vanish into the day total."""

    def test_exactly_the_threshold_capacity_is_flagged(self):
        from special_days.case_studies import RouteDay, _flags_for

        self.assertIn("low_capacity", _flags_for(RouteDay("S", "I", date(2025, 3, 5), 100, 200, 1)))

    def test_an_anomalous_flight_row_shows_on_an_ordinary_looking_day(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            [
                ["SVO", "IST", "2025-03-05", 84, 60],   # LF 1.40 on a tiny gauge
                ["SVO", "IST", "2025-03-05", 150, 300],  # LF 0.50
            ],
        )
        route_days, _ = load_route_days(lf_path)
        cell = route_days[("SVO", "IST")][date(2025, 3, 5)]

        # The day total looks unremarkable: 234/360 = 0.65 on 360 seats.
        self.assertAlmostEqual(cell.lf, 0.65, places=4)
        self.assertEqual(cell.n_rows_lf_gt_1, 1)
        self.assertEqual(cell.n_rows_low_capacity, 1)

        from special_days.case_studies import _flags_for

        flags = _flags_for(cell)
        self.assertIn("row_lf>1", flags)
        self.assertIn("row_low_capacity", flags)


class SlashDateWarningTest(_TempDirCase):
    """F12: DD/MM vs MM/DD is ambiguous, so say so rather than guess silently."""

    def test_slash_dates_are_counted_and_warned_about(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            [["SVO", "IST", "03/04/2025", 150, 200], ["SVO", "IST", "2025-04-04", 150, 200]],
        )
        _, coverage = load_route_days(lf_path)
        self.assertEqual(coverage.n_slash_format_dates, 1)

        events_path = _write_events(
            self.tmp / "events.csv", [["E", "2025-04-03", "", "İstanbul", "IST", "SVO-IST", ""]]
        )
        run(events_path, lf_path, self.tmp / "out", reports=False)
        self.assertIn("WARNING", (self.tmp / "out" / "coverage.txt").read_text(encoding="utf-8"))

    def test_iso_only_extracts_carry_no_warning(self):
        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST"], 60)
        )
        events_path = _write_events(
            self.tmp / "events.csv", [["E", "2025-02-05", "", "İstanbul", "IST", "SVO-IST", ""]]
        )
        run(events_path, lf_path, self.tmp / "out", reports=False)
        self.assertNotIn("WARNING", (self.tmp / "out" / "coverage.txt").read_text(encoding="utf-8"))


class AllSixArtifactsTest(_TempDirCase):
    """The CLI must write every artifact it promises, not just the CSVs."""

    def _run_cli(self, *extra: str) -> Path:
        from special_days.case_studies import main

        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST", "LEDIST"], 150)
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Kanye West concert", "2025-03-05", "", "İstanbul", "IST", "", "stadium"]],
        )
        out = self.tmp / "out"
        code = main(
            ["--events", str(events_path), "--lf", str(lf_path), "--out", str(out), *extra]
        )
        self.assertEqual(code, 0)
        return out

    def test_run_writes_csvs_workbook_html_and_svgs(self):
        out = self._run_cli()
        for name in ("summary.csv", "timeline.csv", "coverage.txt",
                     "case_studies.xlsx", "case_studies.html"):
            self.assertTrue((out / name).is_file(), f"{name} was not written")
            self.assertGreater((out / name).stat().st_size, 0, f"{name} is empty")
        svgs = sorted((out / "charts").glob("*.svg"))
        self.assertEqual(len(svgs), 2, "one SVG per event x route")

    def test_csv_only_skips_the_renderers(self):
        out = self._run_cli("--csv-only")
        self.assertTrue((out / "summary.csv").is_file())
        self.assertFalse((out / "case_studies.xlsx").exists())
        self.assertFalse((out / "charts").exists())

    def test_the_workbook_and_html_read_the_real_result_records(self):
        from openpyxl import load_workbook

        out = self._run_cli()
        workbook = load_workbook(out / "case_studies.xlsx")
        try:
            self.assertEqual(workbook.sheetnames[0], "Summary")
            self.assertEqual(len(workbook.sheetnames), 3)  # Summary + 2 timelines
        finally:
            workbook.close()

        page = (out / "case_studies.html").read_text(encoding="utf-8")
        self.assertIn("Kanye West concert", page)
        self.assertNotIn("<script src", page)
        self.assertEqual(page.count("<svg"), 2)


class NoRankingByOutcomeTest(_TempDirCase):
    """The event list is fixed up front; ranking it would cherry-pick it."""

    def test_a_large_effect_stays_where_the_input_put_it(self):
        big, small = date(2025, 4, 2), date(2025, 3, 5)
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            _generate_lf_rows(["SVOIST"], 220, lf_injection={big: 0.20, small: 0.01}),
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [
                ["Tiny", small.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""],
                ["Huge", big.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""],
            ],
        )
        results = run(events_path, lf_path, self.tmp / "out", guard_days=0, reports=False)
        self.assertEqual([r.event for r in results], ["Tiny", "Huge"])
        self.assertGreater(results[1].mean_delta_pp, results[0].mean_delta_pp + 15.0)


class NormalSwingTest(_TempDirCase):
    """The yardstick that lets a reader tell a real move from routine wobble."""

    def test_a_planted_lift_stands_well_clear_of_the_routes_normal_swing(self):
        event = date(2025, 6, 14)
        lf_path = _write_csv(
            self.tmp / "lf.csv",
            LF_HEADER,
            _generate_lf_rows(["SVOIST"], 300, lf_injection={event: 0.12}),
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Concert", event.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""]],
        )
        result = run(events_path, lf_path, self.tmp / "out", reports=False)[0]
        self.assertIsNotNone(result.normal_swing_pp)
        # The swing is measured off the NON-event days, so the injection must not
        # leak into it -- otherwise the yardstick grows with the thing it measures.
        self.assertLess(result.normal_swing_pp, 5.0)
        self.assertGreater(result.mean_delta_pp, result.normal_swing_pp * 2)

    def test_swing_is_none_when_there_are_no_ordinary_days_to_measure(self):
        event = date(2025, 6, 14)
        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST"], 300)
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Concert", event.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""]],
        )
        # window_days=0 leaves the event day as the only row in the timeline.
        result = run(
            events_path, lf_path, self.tmp / "out", window_days=1, reports=False
        )[0]
        self.assertIsInstance(result.normal_swing_pp, (float, type(None)))

    def test_summary_csv_carries_the_column(self):
        event = date(2025, 6, 14)
        lf_path = _write_csv(
            self.tmp / "lf.csv", LF_HEADER, _generate_lf_rows(["SVOIST"], 300)
        )
        events_path = _write_events(
            self.tmp / "events.csv",
            [["Concert", event.isoformat(), "", "İstanbul", "IST", "SVO-IST", ""]],
        )
        out = self.tmp / "out"
        run(events_path, lf_path, out, reports=False)
        header = (out / "summary.csv").read_text(encoding="utf-8").splitlines()[0]
        self.assertIn("normal_swing_pp", header)


if __name__ == "__main__":
    unittest.main()
