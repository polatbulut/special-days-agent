"""Tests for the case-study report writers.

``special_days.case_studies`` is not imported at runtime by the writers (the
result types are structural), so the fixtures below stand in for it.
"""

import os
import tempfile
import unittest
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, timedelta

from openpyxl import load_workbook

from special_days.case_report import (
    SUMMARY_HEADERS,
    sheet_title,
    slug,
    svg_filename,
    write_html,
    write_svgs,
    write_workbook,
)


@dataclass(frozen=True)
class DayPoint:
    date: date
    lf: float | None
    baseline: float | None
    delta_pp: float | None
    pax: int
    capacity: int
    pax_vs_baseline_pct: float | None
    cap_vs_baseline_pct: float | None
    in_event: bool
    flags: str = ""


@dataclass(frozen=True)
class EventRouteResult:
    event: str
    note: str
    start_date: date
    end_date: date
    route: str
    org: str
    dst: str
    mean_delta_pp: float | None
    mean_pax_pct: float | None
    mean_cap_pct: float | None
    n_days: int
    n_days_no_baseline: int
    verdict: str
    timeline: list = field(default_factory=list)
    normal_swing_pp: float | None = None


def day(offset, lf=0.82, baseline=0.78, in_event=False, flags=""):
    d = date(2026, 7, 10) + timedelta(days=offset)
    delta = None if (lf is None or baseline is None) else (lf - baseline) * 100.0
    return DayPoint(
        date=d, lf=lf, baseline=baseline, delta_pp=delta, pax=160, capacity=196,
        pax_vs_baseline_pct=3.4 if lf is not None else None,
        cap_vs_baseline_pct=1.1 if lf is not None else None,
        in_event=in_event, flags=flags,
    )


def make(event="Kanye West Istanbul", route="SVO-IST", timeline=None,
         delta=4.12, note="concert", verdict="LF rose above baseline"):
    if timeline is None:
        timeline = [day(i, in_event=(i in (3, 4))) for i in range(9)]
    return EventRouteResult(
        event=event, note=note, start_date=date(2026, 7, 13), end_date=date(2026, 7, 14),
        route=route, org=route.split("-")[0], dst=route.split("-")[-1],
        mean_delta_pp=delta, mean_pax_pct=6.2, mean_cap_pct=1.9,
        n_days=len(timeline), n_days_no_baseline=0, verdict=verdict,
        timeline=timeline, normal_swing_pp=1.4,
    )


class SheetNameTest(unittest.TestCase):
    def test_long_name_is_truncated_to_31(self):
        name = sheet_title("E" * 60, "SVO-IST")
        self.assertLessEqual(len(name), 31)
        self.assertTrue(name.endswith("SVO-IST"))

    def test_forbidden_characters_are_stripped(self):
        name = sheet_title("Rock/Pop: Live [2026]?*\\", "SVO-IST")
        self.assertFalse(set(name) & set("[]:*?/\\"))
        self.assertLessEqual(len(name), 31)

    def test_collisions_are_unique(self):
        used = []
        for _ in range(3):
            used.append(sheet_title("E" * 60, "SVO-IST", used))
        self.assertEqual(len(set(used)), 3)
        self.assertTrue(all(len(n) <= 31 for n in used))

    def test_slug_and_filename(self):
        self.assertEqual(slug("Kanye West İstanbul"), "kanye_west_istanbul")
        self.assertEqual(svg_filename("Kanye West", "SVO-IST"), "kanye_west_svo-ist.svg")
        used = ["kanye_west_svo-ist.svg"]
        self.assertEqual(
            svg_filename("Kanye West", "SVO-IST", used), "kanye_west_svo-ist_2.svg"
        )


class WorkbookTest(unittest.TestCase):
    def _write(self, results):
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "case_studies.xlsx")
        write_workbook(results, path)
        return path

    def test_summary_sheet_header_rows_and_freeze(self):
        results = [make(), make(event="Tarkan Live", route="AYT-IST")]
        book = load_workbook(self._write(results))
        self.assertEqual(book.sheetnames[0], "Summary")
        self.assertEqual(len(book.sheetnames), 3)  # Summary + one per result

        summary = book["Summary"]
        self.assertEqual([c.value for c in summary[1]], list(SUMMARY_HEADERS))
        self.assertTrue(all(c.font.bold for c in summary[1]))
        self.assertEqual(summary.freeze_panes, "A2")
        self.assertEqual(summary.auto_filter.ref, "A1:L3")  # +1 col: Normal swing pp
        self.assertEqual(summary.max_row, len(results) + 1)

    def test_summary_values_are_numeric_and_dates_real(self):
        book = load_workbook(self._write([make()]))
        summary = book["Summary"]
        self.assertEqual(summary["A2"].value, "Kanye West Istanbul")
        self.assertEqual(summary["E2"].value, "SVO-IST")
        self.assertIsInstance(summary["F2"].value, float)
        self.assertEqual(summary["G2"].value, 1.4)  # Normal swing pp
        self.assertEqual(summary["J2"].value, 9)    # Days (shifted by the new col)
        start = summary["C2"].value
        self.assertEqual((start.year, start.month, start.day), (2026, 7, 13))
        self.assertEqual(summary["C2"].number_format, "yyyy-mm-dd")

    def test_input_order_is_preserved(self):
        results = [make(delta=-3.0, event="B event"), make(delta=9.0, event="A event")]
        summary = load_workbook(self._write(results))["Summary"]
        self.assertEqual(
            [summary.cell(row=r, column=1).value for r in (2, 3)],
            ["B event", "A event"],
        )

    def test_timeline_sheet_has_rows_and_chart(self):
        book = load_workbook(self._write([make()]))
        sheet = book[book.sheetnames[1]]
        self.assertEqual(sheet.max_row, 10)  # header + 9 days
        self.assertEqual(sheet["I5"].value, 1)  # in_event flag on day index 3
        self.assertEqual(sheet["I2"].value, 0)
        self.assertEqual(len(sheet._charts), 1)

    def test_duplicate_event_route_sheets_are_unique(self):
        book = load_workbook(self._write([make(), make(), make()]))
        self.assertEqual(len(set(book.sheetnames)), 4)

    def test_long_and_illegal_names_survive_a_real_save(self):
        results = [make(event="E" * 60), make(event="Rock/Pop: Live?")]
        names = load_workbook(self._write(results)).sheetnames
        self.assertTrue(all(len(n) <= 31 for n in names))
        self.assertFalse(any(set(n) & set("[]:*?/\\") for n in names))

    def test_empty_results(self):
        book = load_workbook(self._write([]))
        self.assertEqual(book.sheetnames, ["Summary"])
        self.assertEqual(book["Summary"].max_row, 1)


class HtmlTest(unittest.TestCase):
    def _write(self, results):
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "case_studies.html")
        write_html(results, path)
        with open(path, encoding="utf-8") as handle:
            return handle.read()

    def test_self_contained(self):
        text = self._write([make(), make(event="Tarkan Live", route="AYT-IST")])
        self.assertNotIn("http://", text)
        self.assertNotIn("https://", text)
        self.assertNotIn("<script src", text)
        self.assertNotIn("<link ", text)
        self.assertIn("prefers-color-scheme:dark", text)  # renders in both themes
        self.assertIn("<style>", text)

    def test_one_card_per_result(self):
        text = self._write([make(), make(event="Tarkan Live", route="AYT-IST")])
        self.assertEqual(text.count('<section class="card"'), 2)
        self.assertEqual(text.count("<svg"), 2)
        self.assertIn("Tarkan Live", text)
        self.assertIn("Kanye West Istanbul", text)

    def test_summary_table_in_input_order(self):
        text = self._write([make(event="B event"), make(event="A event")])
        self.assertLess(text.index("B event"), text.index("A event"))

    def test_markup_is_escaped(self):
        text = self._write([make(event="Rock & Roll <b>")])
        self.assertIn("Rock &amp; Roll &lt;b&gt;", text)
        self.assertNotIn("Roll <b>", text)

    def test_empty_results(self):
        text = self._write([])
        self.assertIn("No case studies to report.", text)
        self.assertNotIn("<section class=\"card\"", text)


class SvgTest(unittest.TestCase):
    def _write(self, results):
        tmp = tempfile.mkdtemp()
        return write_svgs(results, os.path.join(tmp, "charts"))

    def test_files_are_well_formed_and_named(self):
        paths = self._write([make(), make(event="Tarkan Live", route="AYT-IST")])
        self.assertEqual(len(paths), 2)
        for path, event in zip(paths, ["Kanye West Istanbul", "Tarkan Live"]):
            self.assertTrue(os.path.exists(path))
            root = ET.parse(path).getroot()
            self.assertTrue(root.tag.endswith("svg"))
            with open(path, encoding="utf-8") as handle:
                self.assertIn(event, handle.read())
        self.assertTrue(paths[0].endswith("kanye_west_istanbul_svo-ist.svg"))

    def test_duplicate_names_do_not_overwrite(self):
        paths = self._write([make(), make()])
        self.assertEqual(len(set(paths)), 2)

    def test_event_days_are_shaded(self):
        path = self._write([make()])[0]
        with open(path, encoding="utf-8") as handle:
            self.assertIn('class="band"', handle.read())

    def test_empty_results_creates_dir_and_no_files(self):
        tmp = tempfile.mkdtemp()
        out = os.path.join(tmp, "charts")
        self.assertEqual(write_svgs([], out), [])
        self.assertTrue(os.path.isdir(out))
        self.assertEqual(os.listdir(out), [])


class MissingDataTest(unittest.TestCase):
    """None LFs, empty timelines and odd values must not raise in any writer."""

    def _cases(self):
        gappy = [
            day(0),
            day(1, lf=None, baseline=None),
            day(2, lf=None, baseline=0.80),
            day(3, in_event=True),
            day(4, lf=1.04, baseline=0.99, in_event=True, flags="lf_gt_1"),
            day(5, lf=None, baseline=None),
            day(6),
        ]
        return [
            make(timeline=gappy, delta=None),
            make(event="No timeline", timeline=[], delta=None, verdict="no data"),
            make(event="Single day", timeline=[day(0, in_event=True)], delta=0.0),
            make(event="Flat", timeline=[day(i, lf=0.8, baseline=0.8) for i in range(4)],
                 delta=0.0),
        ]

    def test_all_writers_tolerate_none(self):
        results = self._cases()
        tmp = tempfile.mkdtemp()

        xlsx = os.path.join(tmp, "c.xlsx")
        write_workbook(results, xlsx)
        book = load_workbook(xlsx)
        self.assertEqual(len(book.sheetnames), len(results) + 1)
        self.assertIsNone(book["Summary"]["F2"].value)  # mean_delta_pp = None

        page = os.path.join(tmp, "c.html")
        write_html(results, page)
        with open(page, encoding="utf-8") as handle:
            text = handle.read()
        self.assertEqual(text.count('<section class="card"'), len(results))
        self.assertIn("n/a", text)
        self.assertIn("no data", text)

        paths = write_svgs(results, os.path.join(tmp, "charts"))
        self.assertEqual(len(paths), len(results))
        for path in paths:
            ET.parse(path)

    def test_gap_breaks_the_line_rather_than_plotting_zero(self):
        gappy = [day(0), day(1), day(2, lf=None, baseline=None), day(3), day(4)]
        path = write_svgs([make(timeline=gappy)], tempfile.mkdtemp())[0]
        root = ET.parse(path).getroot()
        ns = "{http://www.w3.org/2000/svg}"
        lf_lines = [
            el for el in root.iter(f"{ns}polyline")
            if "lf" in (el.get("class") or "").split()
        ]
        self.assertEqual(len(lf_lines), 2)  # one polyline either side of the gap
        for element in lf_lines:
            self.assertEqual(len(element.get("points").split()), 2)


class RealResultTypesTest(unittest.TestCase):
    """The writers must accept the actual records from special_days.case_studies."""

    def _result(self):
        from special_days.case_studies import DayPoint as Real
        from special_days.case_studies import EventRouteResult as RealResult
        from special_days.case_studies import EventSpec

        spec = EventSpec(
            row_id=1, event="Kanye West İstanbul", start_date=date(2026, 7, 14),
            end_date=date(2026, 7, 15), city="İstanbul", airport="IST",
            routes=("SVO-IST",), note="stadium concert",
        )
        lfs = [0.71, 0.74, 0.73, 0.79, 0.88, 1.04, 0.91, 0.80, 0.76]
        timeline = [
            Real(
                day=date(2026, 7, 10) + timedelta(days=i), weekday="Fri",
                in_event=(i in (4, 5)), lf=lfs[i],
                baseline_lf=(None if i == 2 else 0.78),
                delta_pp=(None if i == 2 else (lfs[i] - 0.78) * 100),
                pax=160, capacity=196, baseline_pax=150.0, baseline_capacity=190.0,
                pax_vs_baseline_pct=6.6, cap_vs_baseline_pct=3.1,
                n_baseline_obs=4, n_flights=2,
                data_flags=("lf_gt_1" if lfs[i] > 1 else ""),
            )
            for i in range(9)
        ]
        return RealResult(
            spec=spec, org="SVO", dst="IST", timeline=timeline, mean_delta_pp=4.9,
            mean_pax_vs_baseline_pct=6.6, mean_cap_vs_baseline_pct=3.1,
            n_days=9, n_days_no_baseline=1, n_route_days=9, n_flights=18,
            verdict="LF rose 4.9 pp above baseline",
        )

    def test_all_three_writers_accept_real_records(self):
        results = [self._result()]
        tmp = tempfile.mkdtemp()

        xlsx = os.path.join(tmp, "case_studies.xlsx")
        write_workbook(results, xlsx)
        book = load_workbook(xlsx)
        self.assertEqual(len(book.sheetnames), 2)
        row = [c.value for c in book["Summary"][2]]
        self.assertEqual(row[0], "Kanye West İstanbul")
        self.assertEqual(row[4], "SVO-IST")
        self.assertEqual(row[5], 4.9)
        self.assertEqual(book[book.sheetnames[1]].max_row, 10)

        page = os.path.join(tmp, "case_studies.html")
        write_html(results, page)
        with open(page, encoding="utf-8") as handle:
            text = handle.read()
        self.assertEqual(text.count('<section class="card"'), 1)
        self.assertNotIn("https://", text)

        paths = write_svgs(results, os.path.join(tmp, "charts"))
        ET.parse(paths[0])
        self.assertTrue(paths[0].endswith("kanye_west_istanbul_svo-ist.svg"))


if __name__ == "__main__":
    unittest.main()
