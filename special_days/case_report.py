"""Report writers for the event/route case studies: Excel, HTML and SVG.

Consumes the ``EventRouteResult`` / ``DayPoint`` records produced by
:mod:`special_days.case_studies` and renders them three ways:

* :func:`write_workbook` -- an ``.xlsx`` with a ``Summary`` sheet plus one
  timeline sheet per event x route, each carrying a native line chart.
* :func:`write_html` -- one self-contained page (no CDN, no external CSS,
  no JS) with a card and an inline SVG chart per event x route.
* :func:`write_svgs` -- one standalone, hand-built ``.svg`` per event x route.

Input order is preserved everywhere. Nothing here sorts, ranks or filters:
flat and negative results are reported exactly as they arrive.

The result types are imported from :mod:`special_days.case_studies` for typing
only. The writers read a denormalized point (``date``/``baseline``/``flags``,
``event``/``route``/``mean_pax_pct``), which the real records expose as flat
aliases, so nothing here reaches through two levels and the writers can also be
exercised against a stub.
"""

from __future__ import annotations

import html
import os
import unicodedata
from datetime import date
from typing import TYPE_CHECKING, Iterable, Sequence

if TYPE_CHECKING:  # pragma: no cover - typing only
    from .case_studies import DayPoint, EventRouteResult

SUMMARY_SHEET = "Summary"
SUMMARY_HEADERS = [
    "Event",
    "Note",
    "Start date",
    "End date",
    "Route",
    "Mean delta pp",
    "Normal swing pp",
    "Mean pax %",
    "Mean cap %",
    "Days",
    "Days no baseline",
    "Verdict",
]
TIMELINE_HEADERS = [
    "Date",
    "LF",
    "Baseline LF",
    "Delta pp",
    "Pax",
    "Capacity",
    "Pax vs baseline %",
    "Cap vs baseline %",
    "In event",
    "Flags",
]

_SUMMARY_WIDTHS = {
    "A": 34, "B": 40, "C": 12, "D": 12, "E": 10, "F": 14,
    "G": 12, "H": 12, "I": 8, "J": 17, "K": 26,
}
_TIMELINE_WIDTHS = {
    "A": 12, "B": 10, "C": 12, "D": 10, "E": 10, "F": 10,
    "G": 18, "H": 18, "I": 9, "J": 14,
}

_SHEET_FORBIDDEN = set("[]:*?/\\")
_SHEET_MAX = 31
_TURKISH = str.maketrans({"ı": "i", "İ": "I", "ş": "s", "Ş": "S",
                          "ğ": "g", "Ğ": "G", "ç": "c", "Ç": "C",
                          "ö": "o", "Ö": "O", "ü": "u", "Ü": "U"})

# Chart geometry, shared by the inline and the standalone SVG.
_W, _H = 760.0, 320.0
_PAD_L, _PAD_R, _PAD_T, _PAD_B = 58.0, 18.0, 46.0, 46.0


# --------------------------------------------------------------------------
# naming
# --------------------------------------------------------------------------

def sheet_title(event: str, route: str, used: Iterable[str] = ()) -> str:
    """Return a legal, unique Excel sheet name for ``event`` x ``route``."""
    taken = {t.casefold() for t in used}
    route_part = _strip_sheet_chars(route)
    event_part = _strip_sheet_chars(event) or "Event"

    def build(base_event: str, suffix: str) -> str:
        room = _SHEET_MAX - len(suffix) - (len(route_part) + 1 if route_part else 0)
        head = base_event[:max(room, 1)].strip() or "Event"
        name = f"{head} {route_part}" if route_part else head
        return (name + suffix)[:_SHEET_MAX].strip() or "Event"

    candidate = build(event_part, "")
    counter = 2
    while candidate.casefold() in taken:
        candidate = build(event_part, f"~{counter}")
        counter += 1
    return candidate


def _strip_sheet_chars(text: str) -> str:
    cleaned = "".join(" " if ch in _SHEET_FORBIDDEN else ch for ch in text)
    return " ".join(cleaned.split()).strip("'")


def slug(text: str) -> str:
    """Lowercase ASCII slug safe for a filename."""
    folded = unicodedata.normalize("NFKD", text.translate(_TURKISH))
    ascii_text = folded.encode("ascii", "ignore").decode("ascii")
    out = "".join(ch if (ch.isalnum() or ch == "-") else "_" for ch in ascii_text.lower())
    return "_".join(part for part in out.split("_") if part) or "event"


def svg_filename(event: str, route: str, used: Iterable[str] = ()) -> str:
    """Return a unique ``<event>_<route>.svg`` filename."""
    taken = {n.casefold() for n in used}
    base = f"{slug(event)}_{slug(route)}"[:90]
    name = f"{base}.svg"
    counter = 2
    while name.casefold() in taken:
        name = f"{base}_{counter}.svg"
        counter += 1
    return name


# --------------------------------------------------------------------------
# formatting
# --------------------------------------------------------------------------

def _fmt_pp(value: float | None) -> str:
    return "n/a" if value is None else f"{value:+.2f} pp"


def _fmt_pct(value: float | None) -> str:
    return "n/a" if value is None else f"{value:+.2f}%"


def _iso(value: date | None) -> str:
    return "" if value is None else value.isoformat()


def _ensure_parent(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    os.makedirs(parent, exist_ok=True)


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------

def write_workbook(results: Sequence[EventRouteResult], path: str) -> None:
    """Write ``results`` to an ``.xlsx`` workbook at ``path``.

    Sheet ``Summary`` holds one row per event x route in the given order; each
    result also gets its own timeline sheet with a native line chart.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    summary = workbook.active
    summary.title = SUMMARY_SHEET
    summary.append(list(SUMMARY_HEADERS))
    for cell in summary[1]:
        cell.font = Font(bold=True)

    names: list[str] = []
    for result in results:
        names.append(sheet_title(result.event, result.route, names))
        summary.append(
            [
                result.event,
                result.note,
                result.start_date,  # real date cell
                result.end_date,  # real date cell
                result.route,
                result.mean_delta_pp,  # numeric cell
                result.normal_swing_pp,  # the route's ordinary wobble, for scale
                result.mean_pax_pct,  # numeric cell
                result.mean_cap_pct,  # numeric cell
                result.n_days,
                result.n_days_no_baseline,
                result.verdict,
            ]
        )

    for row in summary.iter_rows(min_row=2):
        row[2].number_format = "yyyy-mm-dd"
        row[3].number_format = "yyyy-mm-dd"
        for index in (5, 6, 7):
            row[index].number_format = "0.00"

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = (
        f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{summary.max_row}"
    )
    for column, width in _SUMMARY_WIDTHS.items():
        summary.column_dimensions[column].width = width

    for result, name in zip(results, names):
        _write_timeline_sheet(workbook, result, name)

    _ensure_parent(path)
    workbook.save(path)


def _write_timeline_sheet(workbook, result: EventRouteResult, name: str) -> None:
    from openpyxl.styles import Font

    sheet = workbook.create_sheet(title=name)
    sheet.append(list(TIMELINE_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for point in result.timeline:
        sheet.append(
            [
                point.date,  # real date cell
                point.lf,
                point.baseline,
                point.delta_pp,
                point.pax,
                point.capacity,
                point.pax_vs_baseline_pct,
                point.cap_vs_baseline_pct,
                1 if point.in_event else 0,
                point.flags,
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd"
        row[1].number_format = "0.0%"
        row[2].number_format = "0.0%"
        for index in (3, 6, 7):
            row[index].number_format = "0.00"

    sheet.freeze_panes = "A2"
    for column, width in _TIMELINE_WIDTHS.items():
        sheet.column_dimensions[column].width = width

    if result.timeline:
        _add_line_chart(sheet, result, len(result.timeline))
    else:
        # The SVG and the HTML card both render a "no data" placeholder here;
        # a header-only sheet would make the three renderers disagree about
        # what a null looks like.
        sheet["A2"] = "no data: the extract had no rows for this route in the window"


def _add_line_chart(sheet, result: EventRouteResult, n_rows: int) -> None:
    """Attach a LF-vs-baseline line chart with in-event on a secondary axis."""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import DateAxis

    last = n_rows + 1
    chart = LineChart()
    chart.title = f"{result.event} - {result.route}"
    chart.style = 2
    chart.height = 9
    chart.width = 22
    chart.y_axis.title = "Load factor"
    chart.y_axis.number_format = "0%"
    low, high, _ = _y_bounds(result)  # keep the axis tight around the signal
    chart.y_axis.scaling.min = low / 100.0
    chart.y_axis.scaling.max = high / 100.0
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.number_format = "yyyy-mm-dd"
    chart.x_axis.majorTimeUnit = "days"
    chart.x_axis.title = "Date"

    chart.add_data(
        Reference(sheet, min_col=2, max_col=3, min_row=1, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=last))

    # The 0/1 in-event flag rides a secondary axis on the right, so the event
    # span is visible without rescaling the load-factor axis.
    marker = LineChart()
    marker.add_data(
        Reference(sheet, min_col=9, min_row=1, max_row=last), titles_from_data=True
    )
    marker.y_axis.axId = 200
    marker.y_axis.title = "In event"
    marker.y_axis.scaling.min = 0
    marker.y_axis.scaling.max = 1
    marker.y_axis.majorGridlines = None
    marker.y_axis.crosses = "max"  # secondary axis on the right, LF stays left
    marker.x_axis.delete = True  # hide the duplicate category axis
    chart += marker

    # Straight segments only: a smoothed spline would invent LF between days.
    for series in chart.series:
        series.smooth = False

    sheet.add_chart(chart, "L2")


# --------------------------------------------------------------------------
# SVG
# --------------------------------------------------------------------------

_NICE_STEPS = (1.0, 2.0, 2.5, 5.0, 10.0, 20.0, 25.0, 50.0, 100.0)


def _y_bounds(result: EventRouteResult) -> tuple[float, float, float]:
    """Return (min, max, tick step) load factor in percent, on round numbers."""
    values = [
        v * 100.0
        for point in result.timeline
        for v in (point.lf, point.baseline)
        if v is not None
    ]
    if not values:
        return 0.0, 100.0, 25.0
    low, high = min(values), max(values)
    if high - low < 8.0:  # a flat series still needs a readable band
        mid = (low + high) / 2.0
        low, high = mid - 4.0, mid + 4.0
    pad = (high - low) * 0.12
    low, high = low - pad, high + pad

    target = (high - low) / 4.0
    step = next((s for s in _NICE_STEPS if s >= target), _NICE_STEPS[-1])
    low = step * (low // step)
    high = step * -(-high // step)
    while (high - low) / step < 3.0:  # keep at least four gridlines
        high += step
    return low, high, step


def _x_positions(count: int) -> tuple[list[float], float]:
    plot_w = _W - _PAD_L - _PAD_R
    if count <= 0:
        return [], plot_w
    if count == 1:
        return [_PAD_L + plot_w / 2.0], plot_w / 2.0
    step = plot_w / (count - 1)
    return [_PAD_L + index * step for index in range(count)], step


def _event_runs(timeline: Sequence[DayPoint]) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start: int | None = None
    for index, point in enumerate(timeline):
        if point.in_event and start is None:
            start = index
        elif not point.in_event and start is not None:
            runs.append((start, index - 1))
            start = None
    if start is not None:
        runs.append((start, len(timeline) - 1))
    return runs


def _segments(
    values: Sequence[float | None], xs: Sequence[float], to_y
) -> tuple[list[list[tuple[float, float]]], list[tuple[float, float]]]:
    """Split ``values`` into contiguous runs; ``None`` breaks the line."""
    runs: list[list[tuple[float, float]]] = []
    current: list[tuple[float, float]] = []
    for index, value in enumerate(values):
        if value is None:
            if current:
                runs.append(current)
                current = []
            continue
        current.append((xs[index], to_y(value * 100.0)))
    if current:
        runs.append(current)
    lines = [run for run in runs if len(run) >= 2]
    dots = [run[0] for run in runs if len(run) == 1]
    return lines, dots


def _svg_markup(result: EventRouteResult, standalone: bool) -> str:
    """Build the chart markup. ``standalone`` adds xmlns, styles and a backdrop."""
    esc = html.escape
    timeline = list(result.timeline)
    low, high, tick_step = _y_bounds(result)
    span = high - low
    plot_h = _H - _PAD_T - _PAD_B
    bottom = _H - _PAD_B

    def to_y(percent: float) -> float:
        return bottom - (percent - low) / span * plot_h

    xs, x_step = _x_positions(len(timeline))
    parts: list[str] = []
    open_tag = (
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {_W:.0f} {_H:.0f}" '
        f'width="{_W:.0f}" height="{_H:.0f}" role="img">'
        if standalone
        else (
            f'<svg viewBox="0 0 {_W:.0f} {_H:.0f}" preserveAspectRatio="xMidYMid meet" '
            f'class="chart" role="img">'
        )
    )
    parts.append(open_tag)
    parts.append(f"<title>{esc(result.event)} - {esc(result.route)}</title>")
    if standalone:
        parts.append(f"<style>{_SVG_STYLE}</style>")
        parts.append(f'<rect x="0" y="0" width="{_W:.0f}" height="{_H:.0f}" class="bg"/>')

    parts.append(
        f'<text x="{_PAD_L:.1f}" y="20" class="ttl">'
        f"{esc(result.event)} &#183; {esc(result.route)}</text>"
    )
    parts.append(
        f'<text x="{_PAD_L:.1f}" y="35" class="sub">'
        f"{esc(_iso(result.start_date))} to {esc(_iso(result.end_date))} &#183; "
        f"mean {esc(_fmt_pp(result.mean_delta_pp))} &#183; "
        f"pax {esc(_fmt_pct(result.mean_pax_pct))} &#183; "
        f"cap {esc(_fmt_pct(result.mean_cap_pct))}</text>"
    )

    if not timeline:
        parts.append(
            f'<text x="{_W / 2:.1f}" y="{_H / 2:.1f}" class="sub" '
            f'text-anchor="middle">no data</text>'
        )
        parts.append("</svg>")
        return "".join(parts)

    # event shading
    for first, last in _event_runs(timeline):
        x0 = max(_PAD_L, xs[first] - x_step / 2.0)
        x1 = min(_W - _PAD_R, xs[last] + x_step / 2.0)
        parts.append(
            f'<rect x="{x0:.1f}" y="{_PAD_T:.1f}" width="{max(x1 - x0, 1.0):.1f}" '
            f'height="{plot_h:.1f}" class="band"/>'
        )

    # y gridlines + labels, on round multiples of the tick step
    decimals = 0 if tick_step >= 1.0 and tick_step == int(tick_step) else 1
    for tick in range(int(round(span / tick_step)) + 1):
        value = low + tick_step * tick
        y = to_y(value)
        parts.append(
            f'<line x1="{_PAD_L:.1f}" y1="{y:.1f}" x2="{_W - _PAD_R:.1f}" '
            f'y2="{y:.1f}" class="grid"/>'
        )
        parts.append(
            f'<text x="{_PAD_L - 8:.1f}" y="{y + 4:.1f}" class="tick" '
            f'text-anchor="end">{value:.{decimals}f}%</text>'
        )

    # x labels (subsampled so they never collide)
    every = max(1, -(-len(timeline) // 8))
    for index, point in enumerate(timeline):
        if index % every and index != len(timeline) - 1:
            continue
        parts.append(
            f'<text x="{xs[index]:.1f}" y="{bottom + 18:.1f}" class="tick" '
            f'text-anchor="middle">{point.date.strftime("%m-%d")}</text>'
        )

    # axes
    parts.append(
        f'<line x1="{_PAD_L:.1f}" y1="{bottom:.1f}" x2="{_W - _PAD_R:.1f}" '
        f'y2="{bottom:.1f}" class="axis"/>'
    )
    parts.append(
        f'<line x1="{_PAD_L:.1f}" y1="{_PAD_T:.1f}" x2="{_PAD_L:.1f}" '
        f'y2="{bottom:.1f}" class="axis"/>'
    )
    parts.append(
        f'<text x="{(_PAD_L + _W - _PAD_R) / 2:.1f}" y="{_H - 8:.1f}" class="tick" '
        f'text-anchor="middle">Date</text>'
    )
    parts.append(
        f'<text transform="translate(14 {_PAD_T + plot_h / 2:.1f}) rotate(-90)" '
        f'class="tick" text-anchor="middle">Load factor</text>'
    )

    for values, css in (
        ([p.baseline for p in timeline], "bl"),
        ([p.lf for p in timeline], "lf"),
    ):
        lines, dots = _segments(values, xs, to_y)
        for run in lines:
            points = " ".join(f"{x:.1f},{y:.1f}" for x, y in run)
            parts.append(f'<polyline points="{points}" class="line {css}"/>')
        for x, y in dots:
            parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.6" class="dot {css}"/>')

    # legend
    legend_x = _W - _PAD_R - 250.0
    parts.append(
        f'<rect x="{legend_x:.1f}" y="10" width="14" height="3" class="line lf swatch"/>'
    )
    parts.append(f'<text x="{legend_x + 20:.1f}" y="16" class="tick">Actual LF</text>')
    parts.append(
        f'<rect x="{legend_x + 90:.1f}" y="10" width="14" height="3" '
        f'class="line bl swatch"/>'
    )
    parts.append(f'<text x="{legend_x + 110:.1f}" y="16" class="tick">Baseline</text>')
    parts.append(
        f'<rect x="{legend_x + 180:.1f}" y="7" width="14" height="10" class="band"/>'
    )
    parts.append(f'<text x="{legend_x + 200:.1f}" y="16" class="tick">Event</text>')

    parts.append("</svg>")
    return "".join(parts)


_SVG_STYLE = """
svg{--bg:#ffffff;--fg:#1b1f24;--muted:#5b636e;--grid:#e3e6ea;
--lf:#1d4ed8;--bl:#b45309;--band:rgba(250,204,21,.30)}
@media (prefers-color-scheme:dark){svg{--bg:#14181d;--fg:#e7eaee;--muted:#9aa4b0;
--grid:#2b3138;--lf:#7aa2ff;--bl:#f0b429;--band:rgba(250,204,21,.18)}}
.bg{fill:var(--bg)}
.ttl{fill:var(--fg);font:600 14px system-ui,sans-serif}
.sub{fill:var(--muted);font:11px system-ui,sans-serif}
.tick{fill:var(--muted);font:10px system-ui,sans-serif}
.grid{stroke:var(--grid);stroke-width:1}
.axis{stroke:var(--muted);stroke-width:1}
.band{fill:var(--band)}
.line{fill:none;stroke-width:2}
.line.lf{stroke:var(--lf)}
.line.bl{stroke:var(--bl);stroke-dasharray:5 4}
.swatch{stroke:none}
.swatch.lf{fill:var(--lf)}
.swatch.bl{fill:var(--bl)}
.dot{stroke:none}
.dot.lf{fill:var(--lf)}
.dot.bl{fill:var(--bl)}
""".strip()


def write_svgs(results: Sequence[EventRouteResult], out_dir: str) -> list[str]:
    """Write one standalone SVG per result into ``out_dir``; return the paths."""
    os.makedirs(out_dir, exist_ok=True)
    names: list[str] = []
    paths: list[str] = []
    for result in results:
        name = svg_filename(result.event, result.route, names)
        names.append(name)
        path = os.path.join(out_dir, name)
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(_svg_markup(result, standalone=True))
            handle.write("\n")
        paths.append(path)
    return paths


# --------------------------------------------------------------------------
# HTML
# --------------------------------------------------------------------------

_PAGE_STYLE = """
*{box-sizing:border-box}
:root{--bg:#ffffff;--fg:#1b1f24;--muted:#5b636e;--line:#e3e6ea;--card:#fafbfc;
--pos:#15803d;--neg:#b91c1c}
@media (prefers-color-scheme:dark){:root{--bg:#14181d;--fg:#e7eaee;--muted:#9aa4b0;
--line:#2b3138;--card:#1a1f25;--pos:#4ade80;--neg:#f87171}}
html,body{max-width:100%;overflow-x:hidden}
body{margin:0;padding:24px 18px 56px;background:var(--bg);color:var(--fg);
font:14px/1.5 system-ui,-apple-system,Segoe UI,Roboto,sans-serif}
main{max-width:1000px;margin:0 auto}
h1{font-size:20px;margin:0 0 4px}
h2{font-size:15px;margin:0 0 2px}
p.lede{color:var(--muted);margin:0 0 24px}
.scroll{overflow-x:auto;border:1px solid var(--line);border-radius:6px;margin-bottom:28px}
table{border-collapse:collapse;width:100%;min-width:860px;font-size:13px}
th,td{padding:7px 10px;text-align:left;border-bottom:1px solid var(--line);
white-space:nowrap}
th{background:var(--card);font-weight:600}
tbody tr:last-child td{border-bottom:none}
td.num{text-align:right;font-variant-numeric:tabular-nums}
.pos{color:var(--pos)}
.neg{color:var(--neg)}
.card{border:1px solid var(--line);border-radius:6px;padding:16px;margin-bottom:18px;
background:var(--card)}
.card .route{color:var(--muted);font-size:12px;margin:0 0 12px}
.stats{display:flex;flex-wrap:wrap;gap:10px 26px;margin:0 0 12px;padding:0;
list-style:none}
.stats li{min-width:120px}
.stats .k{display:block;color:var(--muted);font-size:11px;
text-transform:uppercase;letter-spacing:.04em}
.stats .v{font-size:16px;font-variant-numeric:tabular-nums}
.verdict{margin:0 0 12px;padding:8px 10px;border-left:3px solid var(--muted);
background:var(--bg);border-radius:0 4px 4px 0}
.chartbox{overflow-x:auto}
svg.chart{display:block;width:100%;min-width:640px;height:auto}
.empty{color:var(--muted);padding:20px 0}
""".strip()


def _num_class(value: float | None) -> str:
    """CSS classes for a numeric cell, tinted by sign."""
    if value is None or abs(value) < 1e-9:
        return "num"
    return "num pos" if value > 0 else "num neg"


def write_html(results: Sequence[EventRouteResult], path: str) -> None:
    """Write a single self-contained HTML page with a card per result."""
    esc = html.escape
    rows: list[str] = []
    for result in results:
        rows.append(
            "<tr>"
            f"<td>{esc(result.event)}</td>"
            f"<td>{esc(result.route)}</td>"
            f"<td>{esc(_iso(result.start_date))}</td>"
            f"<td>{esc(_iso(result.end_date))}</td>"
            f'<td class="{_num_class(result.mean_delta_pp)}">'
            f"{esc(_fmt_pp(result.mean_delta_pp))}</td>"
            f'<td class="num">&#177;{esc(_fmt_pp(result.normal_swing_pp))}</td>'
            f'<td class="{_num_class(result.mean_pax_pct)}">'
            f"{esc(_fmt_pct(result.mean_pax_pct))}</td>"
            f'<td class="{_num_class(result.mean_cap_pct)}">'
            f"{esc(_fmt_pct(result.mean_cap_pct))}</td>"
            f'<td class="num">{result.n_days}</td>'
            f'<td class="num">{result.n_days_no_baseline}</td>'
            f"<td>{esc(result.verdict)}</td>"
            "</tr>"
        )

    cards: list[str] = []
    for index, result in enumerate(results):
        cards.append(
            f'<section class="card" id="case-{index}">'
            f"<h2>{esc(result.event)}</h2>"
            f'<p class="route">{esc(result.route)} &#183; '
            f"{esc(_iso(result.start_date))} to {esc(_iso(result.end_date))} &#183; "
            f"{result.n_days} days ({result.n_days_no_baseline} without baseline)"
            f"{' &#183; ' + esc(result.note) if result.note else ''}</p>"
            '<ul class="stats">'
            f'<li><span class="k">Mean LF vs baseline</span>'
            f'<span class="v">{esc(_fmt_pp(result.mean_delta_pp))}</span></li>'
            f'<li><span class="k">Normal swing on this route</span>'
            f'<span class="v">&#177;{esc(_fmt_pp(result.normal_swing_pp))}</span></li>'
            f'<li><span class="k">Mean pax vs baseline</span>'
            f'<span class="v">{esc(_fmt_pct(result.mean_pax_pct))}</span></li>'
            f'<li><span class="k">Mean capacity vs baseline</span>'
            f'<span class="v">{esc(_fmt_pct(result.mean_cap_pct))}</span></li>'
            "</ul>"
            f'<p class="verdict">{esc(result.verdict)}</p>'
            f'<div class="chartbox">{_svg_markup(result, standalone=False)}</div>'
            "</section>"
        )

    body = (
        "<main>"
        "<h1>Event case studies: route load factor around the event</h1>"
        '<p class="lede">One row per event and route, in the order supplied. '
        "LF is capacity-weighted; the baseline is the same weekday off-event. "
        "Read the pax and capacity columns beside every LF number.</p>"
        + (
            '<div class="scroll"><table><thead><tr>'
            "<th>Event</th><th>Route</th><th>Start</th><th>End</th>"
            "<th>Mean delta</th><th>Normal swing</th><th>Mean pax</th><th>Mean cap</th>"
            "<th>Days</th><th>No baseline</th><th>Verdict</th>"
            "</tr></thead><tbody>" + "".join(rows) + "</tbody></table></div>"
            if results
            else '<p class="empty">No case studies to report.</p>'
        )
        + "".join(cards)
        + "</main>"
    )

    document = (
        "<!doctype html>\n<html lang=\"en\">\n<head>\n"
        '<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
        "<title>Event case studies</title>\n"
        f"<style>{_PAGE_STYLE}\n{_SVG_STYLE}</style>\n"
        f"</head>\n<body>\n{body}\n</body>\n</html>\n"
    )
    _ensure_parent(path)
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(document)
