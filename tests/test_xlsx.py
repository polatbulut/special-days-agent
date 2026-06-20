import json
import os
import tempfile
import unittest
from datetime import date

from openpyxl import load_workbook

from special_days.models import SpecialDate
from special_days.xlsx_writer import write_xlsx

HEADER = [
    "Event", "Start date", "End date", "City", "Source", "Nearest airport", "Impact",
    "Predicted attendance", "Bridge start", "Bridge end",
    "Impact by day", "Impact by day (bridge)",
]


def make(event, city, start="2026-07-01", end=None, airport=None, impact=None,
         attendance=None, bstart=None, bend=None, by_day=None, by_day_bridge=None):
    s = date.fromisoformat(start)
    e = date.fromisoformat(end) if end else s
    return SpecialDate(
        event, s, e, city, "event", "TR", "ticketmaster",
        nearest_airport=airport, impact_score=impact, predicted_attendance=attendance,
        bridge_start=date.fromisoformat(bstart) if bstart else None,
        bridge_end=date.fromisoformat(bend) if bend else None,
        impact_by_day=by_day, impact_by_day_bridge=by_day_bridge,
    )


class XlsxWriterTest(unittest.TestCase):
    def _write_and_load(self, rows):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "out.xlsx")
            write_xlsx(rows, path)
            workbook = load_workbook(path)
        return workbook.active

    def test_sheet_name_and_bold_header(self):
        sheet = self._write_and_load([make("Tarkan Live", "İstanbul")])
        self.assertEqual(sheet.title, "Special Dates")
        self.assertEqual([c.value for c in sheet[1]], HEADER)
        self.assertTrue(all(c.font.bold for c in sheet[1]))

    def test_rows_have_real_dates_airport_impact_and_unicode(self):
        sheet = self._write_and_load(
            [make("Tarkan Live", "İstanbul", "2026-07-15", "2026-07-16",
                  airport="IST", impact=72, attendance=15000)]
        )
        self.assertEqual(sheet["A2"].value, "Tarkan Live")
        self.assertEqual(sheet["D2"].value, "İstanbul")  # unicode preserved
        self.assertEqual(sheet["E2"].value, "ticketmaster")  # source
        self.assertEqual(sheet["F2"].value, "IST")  # nearest airport
        self.assertEqual(sheet["G2"].value, 72)  # numeric impact cell
        self.assertEqual(sheet["H2"].value, 15000)  # numeric attendance cell
        start = sheet["B2"].value  # real date cell, not text
        self.assertEqual((start.year, start.month, start.day), (2026, 7, 15))
        self.assertEqual(sheet["B2"].number_format, "yyyy-mm-dd")

    def test_freeze_panes_and_autofilter(self):
        sheet = self._write_and_load([make("A", "X"), make("B", "Y")])
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:L3")  # header + 2 rows, 12 cols

    def test_bridge_dates_and_weight_json_cells(self):
        sheet = self._write_and_load([
            make("Ramazan Bayramı", "Nationwide (TR)", "2027-03-08", "2027-03-11",
                 impact=99, bstart="2027-03-06", bend="2027-03-14",
                 by_day=(("2027-03-08", 99), ("2027-03-09", 75)),
                 by_day_bridge=(("2027-03-06", 99),)),
        ])
        # Bridge dates are real date cells with the date number format (cols I, J).
        i2 = sheet["I2"]
        self.assertEqual((i2.value.year, i2.value.month, i2.value.day), (2027, 3, 6))
        self.assertEqual(i2.number_format, "yyyy-mm-dd")
        self.assertEqual((sheet["J2"].value.year, sheet["J2"].value.month), (2027, 3))
        # Weight lists are JSON strings (cols K, L).
        self.assertEqual(json.loads(sheet["K2"].value), {"2027-03-08": 99, "2027-03-09": 75})
        self.assertEqual(json.loads(sheet["L2"].value), {"2027-03-06": 99})

    def test_empty_rows_still_valid(self):
        sheet = self._write_and_load([])
        self.assertEqual([c.value for c in sheet[1]], HEADER)
        self.assertEqual(sheet.max_row, 1)  # header only


if __name__ == "__main__":
    unittest.main()
